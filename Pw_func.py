# In Pw_func.py

import os
from tkinter import filedialog

import openpyxl
import shutil
import csv
import random
import string
import tkinter as tk
from tkinter import messagebox


def show_permission_error_dialog():
    root = tk.Tk()
    root.withdraw()  # Verstecke das Hauptfenster
    messagebox.showerror("Permission Error",
                         "Bitte schließe die Excel-Datei und klicke dann auf 'OK', um es erneut zu versuchen.")
    root.destroy()


def create_backup(file_path):
    backup_path = file_path.replace('.xlsx', '_backup.xlsx')
    shutil.copyfile(file_path, backup_path)
    return backup_path


def generate_random_password(length=8, special_chars="!#?", special_count=2):
    special_indices = random.sample(range(length), special_count)
    characters = string.ascii_letters + string.digits
    password_list = [random.choice(characters) for _ in range(length)]

    for index in special_indices:
        password_list[index] = random.choice(special_chars)

    return ''.join(password_list)


def generate_fixed_password(length=10, special_chars="!#?", special_count=1):
    special_indices = random.sample(range(length), special_count)
    characters = string.ascii_letters + string.digits
    password_list = [random.choice(characters) for _ in range(length - special_count)]

    for index in special_indices:
        password_list.insert(index, random.choice(special_chars))

    return ''.join(password_list)


def generate_passwords_and_save_to_csv(file_path):
    try:
        output_messages = []  # Liste zum Sammeln der Ausgaben

        # Backup-Datei erstellen
        create_backup(file_path)
        output_messages.append("Backup-Datei erstellt.\n")

        # Excel-Datei öffnen
        workbook = openpyxl.load_workbook(file_path)
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        directory = os.path.dirname(file_path)

        # Festes Passwort für alle Mappen generieren
        fixed_password = generate_fixed_password()
        output_messages.append(f"Festes Passwort für Dozent generiert: {fixed_password}\n")

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            nutzer_column = None
            passwort_column = None

            # CSV-Dateiname festlegen
            csv_file_path = os.path.join(directory, f"{base_name}_{sheet_name}_passwords.csv")

            # Setze festes Passwort für den Dozenten in Zeile 2, Spalte A
            dozent_cell = sheet['A2']
            current_value = dozent_cell.value
            dozent_cell.value = f"{current_value} {fixed_password}"

            with (open(csv_file_path, 'w', newline='') as csv_file):
                writer = csv.writer(csv_file)

                # Spalten 'Nutzer' und 'Passwort' finden
                for i, column in enumerate(sheet.iter_cols(min_row=1, max_row=1)):
                    if column[0].value == "Nutzer":
                        nutzer_column = i
                    elif column[0].value == "Kennwort" or column[0].value == "Passwort":
                        passwort_column = i

                if nutzer_column is None or passwort_column is None:
                    output_messages.append(f"Spalte 'Nutzer' oder 'Passwort' in '{sheet_name}' nicht gefunden.\n")
                    continue

                # Basispasswort für die Mappe generieren
                base_password = generate_random_password()
                tn_counter = 1

                # Passwörter für Nutzer generieren und in CSV speichern
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    nutzer_value = row[nutzer_column].value
                    if nutzer_value and str(nutzer_value).startswith("TN") or nutzer_value and str(nutzer_value).startswith(
                            "PU") or nutzer_value and str(nutzer_value).startswith("FA") or nutzer_value and str(nutzer_value).startswith("LB")or nutzer_value and str(nutzer_value).startswith("RG"):
                        nutzer_value_simple = nutzer_value.split('@')[0]
                        password = base_password + str(tn_counter).zfill(2)
                        row[passwort_column].value = password
                        writer.writerow([nutzer_value_simple, password])
                        output_messages.append(
                            f"Passwort für {nutzer_value_simple} in '{sheet_name}' gesetzt: {password}\n")
                        tn_counter += 1

                output_messages.append(f"CSV-Datei für '{sheet_name}' erstellt: {csv_file_path}\n")

        # Excel-Datei speichern
        workbook.save(file_path)
        output_messages.append("Passwörter erfolgreich generiert und Excel-Datei gespeichert.\n")

        return '\n'.join(output_messages)
    except PermissionError:
        show_permission_error_dialog()
        return generate_passwords_and_save_to_csv(file_path)  # Versuche es erneut

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        return generate_passwords_and_save_to_csv(file_path)
    return ""

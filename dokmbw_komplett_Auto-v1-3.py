import os
import re
import ssl
import smtplib
import requests
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import xml.etree.ElementTree as ET
import pandas as pd
from datetime import datetime, time
from email.message import EmailMessage
from requests.adapters import HTTPAdapter
from requests_ntlm import HttpNtlmAuth
from urllib3.poolmanager import PoolManager
from urllib3.exceptions import InsecureRequestWarning
import urllib3
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
import threading
import time as time_module
import schedule
import sqlite3
import json

urllib3.disable_warnings(InsecureRequestWarning)

# ---------- CONFIGURATION FILE ----------
CONFIG_FILE = Path(__file__).parent / "dokmbw_config.json"
DB_FILE = Path(__file__).parent / "dokmbw_bewertungen.db"
LOG_DIR = Path("C:/Log-Bewertungen")

DEFAULT_CONFIG = {
    "username": "itbg\\sp_farm",
    "password": "!DokM1MLcgi",
    "email_password": "quaSeu2i",
    "absender": "ml-projekte-it@mail.de",
    "empfaenger": [
        "a.borowczak@mlgruppe.de",
        "k.vosen@mlgruppe.de",
        "c.mueller@mlgruppe.de",
        "r.panske@mlgruppe.de",
        "j.kujasch@mlgruppe.de"
    ],
    "smtp_host": "smtp.mail.de",
    "mail_betreff": "{} Bewertung DokMBW",
    "mail_text": """Hallo zusammen,

diese Mail wurde Automatisiert versendet. Bitte nicht auf die Absenderadresse antworten.


Gruß
Die Software von Andy
""",
    "abruf_zeit": "16:30",  # Standard-Abrufzeit (Fallback)
    "server_wochentage": {},  # {server_num: [0=Mo, 1=Di, ..., 4=Fr]} oder {} für Mo-Fr
    "server_abruf_zeiten": {},  # {server_num: "HH:MM"} oder {} für Standard-Zeit
    "kuerzel_mapping": {
        "Anwender-Schulung": "AN",
        "Assistent": "AA",
        "Löschberechtigter": "LOEBE",
        "Registrator": "REG",
        "Anwendungsmanager": "AM"
    }
}


# ---------- CONFIGURATION MANAGER ----------
def load_config():
    """Lädt die Konfiguration aus der JSON-Datei"""
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            config = json.load(f)
            # Merge with defaults for any missing keys
            for key, value in DEFAULT_CONFIG.items():
                if key not in config:
                    config[key] = value
            return config
    return DEFAULT_CONFIG.copy()


def save_config(config):
    """Speichert die Konfiguration in die JSON-Datei"""
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=4, ensure_ascii=False)


# Global config
CONFIG = load_config()


# ---------- DATABASE ----------
def init_database():
    """Initialisiert die SQLite-Datenbank für Bewertungen"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS bewertungen (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            datum TEXT,
            trainer TEXT,
            seminar TEXT,
            ort TEXT,
            beginn TEXT,
            ende TEXT,
            zufriedenheit_weiterempfehlen INTEGER,
            zufriedenheit_seminar INTEGER,
            zufriedenheit_trainer INTEGER,
            bewertung_unterlagen INTEGER,
            bewertung_praxis INTEGER,
            bewertung_praesentation INTEGER,
            bewertung_methodisch INTEGER,
            bewertung_fachlich INTEGER,
            bewertung_struktur INTEGER,
            bewertung_atmosphaere INTEGER,
            gesamtnote REAL,
            themen_interessiert TEXT,
            themen_zu_kurz TEXT,
            vorschlaege TEXT,
            erstellt_von TEXT,
            server TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    conn.commit()
    conn.close()


def save_rating_to_db(rating_data):
    """Speichert eine Bewertung in die Datenbank"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    # Berechne Gesamtnote (Durchschnitt aller Bewertungen)
    bewertungen = [
        rating_data.get('zufriedenheit_weiterempfehlen'),
        rating_data.get('zufriedenheit_seminar'),
        rating_data.get('zufriedenheit_trainer'),
        rating_data.get('bewertung_unterlagen'),
        rating_data.get('bewertung_praxis'),
        rating_data.get('bewertung_praesentation'),
        rating_data.get('bewertung_methodisch'),
        rating_data.get('bewertung_fachlich'),
        rating_data.get('bewertung_struktur'),
        rating_data.get('bewertung_atmosphaere')
    ]
    bewertungen = [b for b in bewertungen if b is not None]
    gesamtnote = sum(bewertungen) / len(bewertungen) if bewertungen else None

    cursor.execute('''
        INSERT INTO bewertungen (
            datum, trainer, seminar, ort, beginn, ende,
            zufriedenheit_weiterempfehlen, zufriedenheit_seminar, zufriedenheit_trainer,
            bewertung_unterlagen, bewertung_praxis, bewertung_praesentation,
            bewertung_methodisch, bewertung_fachlich, bewertung_struktur, bewertung_atmosphaere,
            gesamtnote, themen_interessiert, themen_zu_kurz, vorschlaege, erstellt_von, server
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        rating_data.get('datum'),
        rating_data.get('trainer'),
        rating_data.get('seminar'),
        rating_data.get('ort'),
        rating_data.get('beginn'),
        rating_data.get('ende'),
        rating_data.get('zufriedenheit_weiterempfehlen'),
        rating_data.get('zufriedenheit_seminar'),
        rating_data.get('zufriedenheit_trainer'),
        rating_data.get('bewertung_unterlagen'),
        rating_data.get('bewertung_praxis'),
        rating_data.get('bewertung_praesentation'),
        rating_data.get('bewertung_methodisch'),
        rating_data.get('bewertung_fachlich'),
        rating_data.get('bewertung_struktur'),
        rating_data.get('bewertung_atmosphaere'),
        gesamtnote,
        rating_data.get('themen_interessiert'),
        rating_data.get('themen_zu_kurz'),
        rating_data.get('vorschlaege'),
        rating_data.get('erstellt_von'),
        rating_data.get('server')
    ))

    conn.commit()
    conn.close()


def get_trainer_statistics():
    """Holt Statistiken für alle Trainer"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    cursor.execute('''
        SELECT
            trainer,
            COUNT(*) as anzahl_bewertungen,
            AVG(gesamtnote) as durchschnitt,
            MIN(gesamtnote) as schlechteste,
            MAX(gesamtnote) as beste
        FROM bewertungen
        WHERE trainer IS NOT NULL AND trainer != ''
        GROUP BY trainer
        ORDER BY durchschnitt DESC
    ''')

    results = cursor.fetchall()
    conn.close()

    return results


def get_best_worst_ratings(trainer):
    """Holt beste und schlechteste Bewertungen für einen Trainer"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    # Beste Bewertung
    cursor.execute('''
        SELECT datum, seminar, gesamtnote, themen_interessiert, vorschlaege
        FROM bewertungen
        WHERE trainer = ? AND gesamtnote IS NOT NULL
        ORDER BY gesamtnote DESC
        LIMIT 1
    ''', (trainer,))
    beste = cursor.fetchone()

    # Schlechteste Bewertung
    cursor.execute('''
        SELECT datum, seminar, gesamtnote, themen_zu_kurz, vorschlaege
        FROM bewertungen
        WHERE trainer = ? AND gesamtnote IS NOT NULL
        ORDER BY gesamtnote ASC
        LIMIT 1
    ''', (trainer,))
    schlechteste = cursor.fetchone()

    conn.close()

    return beste, schlechteste


# ---------- HELPER ----------
def get_value_from_rating(rating_str, label):
    """Extrahiert numerische Werte aus Bewertungs-Strings"""
    match = re.search(fr"{re.escape(label)};#(\d+)#", rating_str)
    return int(match.group(1)) if match else None


def extract_block(description_raw, label):
    """Extrahiert beschriftete Textblöcke aus HTML-Beschreibungen"""
    match = re.search(fr"<b>{re.escape(label)}:</b>\s*(.*?)</div>", description_raw.replace("\n", ""))
    return re.sub(r"<.*?>", "", match.group(1).strip()) if match else ""


def berechne_speicherpfad(dateiname: str):
    """Berechnet Speicherpfad basierend auf aktueller KW"""
    dokumente = Path.home() / "Documents"
    hauptordner = dokumente / "Bewertungen"
    kw_ordner = hauptordner / f"KW{datetime.now().isocalendar().week}"
    kw_ordner.mkdir(parents=True, exist_ok=True)
    return kw_ordner / dateiname


def get_anwender_kuerzel(seminar_titel):
    """Konvertiert Seminartitel zu Anwender-Kürzeln"""
    for titel, kuerzel in CONFIG["kuerzel_mapping"].items():
        if titel.lower() in seminar_titel.lower():
            return kuerzel
    return "UNBEKANNT"


# ---------- CONVERTER ----------
def convert_xml_to_excel(xml_file_path, output_excel_path, save_to_db=True):
    """Konvertiert XML zu Excel und speichert optional in DB"""

    tree = ET.parse(xml_file_path)
    root = tree.getroot()
    channel = root.find("channel")
    entries = []

    for item in channel.findall("item"):
        author = "***"
        pub_date = item.find("pubDate").text
        pub_date = datetime.strptime(pub_date, "%a, %d %b %Y %H:%M:%S %Z").strftime("%d.%m.%Y")
        description_raw = item.find("description").text

        seminar = extract_block(description_raw, "Titel des Seminars")
        ort = extract_block(description_raw, "Ort der Schulung")
        beginn = extract_block(description_raw, "Schulungszeitraum (Beginn)")
        ende = extract_block(description_raw, "Schulungszeitraum (Ende)")
        trainer = extract_block(description_raw, "Name des Trainers")
        themen = extract_block(description_raw, "Folgende Themen haben mich besonders interessiert:")
        zu_kurz = extract_block(description_raw, "Diese Themen kamen meiner Meinung nach zu kurz/habe ich nicht verstanden:")
        bewertung = extract_block(description_raw, "Lehrgangsbewertung")
        zufriedenheit = extract_block(description_raw, "Zufriedenheit")
        vorschlaege = extract_block(description_raw, "Haben Sie noch Wünsche, Vorschläge, Anregungen?")

        zufriedenheit_weiterempfehlen = get_value_from_rating(zufriedenheit, "Ich würde das Seminar weiterempfehlen")
        zufriedenheit_seminar = get_value_from_rating(zufriedenheit, "Zufriedenheit mit Seminar")
        zufriedenheit_trainer = get_value_from_rating(zufriedenheit, "Zufriedenheit mit Trainer/in")
        bewertung_unterlagen = get_value_from_rating(bewertung, "Teilnehmerunterlagen")
        bewertung_praxis = get_value_from_rating(bewertung, "Praxisanteil des Seminars (erster Eindruck)")
        bewertung_praesentation = get_value_from_rating(bewertung, "Präsentation der Inhalte (Nachvollziehbarkeit)")
        bewertung_methodisch = get_value_from_rating(bewertung, "Durchführung durch Trainer/in (Methodisch)")
        bewertung_fachlich = get_value_from_rating(bewertung, "Durchführung durch Trainer/in (Fachlich)")
        bewertung_struktur = get_value_from_rating(bewertung, "Struktur des Seminars (Roter Faden)")
        bewertung_atmosphaere = get_value_from_rating(bewertung, "Allgemeine Atmosphäre")

        entries.append({
            "Erstellt von": author,
            "Titel des Seminars": seminar,
            "Ort der Schulung": ort,
            "Schulungszeitraum (Beginn)": beginn,
            "Schulungszeitraum (Ende)": ende,
            "Name des Trainers": trainer,
            "Folgende Themen haben mich besonders interessiert:": themen,
            "Diese Themen kamen meiner Meinung nach zu kurz/habe ich nicht verstanden:": zu_kurz,
            "Haben Sie noch Wünsche, Vorschläge, Anregungen?": vorschlaege,
            "Zufriedenheit_Ich würde das Seminar weiterempfehlen": zufriedenheit_weiterempfehlen,
            "Zufriedenheit_Zufriedenheit mit Seminar": zufriedenheit_seminar,
            "Zufriedenheit_Zufriedenheit mit Trainer/in": zufriedenheit_trainer,
            "Lehrgangsbewertung_Teilnehmerunterlagen": bewertung_unterlagen,
            "Lehrgangsbewertung_Praxisanteil des Seminars (erster Eindruck)": bewertung_praxis,
            "Lehrgangsbewertung_Präsentation der Inhalte (Nachvollziehbarkeit)": bewertung_praesentation,
            "Lehrgangsbewertung_Durchführung durch Trainer/in (Methodisch)": bewertung_methodisch,
            "Lehrgangsbewertung_Durchführung durch Trainer/in (Fachlich)": bewertung_fachlich,
            "Lehrgangsbewertung_Struktur des Seminars (Roter Faden)": bewertung_struktur,
            "Lehrgangsbewertung_Allgemeine Atmosphäre": bewertung_atmosphaere,
            "Elementtyp": "Element",
            "Pfad": ""
        })

        # In DB speichern
        if save_to_db:
            server_num = re.search(r'(\d+)_feed\.xml', xml_file_path)
            server_prefix = server_num.group(1) if server_num else "unknown"

            rating_data = {
                'datum': pub_date,
                'trainer': trainer,
                'seminar': seminar,
                'ort': ort,
                'beginn': beginn,
                'ende': ende,
                'zufriedenheit_weiterempfehlen': zufriedenheit_weiterempfehlen,
                'zufriedenheit_seminar': zufriedenheit_seminar,
                'zufriedenheit_trainer': zufriedenheit_trainer,
                'bewertung_unterlagen': bewertung_unterlagen,
                'bewertung_praxis': bewertung_praxis,
                'bewertung_praesentation': bewertung_praesentation,
                'bewertung_methodisch': bewertung_methodisch,
                'bewertung_fachlich': bewertung_fachlich,
                'bewertung_struktur': bewertung_struktur,
                'bewertung_atmosphaere': bewertung_atmosphaere,
                'themen_interessiert': themen,
                'themen_zu_kurz': zu_kurz,
                'vorschlaege': vorschlaege,
                'erstellt_von': author,
                'server': server_prefix
            }
            save_rating_to_db(rating_data)

    df = pd.DataFrame(entries)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bewertungen"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    stripe_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        for cell in ws[r_idx]:
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif r_idx % 2 == 0:
                cell.fill = stripe_fill

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 15)

    ws.auto_filter.ref = ws.dimensions
    wb.save(output_excel_path)


# ---------- RSS FEED ----------
class LegacySSLAdapter(HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
        ctx.options |= 0x00000004
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        kwargs["ssl_context"] = ctx
        self.poolmanager = PoolManager(*args, **kwargs)


def download_rss_and_save_xml(server_num):
    # Erstelle Log-Verzeichnis falls nicht vorhanden
    LOG_DIR.mkdir(parents=True, exist_ok=True)

    server_prefix = f"{int(server_num):02}"
    base_url = f"https://{server_prefix}.ml-schulung.de"
    overview_url = f"{base_url}/trainerseite/Lists/Lehrgangsbewertung/overview.aspx"

    session = requests.Session()
    session.mount("https://", LegacySSLAdapter())
    session.auth = HttpNtlmAuth(CONFIG["username"], CONFIG["password"])
    session.verify = False
    session.headers.update({"User-Agent": "DokMBW/NTLM"})

    try:
        res = session.get(overview_url, timeout=10)
        res.raise_for_status()
        with open(LOG_DIR / f"{server_prefix}_overview_debug.html", "w", encoding="utf-8") as f:
            f.write(res.text)

        guids = re.findall(r'\{[0-9a-fA-F\-]{36}\}', res.text)
        guids = list(set(guids))
        if not guids:
            raise Exception("Keine GUIDs im HTML gefunden.")

        for guid in guids:
            feed_url = f"{base_url}/trainerseite/_layouts/15/listfeed.aspx?List={guid}"
            try:
                feed_res = session.get(feed_url, timeout=10)
                feed_res.raise_for_status()

                xml_filename = LOG_DIR / f"{server_prefix}_feed.xml"
                with open(xml_filename, "w", encoding="utf-8") as f:
                    f.write(feed_res.text)

                root = ET.fromstring(feed_res.text)
                trainer = extract_block(root.find("channel").find("item").find("description").text, "Name des Trainers")

                return {
                    "server": server_prefix,
                    "trainer": trainer,
                    "xml_file": str(xml_filename)
                }

            except Exception:
                continue

        raise Exception("Keine funktionierende Feed-URL mit gültiger GUID gefunden.")
    except Exception as e:
        raise Exception(f"[{server_prefix}] Fehler: {e}")


# ---------- MAIL ----------
def sende_auswertung_per_mail(excel_dateien, empfaenger):
    datum = datetime.now().strftime("%d.%m.%Y")

    msg = EmailMessage()
    msg["Subject"] = CONFIG["mail_betreff"].format(datum)
    msg["From"] = CONFIG["absender"]
    msg["To"] = empfaenger
    msg.set_content(CONFIG["mail_text"])

    for datei in excel_dateien:
        with open(datei, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="octet-stream",
                filename=os.path.basename(datei)
            )

    try:
        with smtplib.SMTP(CONFIG["smtp_host"], 587, timeout=20) as smtp:
            smtp.starttls()
            smtp.login(CONFIG["absender"], CONFIG["email_password"])
            smtp.send_message(msg)
            print(f"✅ Mail erfolgreich gesendet an {empfaenger}")
    except Exception as e:
        print(f"❌ Fehler beim Senden an {empfaenger}: {e}")


# ---------- AUTOMATISCHE SERVER-ABFRAGE ----------
def should_process_server(server_num, today_weekday):
    """
    Prüft ob Server heute verarbeitet werden soll
    today_weekday: 0=Montag, 1=Dienstag, ..., 6=Sonntag
    """
    server_config = CONFIG.get("server_wochentage", {})
    server_key = str(server_num)

    if server_key in server_config and server_config[server_key]:
        # Spezifische Tage definiert
        return today_weekday in server_config[server_key]
    else:
        # Standard: Mo-Fr (0-4)
        return 0 <= today_weekday <= 4


def run_automatic_process_for_server(server_num):
    """Verarbeitet einen einzelnen Server automatisch"""
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Starte automatische Verarbeitung für Server {server_num:02d}...")

    today_weekday = datetime.now().weekday()

    if not should_process_server(server_num, today_weekday):
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Server {server_num:02d}: Heute nicht geplant (Wochentag-Filter)")
        return

    try:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Prüfe Server {server_num:02d}...")
        result = download_rss_and_save_xml(server_num)

        root = ET.parse(result["xml_file"]).getroot()
        channel = root.find("channel")
        first_item = channel.find("item")

        if first_item is None:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Server {server_num:02d}: Keine Daten gefunden")
            return

        description_raw = first_item.find("description").text
        seminar_titel = extract_block(description_raw, "Titel des Seminars")

        datum = datetime.now().strftime("%Y%m%d")
        anwender_kuerzel = get_anwender_kuerzel(seminar_titel)
        dateiname = f"{datum}-{anwender_kuerzel}-{result['trainer']}.xlsx"
        zielpfad = berechne_speicherpfad(dateiname)

        convert_xml_to_excel(result["xml_file"], zielpfad)
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Server {server_num:02d}: Erfolgreich verarbeitet - {dateiname}")

        # Mail einzeln versenden
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Versende Mail für Server {server_num:02d}...")
        sende_auswertung_per_mail([str(zielpfad)], CONFIG["empfaenger"])
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Server {server_num:02d}: Verarbeitung abgeschlossen!")

    except Exception as e:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Server {server_num:02d}: {str(e)}")


def run_automatic_process():
    """Läuft automatisch alle Server durch und verarbeitet gefundene Daten (für manuellen Test)"""
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Starte automatische Server-Abfrage...")

    erfolge = []
    alle_server = list(range(1, 11))
    today_weekday = datetime.now().weekday()

    for server in alle_server:
        if not should_process_server(server, today_weekday):
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Server {server:02d}: Heute nicht geplant (Wochentag-Filter)")
            continue

        try:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Prüfe Server {server:02d}...")
            result = download_rss_and_save_xml(server)

            root = ET.parse(result["xml_file"]).getroot()
            channel = root.find("channel")
            first_item = channel.find("item")

            if first_item is None:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] Server {server:02d}: Keine Daten gefunden")
                continue

            description_raw = first_item.find("description").text
            seminar_titel = extract_block(description_raw, "Titel des Seminars")

            datum = datetime.now().strftime("%Y%m%d")
            anwender_kuerzel = get_anwender_kuerzel(seminar_titel)
            dateiname = f"{datum}-{anwender_kuerzel}-{result['trainer']}.xlsx"
            zielpfad = berechne_speicherpfad(dateiname)

            convert_xml_to_excel(result["xml_file"], zielpfad)
            erfolge.append(str(zielpfad))
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Server {server:02d}: Erfolgreich verarbeitet - {dateiname}")

        except Exception as e:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Server {server:02d}: {str(e)}")
            continue

    if erfolge:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Versende {len(erfolge)} Datei(en) per Mail...")
        sende_auswertung_per_mail(erfolge, CONFIG["empfaenger"])
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Automatische Verarbeitung abgeschlossen!")
    else:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Keine Daten zur Verarbeitung gefunden.")


def run_manual_test():
    """Manuelle Test-Ausführung"""
    print("=== MANUELLER TEST GESTARTET ===")
    run_automatic_process()
    print("=== MANUELLER TEST BEENDET ===")


def setup_scheduler():
    """Richtet den Zeitplaner ein - für jeden Server individuell"""
    schedule.clear()

    default_zeit = CONFIG.get("abruf_zeit", "16:30")
    server_zeiten = CONFIG.get("server_abruf_zeiten", {})

    # Für jeden Server 1-10 einen eigenen Job erstellen
    for server in range(1, 11):
        server_key = str(server)
        # Individuelle Zeit nutzen, falls vorhanden, sonst Standard
        abruf_zeit = server_zeiten.get(server_key, default_zeit)

        # Job für diesen Server erstellen
        schedule.every().day.at(abruf_zeit).do(run_automatic_process_for_server, server)
        print(f"Zeitplaner: Server {server:02d} um {abruf_zeit} Uhr")

    print(f"\n✅ Zeitplaner für alle Server eingerichtet!")
    print(f"Standard-Zeit: {default_zeit} Uhr")
    print(f"Individuelle Zeiten: {len(server_zeiten)} Server mit eigener Zeit\n")

    while True:
        schedule.run_pending()
        time_module.sleep(60)


def start_scheduler():
    """Startet den Zeitplaner in einem separaten Thread"""
    scheduler_thread = threading.Thread(target=setup_scheduler, daemon=True)
    scheduler_thread.start()


# ---------- GUI: KONFIGURATION ----------
def open_config_window():
    """Öffnet Konfigurationsfenster"""
    config_window = tk.Toplevel()
    config_window.title("Konfiguration")
    config_window.geometry("700x800")

    notebook = ttk.Notebook(config_window)
    notebook.pack(fill="both", expand=True, padx=10, pady=10)

    # Tab 1: E-Mail Konfiguration
    email_frame = ttk.Frame(notebook)
    notebook.add(email_frame, text="E-Mail")

    tk.Label(email_frame, text="Absender E-Mail:", font=("Arial", 10, "bold")).grid(row=0, column=0, sticky="w", padx=5, pady=5)
    absender_entry = tk.Entry(email_frame, width=50)
    absender_entry.insert(0, CONFIG["absender"])
    absender_entry.grid(row=0, column=1, padx=5, pady=5)

    tk.Label(email_frame, text="E-Mail Passwort:", font=("Arial", 10, "bold")).grid(row=1, column=0, sticky="w", padx=5, pady=5)
    email_pw_entry = tk.Entry(email_frame, width=50, show="*")
    email_pw_entry.insert(0, CONFIG["email_password"])
    email_pw_entry.grid(row=1, column=1, padx=5, pady=5)

    tk.Label(email_frame, text="SMTP Host:", font=("Arial", 10, "bold")).grid(row=2, column=0, sticky="w", padx=5, pady=5)
    smtp_entry = tk.Entry(email_frame, width=50)
    smtp_entry.insert(0, CONFIG["smtp_host"])
    smtp_entry.grid(row=2, column=1, padx=5, pady=5)

    tk.Label(email_frame, text="Empfänger (einer pro Zeile):", font=("Arial", 10, "bold")).grid(row=3, column=0, sticky="nw", padx=5, pady=5)
    empfaenger_text = scrolledtext.ScrolledText(email_frame, width=50, height=8)
    empfaenger_text.insert("1.0", "\n".join(CONFIG["empfaenger"]))
    empfaenger_text.grid(row=3, column=1, padx=5, pady=5)

    tk.Label(email_frame, text="E-Mail Betreff:", font=("Arial", 10, "bold")).grid(row=4, column=0, sticky="w", padx=5, pady=5)
    betreff_entry = tk.Entry(email_frame, width=50)
    betreff_entry.insert(0, CONFIG["mail_betreff"])
    betreff_entry.grid(row=4, column=1, padx=5, pady=5)

    tk.Label(email_frame, text="E-Mail Text:", font=("Arial", 10, "bold")).grid(row=5, column=0, sticky="nw", padx=5, pady=5)
    mail_text_text = scrolledtext.ScrolledText(email_frame, width=50, height=10)
    mail_text_text.insert("1.0", CONFIG["mail_text"])
    mail_text_text.grid(row=5, column=1, padx=5, pady=5)

    # Tab 2: Zeitplanung
    time_frame = ttk.Frame(notebook)
    notebook.add(time_frame, text="Zeitplanung")

    tk.Label(time_frame, text="Standard-Abrufzeit (HH:MM):", font=("Arial", 10, "bold")).grid(row=0, column=0, sticky="w", padx=5, pady=5)
    abruf_entry = tk.Entry(time_frame, width=20)
    abruf_entry.insert(0, CONFIG["abruf_zeit"])
    abruf_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")
    tk.Label(time_frame, text="(Wird verwendet wenn keine Server-Zeit definiert ist)", font=("Arial", 8, "italic")).grid(row=0, column=2, sticky="w", padx=5)

    # Scrollbarer Frame für Server-Konfiguration
    canvas = tk.Canvas(time_frame)
    scrollbar = tk.Scrollbar(time_frame, orient="vertical", command=canvas.yview)
    scrollable_frame = tk.Frame(canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.grid(row=1, column=0, columnspan=8, sticky="nsew", padx=5, pady=5)
    scrollbar.grid(row=1, column=8, sticky="ns", pady=5)

    time_frame.grid_rowconfigure(1, weight=1)
    time_frame.grid_columnconfigure(0, weight=1)

    tk.Label(scrollable_frame, text="Server-Konfiguration", font=("Arial", 12, "bold")).grid(row=0, column=0, columnspan=9, pady=10)
    tk.Label(scrollable_frame, text="(Wochentage: Leer = Mo-Fr)", font=("Arial", 9, "italic")).grid(row=1, column=0, columnspan=9)

    # Header
    tk.Label(scrollable_frame, text="Server", font=("Arial", 10, "bold")).grid(row=2, column=0, padx=5)
    tk.Label(scrollable_frame, text="Abrufzeit", font=("Arial", 10, "bold")).grid(row=2, column=1, padx=5)

    weekday_names = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]
    for i, day in enumerate(weekday_names):
        tk.Label(scrollable_frame, text=day, font=("Arial", 10, "bold")).grid(row=2, column=i+2)

    # Server-Konfiguration (Zeit + Wochentage)
    server_weekday_vars = {}
    server_time_entries = {}

    for server in range(1, 11):
        # Server-Nummer
        tk.Label(scrollable_frame, text=f"{server:02}").grid(row=server+2, column=0, padx=5, pady=2)

        # Zeit-Eingabefeld
        current_time = CONFIG.get("server_abruf_zeiten", {}).get(str(server), "")
        time_entry = tk.Entry(scrollable_frame, width=10)
        time_entry.insert(0, current_time)
        time_entry.grid(row=server+2, column=1, padx=5, pady=2)
        server_time_entries[server] = time_entry

        # Wochentage
        server_weekday_vars[server] = []
        current_days = CONFIG["server_wochentage"].get(str(server), [])

        for day_idx in range(7):
            var = tk.IntVar(value=1 if day_idx in current_days else 0)
            cb = tk.Checkbutton(scrollable_frame, variable=var)
            cb.grid(row=server+2, column=day_idx+2)
            server_weekday_vars[server].append(var)

    # Tab 3: Kürzel-Mapping
    kuerzel_frame = ttk.Frame(notebook)
    notebook.add(kuerzel_frame, text="Kürzel-Mapping")

    tk.Label(kuerzel_frame, text="Seminartitel → Kürzel Zuordnung:", font=("Arial", 12, "bold")).pack(pady=10)

    kuerzel_entries = []
    for titel, kuerzel in CONFIG["kuerzel_mapping"].items():
        frame = tk.Frame(kuerzel_frame)
        frame.pack(fill="x", padx=10, pady=5)

        tk.Label(frame, text="Titel:", width=15, anchor="w").pack(side="left")
        titel_entry = tk.Entry(frame, width=30)
        titel_entry.insert(0, titel)
        titel_entry.pack(side="left", padx=5)

        tk.Label(frame, text="→", width=3).pack(side="left")

        tk.Label(frame, text="Kürzel:", width=8, anchor="w").pack(side="left")
        kuerzel_entry = tk.Entry(frame, width=15)
        kuerzel_entry.insert(0, kuerzel)
        kuerzel_entry.pack(side="left", padx=5)

        kuerzel_entries.append((titel_entry, kuerzel_entry))

    # Neue Zeile hinzufügen
    def add_kuerzel_row():
        frame = tk.Frame(kuerzel_frame)
        frame.pack(fill="x", padx=10, pady=5)

        tk.Label(frame, text="Titel:", width=15, anchor="w").pack(side="left")
        titel_entry = tk.Entry(frame, width=30)
        titel_entry.pack(side="left", padx=5)

        tk.Label(frame, text="→", width=3).pack(side="left")

        tk.Label(frame, text="Kürzel:", width=8, anchor="w").pack(side="left")
        kuerzel_entry = tk.Entry(frame, width=15)
        kuerzel_entry.pack(side="left", padx=5)

        kuerzel_entries.append((titel_entry, kuerzel_entry))

    tk.Button(kuerzel_frame, text="+ Neue Zuordnung", command=add_kuerzel_row).pack(pady=10)

    # Speichern Button
    def save_all_config():
        # E-Mail Konfiguration
        CONFIG["absender"] = absender_entry.get()
        CONFIG["email_password"] = email_pw_entry.get()
        CONFIG["smtp_host"] = smtp_entry.get()
        CONFIG["empfaenger"] = [line.strip() for line in empfaenger_text.get("1.0", "end").split("\n") if line.strip()]
        CONFIG["mail_betreff"] = betreff_entry.get()
        CONFIG["mail_text"] = mail_text_text.get("1.0", "end").rstrip()

        # Zeitplanung
        CONFIG["abruf_zeit"] = abruf_entry.get()

        # Server-Abrufzeiten
        new_server_zeiten = {}
        for server, time_entry in server_time_entries.items():
            zeit = time_entry.get().strip()
            if zeit:  # Nur speichern wenn nicht leer
                # Validierung: Zeit im Format HH:MM
                if re.match(r"^\d{1,2}:\d{2}$", zeit):
                    new_server_zeiten[str(server)] = zeit
                else:
                    messagebox.showwarning("Ungültige Zeit", f"Server {server:02d}: Bitte Zeit im Format HH:MM eingeben (z.B. 11:30)")
                    return
        CONFIG["server_abruf_zeiten"] = new_server_zeiten

        # Server-Wochentage
        new_server_wochentage = {}
        for server, day_vars in server_weekday_vars.items():
            selected_days = [i for i, var in enumerate(day_vars) if var.get() == 1]
            if selected_days:  # Nur speichern wenn nicht leer
                new_server_wochentage[str(server)] = selected_days
        CONFIG["server_wochentage"] = new_server_wochentage

        # Kürzel-Mapping
        new_mapping = {}
        for titel_entry, kuerzel_entry in kuerzel_entries:
            titel = titel_entry.get().strip()
            kuerzel = kuerzel_entry.get().strip()
            if titel and kuerzel:
                new_mapping[titel] = kuerzel
        CONFIG["kuerzel_mapping"] = new_mapping

        save_config(CONFIG)
        messagebox.showinfo("Gespeichert", "Konfiguration erfolgreich gespeichert!")
        config_window.destroy()

    tk.Button(config_window, text="💾 Speichern", command=save_all_config, bg="green", fg="white", font=("Arial", 12, "bold")).pack(pady=10)


# ---------- GUI: STATISTIK ----------
def open_statistics_window():
    """Öffnet Statistik-Fenster"""
    stats_window = tk.Toplevel()
    stats_window.title("Trainer-Statistiken")
    stats_window.geometry("900x600")

    # Treeview für Statistiken
    tree_frame = tk.Frame(stats_window)
    tree_frame.pack(fill="both", expand=True, padx=10, pady=10)

    scrollbar = tk.Scrollbar(tree_frame)
    scrollbar.pack(side="right", fill="y")

    tree = ttk.Treeview(tree_frame, columns=("Trainer", "Anzahl", "Durchschnitt", "Beste", "Schlechteste"),
                        show="headings", yscrollcommand=scrollbar.set)
    scrollbar.config(command=tree.yview)

    tree.heading("Trainer", text="Trainer")
    tree.heading("Anzahl", text="Anzahl Bewertungen")
    tree.heading("Durchschnitt", text="Durchschnittsnote")
    tree.heading("Beste", text="Beste Note")
    tree.heading("Schlechteste", text="Schlechteste Note")

    tree.column("Trainer", width=200)
    tree.column("Anzahl", width=150, anchor="center")
    tree.column("Durchschnitt", width=150, anchor="center")
    tree.column("Beste", width=150, anchor="center")
    tree.column("Schlechteste", width=150, anchor="center")

    tree.pack(fill="both", expand=True)

    # Daten laden
    stats = get_trainer_statistics()
    for stat in stats:
        trainer, anzahl, durchschnitt, schlechteste, beste = stat
        tree.insert("", "end", values=(
            trainer,
            anzahl,
            f"{durchschnitt:.2f}" if durchschnitt else "N/A",
            f"{beste:.2f}" if beste else "N/A",
            f"{schlechteste:.2f}" if schlechteste else "N/A"
        ))

    # Details anzeigen
    def show_details():
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Auswahl erforderlich", "Bitte einen Trainer auswählen.")
            return

        item = tree.item(selected[0])
        trainer = item["values"][0]

        beste, schlechteste = get_best_worst_ratings(trainer)

        details_window = tk.Toplevel(stats_window)
        details_window.title(f"Details: {trainer}")
        details_window.geometry("600x400")

        text = scrolledtext.ScrolledText(details_window, width=70, height=20, wrap=tk.WORD)
        text.pack(padx=10, pady=10, fill="both", expand=True)

        if beste:
            text.insert("end", "🏆 BESTE BEWERTUNG\n", "header")
            text.insert("end", "=" * 60 + "\n")
            text.insert("end", f"Datum: {beste[0]}\n")
            text.insert("end", f"Seminar: {beste[1]}\n")
            text.insert("end", f"Gesamtnote: {beste[2]:.2f}\n")
            text.insert("end", f"\nInteressante Themen:\n{beste[3]}\n")
            text.insert("end", f"\nVorschläge:\n{beste[4]}\n\n")

        if schlechteste:
            text.insert("end", "\n\n❌ SCHLECHTESTE BEWERTUNG\n", "header")
            text.insert("end", "=" * 60 + "\n")
            text.insert("end", f"Datum: {schlechteste[0]}\n")
            text.insert("end", f"Seminar: {schlechteste[1]}\n")
            text.insert("end", f"Gesamtnote: {schlechteste[2]:.2f}\n")
            text.insert("end", f"\nThemen zu kurz gekommen:\n{schlechteste[3]}\n")
            text.insert("end", f"\nVorschläge:\n{schlechteste[4]}\n")

        text.tag_config("header", font=("Arial", 12, "bold"))
        text.config(state="disabled")

    tk.Button(stats_window, text="📊 Details anzeigen", command=show_details,
              bg="blue", fg="white", font=("Arial", 10, "bold")).pack(pady=10)

    tk.Button(stats_window, text="🔄 Aktualisieren", command=lambda: [tree.delete(*tree.get_children()),
              [tree.insert("", "end", values=(s[0], s[1], f"{s[2]:.2f}" if s[2] else "N/A",
              f"{s[4]:.2f}" if s[4] else "N/A", f"{s[3]:.2f}" if s[3] else "N/A")) for s in get_trainer_statistics()]],
              bg="orange", fg="white", font=("Arial", 10, "bold")).pack(pady=5)


# ---------- GUI: MANUAL PROCESSING ----------
def run_gui_process(server_list, versand_aktiv):
    erfolge = []
    fehler = []

    if not versand_aktiv:
        zielordner = filedialog.askdirectory(title="Speicherort für Präsenzbewertung auswählen")
        if not zielordner:
            messagebox.showinfo("Abgebrochen", "Kein Speicherort ausgewählt. Vorgang abgebrochen.")
            return
        zielordner = Path(zielordner)
    else:
        zielordner = None

    for server in server_list:
        try:
            result = download_rss_and_save_xml(server)

            root = ET.parse(result["xml_file"]).getroot()
            channel = root.find("channel")
            first_item = channel.find("item")
            description_raw = first_item.find("description").text
            seminar_titel = extract_block(description_raw, "Titel des Seminars")

            datum = datetime.now().strftime("%Y%m%d")
            anwender_kuerzel = get_anwender_kuerzel(seminar_titel)
            dateiname = f"{datum}-{anwender_kuerzel}-{result['trainer']}.xlsx"
            if versand_aktiv:
                zielpfad = berechne_speicherpfad(dateiname)
            else:
                zielpfad = zielordner / dateiname

            convert_xml_to_excel(result["xml_file"], zielpfad)
            erfolge.append(str(zielpfad))
        except Exception as e:
            fehler.append(str(e))

    if erfolge:
        if versand_aktiv:
            sende_auswertung_per_mail(erfolge, CONFIG["empfaenger"])
        else:
            messagebox.showinfo("Hinweis", f"{len(erfolge)} Datei(en) gespeichert. Kein Versand (Präsenzschulung).")
    else:
        messagebox.showerror("Fehler", "\n".join(fehler))


# ---------- MAIN GUI ----------
def start_gui():
    root = tk.Tk()
    root.title("DokMBW v1.3 - Serverauswahl & Konfiguration")
    root.geometry("500x700")

    # Menü-Buttons
    menu_frame = tk.Frame(root, bg="lightgray")
    menu_frame.pack(fill="x", padx=10, pady=10)

    tk.Button(menu_frame, text="⚙️ Konfiguration", command=open_config_window,
              bg="blue", fg="white", font=("Arial", 11, "bold"), width=20).pack(side="left", padx=5)

    tk.Button(menu_frame, text="📊 Statistiken", command=open_statistics_window,
              bg="purple", fg="white", font=("Arial", 11, "bold"), width=20).pack(side="left", padx=5)

    # Automatische Funktionen
    tk.Label(root, text="=== AUTOMATISCHE FUNKTIONEN ===", font=("Arial", 12, "bold")).pack(pady=(10, 5))

    def manual_test():
        import subprocess
        import sys
        subprocess.Popen([sys.executable, "-c", f"""
import sys
sys.path.insert(0, r'{os.path.dirname(__file__)}')
from pathlib import Path
import importlib.util
spec = importlib.util.spec_from_file_location("dokmbw", r'{__file__}')
module = importlib.util.module_from_spec(spec)
spec.loader.exec_module(module)
module.run_manual_test()
input('Drücke Enter zum Beenden...')
"""], creationflags=subprocess.CREATE_NEW_CONSOLE if os.name == 'nt' else 0)
        messagebox.showinfo("Test gestartet", "Manueller Test läuft in separatem Fenster!")

    def start_auto_mode():
        start_scheduler()
        server_zeiten = CONFIG.get("server_abruf_zeiten", {})
        if server_zeiten:
            info_text = f"Automatische Verarbeitung ist aktiv!\n\nStandard-Zeit: {CONFIG['abruf_zeit']} Uhr\nIndividuelle Zeiten für {len(server_zeiten)} Server konfiguriert"
        else:
            info_text = f"Automatische Verarbeitung ist aktiv!\nAlle Server: {CONFIG['abruf_zeit']} Uhr"
        messagebox.showinfo("Automatik gestartet", info_text)

    tk.Button(root, text="🧪 Manueller Test (alle Server)", command=manual_test,
              bg="orange", fg="white", font=("Arial", 10, "bold")).pack(pady=5, fill="x", padx=20)

    server_zeiten_count = len(CONFIG.get("server_abruf_zeiten", {}))
    auto_text = f"⏰ Automatik starten (Standard: {CONFIG['abruf_zeit']} Uhr"
    if server_zeiten_count > 0:
        auto_text += f", {server_zeiten_count} Server individuell"
    auto_text += ")"

    tk.Button(root, text=auto_text, command=start_auto_mode,
              bg="green", fg="white", font=("Arial", 10, "bold")).pack(pady=5, fill="x", padx=20)

    # Trennlinie
    tk.Label(root, text="=== MANUELLE SERVERAUSWAHL ===", font=("Arial", 12, "bold")).pack(pady=(20, 5))

    tk.Label(root, text="Wähle die Server aus:").pack(pady=(10, 0))

    vars = []
    for i in range(1, 11):
        var = tk.IntVar()
        tk.Checkbutton(root, text=f"{i:02}", variable=var).pack(anchor="w", padx=50)
        vars.append((i, var))

    versand_var = tk.IntVar(value=1)
    tk.Radiobutton(root, text="Versand aktiv (Online-Schulung)", variable=versand_var, value=1).pack(anchor="w", pady=(10, 0), padx=50)
    tk.Radiobutton(root, text="Nur speichern (Präsenzschulung)", variable=versand_var, value=0).pack(anchor="w", padx=50)

    def starten():
        selected = [int(i) for i, var in vars if var.get()]
        if not selected:
            messagebox.showwarning("Hinweis", "Bitte mindestens einen Server auswählen.")
            return
        run_gui_process(selected, versand_var.get() == 1)

    tk.Button(root, text="Ausgewählte Server verarbeiten", command=starten, font=("Arial", 11, "bold")).pack(pady=10)
    root.mainloop()


if __name__ == "__main__":
    # Datenbank initialisieren
    init_database()

    # GUI starten
    start_gui()

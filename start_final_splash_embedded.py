__version__ = "v0.1"

# In Pw_func.py

import os
from tkinter import filedialog

import openpyxl
import shutil
import csv
import random
import string
import tkinter as tk
from tkinter import messagebox


def show_permission_error_dialog():
    root = tk.Tk()
    root.withdraw()  # Verstecke das Hauptfenster
    messagebox.showerror("Permission Error",
                         "Bitte schließe die Excel-Datei und klicke dann auf 'OK', um es erneut zu versuchen.")
    root.destroy()


def create_backup(file_path):
    backup_path = file_path.replace('.xlsx', '_backup.xlsx')
    shutil.copyfile(file_path, backup_path)
    return backup_path


def generate_random_password(length=8, special_chars="!#?", special_count=2):
    special_indices = random.sample(range(length), special_count)
    characters = string.ascii_letters + string.digits
    password_list = [random.choice(characters) for _ in range(length)]

    for index in special_indices:
        password_list[index] = random.choice(special_chars)

    return ''.join(password_list)


def generate_fixed_password(length=10, special_chars="!#?", special_count=1):
    special_indices = random.sample(range(length), special_count)
    characters = string.ascii_letters + string.digits
    password_list = [random.choice(characters) for _ in range(length - special_count)]

    for index in special_indices:
        password_list.insert(index, random.choice(special_chars))

    return ''.join(password_list)


def generate_passwords_and_save_to_csv(file_path):
    try:
        output_messages = []  # Liste zum Sammeln der Ausgaben

        # Backup-Datei erstellen
        create_backup(file_path)
        output_messages.append("Backup-Datei erstellt.\n")

        # Excel-Datei öffnen
        workbook = openpyxl.load_workbook(file_path)
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        directory = os.path.dirname(file_path)

        # Festes Passwort für alle Mappen generieren
        fixed_password = generate_fixed_password()
        output_messages.append(f"Festes Passwort für Dozent generiert: {fixed_password}\n")

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            nutzer_column = None
            passwort_column = None

            # CSV-Dateiname festlegen
            csv_file_path = os.path.join(directory, f"{base_name}_{sheet_name}_passwords.csv")

            # Setze festes Passwort für den Dozenten in Zeile 2, Spalte A
            dozent_cell = sheet['A2']
            current_value = dozent_cell.value
            dozent_cell.value = f"{current_value} {fixed_password}"

            with (open(csv_file_path, 'w', newline='') as csv_file):
                writer = csv.writer(csv_file)

                # Spalten 'Nutzer' und 'Passwort' finden
                for i, column in enumerate(sheet.iter_cols(min_row=1, max_row=1)):
                    if column[0].value == "Nutzer":
                        nutzer_column = i
                    elif column[0].value == "Kennwort" or column[0].value == "Passwort":
                        passwort_column = i

                if nutzer_column is None or passwort_column is None:
                    output_messages.append(f"Spalte 'Nutzer' oder 'Passwort' in '{sheet_name}' nicht gefunden.\n")
                    continue

                # Basispasswort für die Mappe generieren
                base_password = generate_random_password()
                tn_counter = 1

                # Passwörter für Nutzer generieren und in CSV speichern
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    nutzer_value = row[nutzer_column].value
                    if nutzer_value and str(nutzer_value).startswith("TN") or nutzer_value and str(nutzer_value).startswith(
                            "PU") or nutzer_value and str(nutzer_value).startswith("FA") or nutzer_value and str(nutzer_value).startswith("LB")or nutzer_value and str(nutzer_value).startswith("RG"):
                        nutzer_value_simple = nutzer_value.split('@')[0]
                        password = base_password + str(tn_counter).zfill(2)
                        row[passwort_column].value = password
                        writer.writerow([nutzer_value_simple, password])
                        output_messages.append(
                            f"Passwort für {nutzer_value_simple} in '{sheet_name}' gesetzt: {password}\n")
                        tn_counter += 1

                output_messages.append(f"CSV-Datei für '{sheet_name}' erstellt: {csv_file_path}\n")

        # Excel-Datei speichern
        workbook.save(file_path)
        output_messages.append("Passwörter erfolgreich generiert und Excel-Datei gespeichert.\n")

        return '\n'.join(output_messages)
    except PermissionError:
        show_permission_error_dialog()
        return generate_passwords_and_save_to_csv(file_path)  # Versuche es erneut

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        return generate_passwords_and_save_to_csv(file_path)
    return ""


from tkinter import (Tk, Button, Text, Menu, Label, Entry, StringVar,
                     Toplevel, END, OptionMenu, Radiobutton, filedialog, WORD)
from tkinter import scrolledtext
import Pw_func
import os


# In Pw_func.py

import os
from tkinter import filedialog

import openpyxl
import shutil
import csv
import random
import string
import tkinter as tk
from tkinter import messagebox


def show_permission_error_dialog():
    root = tk.Tk()
    root.withdraw()  # Verstecke das Hauptfenster
    messagebox.showerror("Permission Error",
                         "Bitte schließe die Excel-Datei und klicke dann auf 'OK', um es erneut zu versuchen.")
    root.destroy()


def create_backup(file_path):
    backup_path = file_path.replace('.xlsx', '_backup.xlsx')
    shutil.copyfile(file_path, backup_path)
    return backup_path


def generate_random_password(length=8, special_chars="!#?", special_count=2):
    special_indices = random.sample(range(length), special_count)
    characters = string.ascii_letters + string.digits
    password_list = [random.choice(characters) for _ in range(length)]

    for index in special_indices:
        password_list[index] = random.choice(special_chars)

    return ''.join(password_list)


def generate_fixed_password(length=10, special_chars="!#?", special_count=1):
    special_indices = random.sample(range(length), special_count)
    characters = string.ascii_letters + string.digits
    password_list = [random.choice(characters) for _ in range(length - special_count)]

    for index in special_indices:

       return ''.join(password_list)
    return None


def generate_passwords_and_save_to_csv(file_path):
    try:
        output_messages = []  # Liste zum Sammeln der Ausgaben

        # Backup-Datei erstellen
        create_backup(file_path)
        output_messages.append("Backup-Datei erstellt.\n")

        # Excel-Datei öffnen
        workbook = openpyxl.load_workbook(file_path)
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        directory = os.path.dirname(file_path)

        # Festes Passwort für alle Mappen generieren
        fixed_password = generate_fixed_password()
        output_messages.append(f"Festes Passwort für Dozent generiert: {fixed_password}\n")

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            nutzer_column = None
            passwort_column = None

            # CSV-Dateiname festlegen
            csv_file_path = os.path.join(directory, f"{base_name}_{sheet_name}_passwords.csv")

            # Setze festes Passwort für den Dozenten in Zeile 2, Spalte A
            dozent_cell = sheet['A2']
            current_value = dozent_cell.value
            dozent_cell.value = f"{current_value} {fixed_password}"

            with (open(csv_file_path, 'w', newline='') as csv_file):
                writer = csv.writer(csv_file)

                # Spalten 'Nutzer' und 'Passwort' finden
                for i, column in enumerate(sheet.iter_cols(min_row=1, max_row=1)):
                    if column[0].value == "Nutzer":
                        nutzer_column = i
                    elif column[0].value == "Kennwort" or column[0].value == "Passwort":
                        passwort_column = i

                if nutzer_column is None or passwort_column is None:
                    output_messages.append(f"Spalte 'Nutzer' oder 'Passwort' in '{sheet_name}' nicht gefunden.\n")
                    continue

                # Basispasswort für die Mappe generieren
                base_password = generate_random_password()
                tn_counter = 1

                # Passwörter für Nutzer generieren und in CSV speichern
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    nutzer_value = row[nutzer_column].value
                    if nutzer_value and str(nutzer_value).startswith("TN") or nutzer_value and str(nutzer_value).startswith(
                            "PU") or nutzer_value and str(nutzer_value).startswith("FA") or nutzer_value and str(nutzer_value).startswith("LB")or nutzer_value and str(nutzer_value).startswith("RG"):
                        nutzer_value_simple = nutzer_value.split('@')[0]
                        password = base_password + str(tn_counter).zfill(2)
                        row[passwort_column].value = password
                        writer.writerow([nutzer_value_simple, password])
                        output_messages.append(
                            f"Passwort für {nutzer_value_simple} in '{sheet_name}' gesetzt: {password}\n")
                        tn_counter += 1

                output_messages.append(f"CSV-Datei für '{sheet_name}' erstellt: {csv_file_path}\n")

        # Excel-Datei speichern
        workbook.save(file_path)
        output_messages.append("Passwörter erfolgreich generiert und Excel-Datei gespeichert.\n")

        return '\n'.join(output_messages)
    except PermissionError:
        show_permission_error_dialog()
        return generate_passwords_and_save_to_csv(file_path)  # Versuche es erneut

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        return generate_passwords_and_save_to_csv(file_path)
    return ""


import os
import re
import ssl
import smtplib
import requests
import tkinter as tk
import xml.etree.ElementTree as ET
import pandas as pd
from datetime import datetime
from tkinter import messagebox
from email.message import EmailMessage
from requests.adapters import HTTPAdapter
from requests_ntlm import HttpNtlmAuth
from urllib3.poolmanager import PoolManager
from urllib3.exceptions import InsecureRequestWarning
import urllib3
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from tkinter import filedialog

urllib3.disable_warnings(InsecureRequestWarning)

# ---------- KONFIGURATION ----------
USERNAME = "itbg\\sp_farm"
PASSWORT = "!DokM1MLcgi"
EMAIL_PASSWORT = "quaSeu2i"
ABSENDER = "ml-projekte-it@mail.de"
EMPFAENGER = "c.mueller@mlgruppe.de, k.vosen@mlgruppe.de, r.panske@mlgruppe.de, a.borowczak@mlgruppe.de"
SMTP_HOST = "smtp.mail.de"
MAIL_BETREFF_VORLAGE = "{} Bewertung DokMBW"
MAIL_TEXT = """Hallo zusammen,

diese Mail wurde Automatisiert versendet. Bitte nicht auf die Absenderadresse antworten.


Gruß
Die Software von Andy
"""

# ---------- HELPER ----------
def get_value_from_rating(rating_str, label):
    """
    Extracts a numerical value corresponding to a specified label from the given
    rating string. The rating string is assumed to contain label-value pairs in
    a specific format, and this function retrieves the integer value associated
    with the provided label.

    :param rating_str: The input string containing label-value pairs in the format
        `label;#value#`. This string represents ratings or similar data, which
        contains labels and their corresponding numerical values.
    :type rating_str: str
    :param label: The label whose corresponding numerical value needs to be
        extracted from the rating string.
    :type label: str
    :return: The integer value associated with the specified label if found in
        the rating string, otherwise None.
    :rtype: Optional[int]
    """
    match = re.search(fr"{re.escape(label)};#(\d+)#", rating_str)
    return int(match.group(1)) if match else None

def extract_block(description_raw, label):
    """
    Extracts and cleans a specific labeled block of text from a raw HTML-like description.

    This function searches for a block of text in the provided raw description
    that is associated with a specific label. It uses regex to locate the block in
    the format `<b>{label}:</b> ... </div>`, removes all HTML tags within the block,
    and returns the cleaned text. If the label is not found or the format is
    incorrect, the function returns an empty string.

    :param description_raw: A raw string containing HTML-like content to be
        searched through for a labeled block.
    :type description_raw: str
    :param label: The label to search for within the raw description that identifies
        the block to extract.
    :type label: str
    :return: The cleaned text content of the block associated with the specified
        label. Returns an empty string if no matching block is found.
    :rtype: str
    """
    match = re.search(fr"<b>{label}:</b>(.*?)</div>", description_raw.replace("\n", ""))
    return re.sub(r"<.*?>", "", match.group(1).strip()) if match else ""

def berechne_speicherpfad(dateiname: str):
    """
    Constructs and returns a file system path for a specific file by creating a directory structure
    based on the current week of the ISO calendar year. The function ensures all necessary directories
    are created before returning the final file path.

    :param dateiname: The name of the file for which the storage path is constructed.
    :type dateiname: str

    :rtype: Path
    """
    dokumente = Path.home() / "Documents"
    hauptordner = dokumente / "Bewertungen"
    kw_ordner = hauptordner / f"KW{datetime.now().isocalendar().week}"
    kw_ordner.mkdir(parents=True, exist_ok=True)
    return kw_ordner / dateiname

# ---------- CONVERTER ----------
def convert_xml_to_excel(xml_file_path, output_excel_path):
    """
    Converts data from an XML file to an Excel spreadsheet.

    This function parses an XML file containing structured data about seminar ratings
    as an Excel file. The generated spreadsheet contains an organized layout with formatting,
    headers, alternating row colors, and auto-filter functionality for ease of analysis.

    :param xml_file_path: Path to the input XML file.
    :type xml_file_path: str
    :param output_excel_path: Path where the generated Excel file will be saved.
    :type output_excel_path: str
    :return: None
    """
    def get_value_from_rating(rating_str, label):
        """
        Converts an XML file to an Excel file.

        This function reads data from an XML file and writes it to an Excel file. It
        includes a nested helper function for extracting numerical values from a
        formatted rating string based on a specific label.

        :param xml_file_path: Path to the input XML file.
        :type xml_file_path: str
        :param output_excel_path: Path to the output Excel file.
        :type output_excel_path: str
        """
        match = re.search(fr"{re.escape(label)};#(\d+)#", rating_str)
        return int(match.group(1)) if match else None

    def extract_block(description_raw, label):
        """
        Convert an XML document into an Excel file by extracting specific data blocks.

        This function processes an XML input file, extracts structured data using
        certain labels, and writes the extracted information into an Excel file.
        The parsing and extraction are implemented based on predefined patterns.

        :param xml_file_path: Path to the source XML file to be processed.
        :type xml_file_path: str
        :param output_excel_path: Path where the resulting Excel file will be saved.
        :type output_excel_path: str
        """
        match = re.search(fr"<b>{label}:</b>(.*?)</div>", description_raw.replace("\n", ""))
        return re.sub(r"<.*?>", "", match.group(1).strip()) if match else ""

    tree = ET.parse(xml_file_path)
    root = tree.getroot()
    channel = root.find("channel")
    entries = []

    for item in channel.findall("item"):
        author = "***"
        pub_date = item.find("pubDate").text
        pub_date = datetime.strptime(pub_date, "%a, %d %b %Y %H:%M:%S %Z").strftime("%d.%m.%Y")
        description_raw = item.find("description").text

        seminar = extract_block(description_raw, "Titel des Seminars")
        ort = extract_block(description_raw, "Ort der Schulung")
        beginn = extract_block(description_raw, "Schulungszeitraum (Beginn)")
        ende = extract_block(description_raw, "Schulungszeitraum (Ende)")
        trainer = extract_block(description_raw, "Name des Trainers")
        themen = extract_block(description_raw, "Folgende Themen haben mich besonders interessiert:")
        zu_kurz = extract_block(description_raw, "Diese Themen kamen meiner Meinung nach zu kurz/habe ich nicht verstanden:")
        bewertung = extract_block(description_raw, "Lehrgangsbewertung")
        zufriedenheit = extract_block(description_raw, "Zufriedenheit")
        vorschlaege = extract_block(description_raw, "Haben Sie noch Wünsche, Vorschläge, Anregungen?")

        entries.append({
            "Erstellt von": author,
            "Titel des Seminars": seminar,
            "Ort der Schulung": ort,
            "Schulungszeitraum (Beginn)": beginn,
            "Schulungszeitraum (Ende)": ende,
            "Name des Trainers": trainer,
            "Folgende Themen haben mich besonders interessiert:": themen,
            "Diese Themen kamen meiner Meinung nach zu kurz/habe ich nicht verstanden:": zu_kurz,
            "Haben Sie noch Wünsche, Vorschläge, Anregungen?": vorschlaege,
            "Zufriedenheit_Ich würde das Seminar weiterempfehlen": get_value_from_rating(zufriedenheit, "Ich würde das Seminar weiterempfehlen"),
            "Zufriedenheit_Zufriedenheit mit Seminar": get_value_from_rating(zufriedenheit, "Zufriedenheit mit Seminar"),
            "Zufriedenheit_Zufriedenheit mit Trainer/in": get_value_from_rating(zufriedenheit, "Zufriedenheit mit Trainer/in"),
            "Lehrgangsbewertung_Teilnehmerunterlagen": get_value_from_rating(bewertung, "Teilnehmerunterlagen"),
            "Lehrgangsbewertung_Praxisanteil des Seminars (erster Eindruck)": get_value_from_rating(bewertung, "Praxisanteil des Seminars (erster Eindruck)"),
            "Lehrgangsbewertung_Präsentation der Inhalte (Nachvollziehbarkeit)": get_value_from_rating(bewertung, "Präsentation der Inhalte (Nachvollziehbarkeit)"),
            "Lehrgangsbewertung_Durchführung durch Trainer/in (Methodisch)": get_value_from_rating(bewertung, "Durchführung durch Trainer/in (Methodisch)"),
            "Lehrgangsbewertung_Durchführung durch Trainer/in (Fachlich)": get_value_from_rating(bewertung, "Durchführung durch Trainer/in (Fachlich)"),
            "Lehrgangsbewertung_Struktur des Seminars (Roter Faden)": get_value_from_rating(bewertung, "Struktur des Seminars (Roter Faden)"),
            "Lehrgangsbewertung_Allgemeine Atmosphäre": get_value_from_rating(bewertung, "Allgemeine Atmosphäre"),
            "Elementtyp": "Element",
            "Pfad": ""
        })

    df = pd.DataFrame(entries)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bewertungen"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    stripe_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        for cell in ws[r_idx]:
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif r_idx % 2 == 0:
                cell.fill = stripe_fill

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 15)

    ws.auto_filter.ref = ws.dimensions
    wb.save(output_excel_path)
# ---------- RSS FEED ----------
class LegacySSLAdapter(HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
        ctx.options |= 0x00000004
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        kwargs["ssl_context"] = ctx
        self.poolmanager = PoolManager(*args, **kwargs)

def download_rss_and_save_xml(server_num):
    server_prefix = f"{int(server_num):02}"
    base_url = f"https://{server_prefix}.ml-schulung.de"
    overview_url = f"{base_url}/trainerseite/Lists/Lehrgangsbewertung/overview.aspx"

    session = requests.Session()
    session.mount("https://", LegacySSLAdapter())
    session.auth = HttpNtlmAuth(USERNAME, PASSWORT)
    session.verify = False

    try:
        res = session.get(overview_url, timeout=10)
        res.raise_for_status()
        with open(f"{server_prefix}_overview_debug.html", "w", encoding="utf-8") as f:
            f.write(res.text)

        guids = re.findall(r'\{[0-9a-fA-F\-]{36}\}', res.text)
        guids = list(set(guids))
        if not guids:
            raise Exception("Keine GUIDs im HTML gefunden.")

        for guid in guids:
            feed_url = f"{base_url}/trainerseite/_layouts/15/listfeed.aspx?List={guid}"
            try:
                feed_res = session.get(feed_url, timeout=10)
                feed_res.raise_for_status()

                xml_filename = f"{server_prefix}_feed.xml"
                with open(xml_filename, "w", encoding="utf-8") as f:
                    f.write(feed_res.text)

                root = ET.fromstring(feed_res.text)
                trainer = extract_block(root.find("channel").find("item").find("description").text, "Name des Trainers")

                return {
                    "server": server_prefix,
                    "trainer": trainer,
                    "xml_file": xml_filename
                }

            except Exception:
                continue

        raise Exception("Keine funktionierende Feed-URL mit gültiger GUID gefunden.")
    except Exception as e:
        raise Exception(f"[{server_prefix}] Fehler: {e}")

# ---------- MAIL ----------
def sende_auswertung_per_mail(excel_dateien, empfaenger):
    datum = datetime.now().strftime("%d.%m.%Y")

    msg = EmailMessage()
    msg["Subject"] = MAIL_BETREFF_VORLAGE.format(datum)
    msg["From"] = ABSENDER
    msg["To"] = empfaenger
    msg.set_content(MAIL_TEXT)

    for datei in excel_dateien:
        with open(datei, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="octet-stream",
                filename=os.path.basename(datei)
            )

    try:
        with smtplib.SMTP(SMTP_HOST, 587, timeout=20) as smtp:
            smtp.starttls()
            smtp.login(ABSENDER, EMAIL_PASSWORT)
            smtp.send_message(msg)
            print(f"✅ Mail erfolgreich gesendet an {empfaenger}")
    except Exception as e:
        print(f"❌ Fehler beim Senden an {empfaenger}: {e}")

# ---------- GUI ----------


def run_gui_process(server_list, versand_aktiv):
    erfolge = []
    fehler = []

    if not versand_aktiv:
        zielordner = filedialog.askdirectory(title="Speicherort für Präsenzbewertung auswählen")
        if not zielordner:
            messagebox.showinfo("Abgebrochen", "Kein Speicherort ausgewählt. Vorgang abgebrochen.")
            return
        zielordner = Path(zielordner)
    else:
        zielordner = None

    for server in server_list:
        try:
            result = download_rss_and_save_xml(server)
            datum = datetime.now().strftime("%d%m%Y")
            dateiname = f"{datum}_{result['server']}_{result['trainer']}.xlsx"
            if versand_aktiv:
                zielpfad = berechne_speicherpfad(dateiname)
            else:
                zielpfad = zielordner / dateiname

            convert_xml_to_excel(result["xml_file"], zielpfad)
            erfolge.append(str(zielpfad))
        except Exception as e:
            fehler.append(str(e))

    if erfolge:
        if versand_aktiv:
            sende_auswertung_per_mail(erfolge, EMPFAENGER)
        else:
            messagebox.showinfo("Hinweis", f"{len(erfolge)} Datei(en) gespeichert. Kein Versand (Präsenzschulung).")
    else:
        messagebox.showerror("Fehler", "\n".join(fehler))

def start_gui():
    root = tk.Toplevel()
    root.title("DokMBW Serverauswahl")
    server_vars = []
    versand_var = tk.IntVar(value=1)

    tk.Label(root, text="Wähle die Server aus:").pack(pady=(10, 0))

    for i in range(1, 11):
        var = tk.IntVar()
        tk.Checkbutton(root, text=f"{i:02}", variable=var).pack(anchor="w")
        server_vars.append((i, var))

    tk.Radiobutton(root, text="Versand aktiv (Online-Schulung)", variable=versand_var, value=1).pack(anchor="w", pady=(10, 0))
    tk.Radiobutton(root, text="Nur speichern (Präsenzschulung)", variable=versand_var, value=0).pack(anchor="w")

    def starten():
        selected = [int(i) for i, var in server_vars if var.get()]
        if not selected:
            messagebox.showwarning("Hinweis", "Bitte mindestens einen Server auswählen.")
            return
        run_gui_process(selected, versand_var.get() == 1)

    tk.Button(root, text="Auswertung starten", command=starten).pack(pady=10)
    root.mainloop()
    


from tkinter import Menu, Tk, Label

def show_output_in_window(output_text):
    output_window = tk.Toplevel()
    output_window.title("Ausgabe")
    text_area = tk.Text(output_window, wrap="word", width=80, height=20)
    text_area.pack(padx=10, pady=10)
    text_area.insert("end", output_text)
    tk.Button(output_window, text="Schließen", command=output_window.destroy).pack(pady=10)

def select_file_and_show_output():
    output = select_file()
    if output:
        show_output_in_window(output)

def create_menu(root):
    menubar = Menu(root)

    # Menü für PW-Funktionen
    pw_menu = Menu(menubar, tearoff=0)
    pw_menu.add_command(label="DokMBw Passwörter", command=select_file_and_show_output)
    menubar.add_cascade(label="PW-Funktionen", menu=pw_menu)

    # Menü für DokMBw Bewertungs-Tool
    bew_menu = Menu(menubar, tearoff=0)
    bew_menu.add_command(label="DokMBw Bewertungen", command=start_gui)
    menubar.add_cascade(label="Bewertungen", menu=bew_menu)

    # Info-Menü
    def show_info():
        messagebox.showinfo(
            "Version",
            f"DokMBw Tool – Version {__version__}\n"
            "Build: 2025-05-21\n"
            "Download: github.com/ML_PIT/dokmbw\n"
            "© MLGruppe 2025"
        )

    info_menu = Menu(menubar, tearoff=0)
    info_menu.add_command(label="Info", command=show_info)
    menubar.add_cascade(label="Hilfe", menu=info_menu)

    root.config(menu=menubar)

    


from tkinter import PhotoImage

def show_splash():
    splash = Tk()
    splash.overrideredirect(True)
    splash.geometry("320x380+500+200")
    try:
        img = PhotoImage(file="ml_splash_with_text.png")
    except Exception as e:
        print("Splashbild konnte nicht geladen werden:", e)
        return
    label = Label(splash, image=img)
    label.image = img
    label.pack()
    splash.after(3000, splash.destroy)
    splash.mainloop()




if __name__ == "__main__":
    show_splash()
    root = tk.Tk()
    root.title(f"DokMBw Tool – Version {__version__}")
    root.geometry("400x200")
    tk.Label(root, text="Bitte wählen Sie eine Funktion über das Menü aus.", font=("Arial", 12)).pack(pady=60)
    create_menu(root)
    root.mainloop()




import base64
from io import BytesIO
from PIL import Image, ImageTk

SPLASH_BASE64 = """iVBORw0KGgoAAAANSUhEUgAAAUAAAAF8CAYAAAC+KGJ8AADg20lEQVR4nOz9eZRl1ZXmCf72Oefe956ZjzjzIDQPaECABhCSmGcHnFlCCkUoIjMqO6urO3tld1Z1ZVet6srVnd0rcmX26sqVnZERkRESYhAzCHAGMSNGCQkBmpCExCRmfDKz9+495+z+45x73zNzc3ckheQ4dj8tZOb2pvvu8N199v72t0VVlQ4dOnRYgjA7ewM6dOjQYWehI8AOHTosWXQE2KFDhyWLjgA7dOiwZNERYIcOHZYsOgLs0KHDkkVHgB06dFiy6AiwQ4cOSxYdAXbo0GHJoiPADh06LFl0BNihQ4cli44AO3TosGTREWCHDh2WLDoC7NChw5JFR4AdOnRYsugIsEOHDksWbmdvQIc/HkSk/X1X9cGd/A7w+32P32V/NK/ZVfdfh/noCHAXx29DCKq61fMn32PytYs9b0fv/2a38fd5r+Y123rP33abmn0iItvdnt/38zq8NdEtgd8mWIwYmgt7Wxfvwgu/ed5ihDj5nMn3nPy5PZJY7L12tJ072v4dYbFt/V3fS1W7qO9tiI4A36ZYSGKLkcvk47/N+23rNb/N3xdu145I+Hchro6wOuwIHQEuYSxGEDtari6MMLf3mrcCtkW0HTpAR4BLGtvKH74ZkuiIpMPbAdKNxdy1sb0iyGJVzoXJ/4WvbXKCb6ZwMfncHZ1Gv+127mj7t7dd2/r8xXKP2/vcP1RRqMNbBx0BdujQYcmiWwJ36NBhyaIjwA4dOixZdATYoUOHJYuOADt06LBk0RFghw4dliw6AuzQocOSRUeAHTp0WLLoCLBDhw5LFh0BdujQYcmiI8AOHTosWXQE2KFDhyWLjgA7dOiwZNERYIcOHZYsOgLs0KHDkkVHgB06dFiy6AiwQ4cOSxYdAXbo0GHJoiPADh06LFl0BNihQ4cli44AO3TosGTREWCHDh2WLDoC7NChw5JFR4AdOnRYsugIsEOHDksWHQF26NBhyaIjwA4dOixZdATYoUOHJYuOADt06LBk0RFghw4dliw6AuzQocOSRUeAHTp0WLLoCLBDhw5LFh0BdujQYcmiI8AOHTosWXQE2KFDhyWLjgA7dOiwZNERYIcOHZYsOgLs0KHDkkVHgB06dFiy6AiwQ4cOSxYdAXbo0GHJoiPADh06LFl0BNihQ4cli44AO3TosGTREWCHDh2WLDoC7NChw5JFR4AdOnRYsugIsEOHDksWHQF26NBhyaIjwA4dOixZuJ29AR12LnSRv8nv9Y5xkb+Z9EG/yxtPbqBs/WfZ6vPM+Am/3xfpsATQEeAShpLoaiFXGLbNHVvx0bw/RJSAqmJkYnGhClg0RMSmv4cIxoxfLwIhBEQUY0x+Tf4UVRBBY0SMQwGvYAQMEfmt2G6CINsv0WGpolsCd9iK/N7scwFiCBAjqBIjCBYjDjAoSgyKqoJExBhUM/cYCEAUmK0qFDDW5i0wiRGBGEP+UNOSH0DzTFgkilXzJr5Jhw4gqrrYKqjDksFiS9btIRFLirlier0CkkhHgTq/pTXNsyMxBIy1KMLMXI0RS69vqZutqALWGPpWQCESE3FaS4gBYywCiIKgaPSIKBhHBAxmPjlPBIXzT/D07C7w6wAdAS5xRNCYWEUXUoLZxvJwTIDp9SNEDKoOxaalrKZVqzEQFUKosc4Bwkzl6ZeJtH7w41+zZW6WQw75EFMCGqDMn2Bt+rSQPiWRH2AUjOTtBlTSe0laJSMTZ3PzlSYp3uR/ycR36bB00eUAlzpEM5k0ZCDzfqBbE+HkHdMrODGoREJQrHHYzFZzVaAsLcEY5kaeslcQS8d3nniO62+6jZtuu5PXN27huOOP5sJzTuOzh70HA8Tg6anB+4iq0Ctt+7lRA0YURNGoKUWYN1d1YrPfVIg3+b07LEV0EeCSxiIRYFu8WJwYtl5OBgKKYABHzNGfzYXfUUxvWQOPP/UKl12znutv+TYvvr6ZwbJVmGKKmS0b6bsRJ33+U/zFl87hMx99Nw4IdWRQpOVq7RUn4KwSYo2opshTynZrmiiw3dZFI8D0l3H1uIsBljI6AlzymJ8D1DcZEUn72sjcaEhR9BBT4FWJCEYS+amBh37wG6649iauv+0uXtk8y9Rue6Bln6GPOFsSQ82UdYy2vM6UeE495jP86YWn8+kP749EGBiIdcBZwRiDaiCqxxqLqm23RiaLwTIm644AO2wLHQEucSw8+NtQ8bWYv7KMhBCIpGJEFYWAYC1squDJn7zEJVdcz/o7vsMbWyqW77E3w2gYCgRr0ZAYqygKYh0wvmbglLkNL7F6WjjpqE/yFxeezcc+sD9TBqJXjCjGgIhAjNgcsY4JTdK3mIhkF9c6xvbxDksXHQEuYSymA2x+N0DUiGAwAnVdoxooiz4hVKmooUpUJaqlEkENzAL3Pvw037j8Ou78zvfwDMBOQdnHI9SqhBiy2sVCtFgxiChKwCA4IlrNwWgjq6aE4z93GF85by2f+tg7sYCGSM8YNEJpIfgKa0CMjAWGmQcxk4KZDh3moyPAJYwkWVGsGddEo0ZEpI2bVJW6rinLJtc2JpMQwQPBwJzCA4/+mr+/+CrueugHzHmLlsuIMiAalzR8AjFGosSkC8Qi0WIQIgHVkD7bGCQq+CESExGuHCgnHvVp/uzCcznsoP0oSFrAMPJM9xxoElEnDY1PX866lAgUs2gxp0OHjgCXOBJlKDFGnGnCpiRwFhHEjEXFISjGWqoAzsJIoRa4+b6f8ndfu4z7v/ck9JbjBqsYRYsUPeqoqUgS02kmjcBZwKhBNPWdqEZi9IgJGCOICKqKiOBQHIHhpleZsjWnHn8kX71wHYcctA8DUhZvVHkGhcEQsY0WRjVte0OADToi7JDREWAHIBJjKgyIya1nCIikJg8cgdx+ZhNFbpyBux74Pl+/Zj3fffLnbJkNDFauYXYojLzi+tP44JEiFS0g5iptQ0b5PxXE5AKG1qT8nSZiFkHFoj5CVFZM97BxjjdefJp99ljGiUcfzhfOPJVPfuwASnL1OSqlUbSuKMscAeq4ua/VbHfoQEeASxqqqaPC2ByVhZDExGLwQal8oOgPmBmB66Xl7pYK7rjncf7hG5fy0PefZFSuxE6touhNUXmoa6HolYgI1WgOsSnTKEZBDaoCalHcPM2KlYDRiGjuAFFLMKDWIcYQYw3VCOOEQWGgqhhufoOVZWTdKcew7vQTOfKQ/SkAX0X6VulbO7+nuPmtiwA7ZHQEuETRHHYRJYYaY1zWAypRhWgsPtdKPbCxgptu/z5XXHszj/zgCUZRKJftRuytYK7yRLW4okAxhFCPc3Ix/WyWvqqGiAF1mZNSPtBoivLaTg4pwAiBRlhIMkmIqffYREMhSo9Z6rkNTBWBdaccw1e/cBYfevdu9IFQBXqlxej8VW9HgB0adAS4RDEmQEmtarZAgaquwDjEGkbAxjm4+qa7+Po3r+fxnz+L9FYSi2mKqZVUdSTGVNgwRYrUQj1MH1CYbJIAaJP7M8w/3bIURTT3u8n8fJ1YCAFTFBAVDQFjwEgyWBCNhLpiUIANM/jZV1kzpRxz+CH883/yFT76/ndQMG6haz5LkE4G0wHoVKBvKbQtXYs89tsFLRMtXgvfTCafAzEmYvKAD0BREoBfvzTHnfc/xKVXfYsf/PhppFxBuWofgkwxUkM9FzNhCc45fKihrqGwIBFCyH5XyfNKo2z9+ZONu0ZALIJFNXf/agQLsRqCtYgVQl2BFQrrqGqPc44UoybFX89arCqOiJvYp1vvn8mfO8Kb6Yp588eosyp866CLAHcyFsZDhh2LkaER8pr2DZRmaZc8+Yi5gqsmk5FNLgVOQCK1HyXiqiOKwxuLWnhhA1x23S1cfu3N/PL5F3GDFXgpgJLaJMMDj03vq4qRicpxu6Td3hduvkl+nkmCZsRgsFAHVAXjBDWSJDOhIod+SB0geEqx9KyhrmeIw03svbJk7bFH8JVzT+Xg9+9DCVTVDEVRYMRR+UTIhbH4ekThxkv+xmsQyFVvWvcaSJtnTIoVQhO0ytb9xk2XSTo+C772xBFcpPO6w05CR4BvAeyoG2Nr8mueZfLFufib+rrGFUVbCFCVljOrUIOxGJP8UZ55peaqm27nsutu5idP/wbpL8cNllOrIeBS320jh2F89RsUSb0g4+1f7KqeNFdo/xbBe6TfBx/Q2tMve8QI1XCYGoodSe3sa6grej3HlDXUMzNUMxvZe48VnL/uFM49/UQ+sH+PASBRMWGIKxxo1kbbYoJ4UsW7lfpkyY3KWKYj7QaPO0pCCFhr2+8agmLtYl92fATHisoFu6MLA98S6Ahwp2JyKba1CcE2TYvbByJQE1FiFIwpIcjYaXle5j+bE4SxNngE/PrFwJU33MIV197AL55/GTe1inJ6NUEcc1VERYiYcQ6t3eL0B5PFx6ILCCMTXWtIsPBLSPrupXEEXxGCx7okgvZB0Wgwrk+MmigkBqZ6oKPNzLz6HO98996cedLR/Nn5Z3LAnr3WRivWkdIYXO5zsdZS+4hxhpgrzIWR1EucPQyrqsIYky27kpGrZp9p7wM9Z+cfh6hE9SlC1Hz8JsTWkz3IW7cRjq280h8789adiY4AdyoWy0Vt42LQRf4ukcAcBgMUgEND9uJjfG3VIRU2ohhGmRh/8dyIK6+/lStuup1fv/ga0ZVMr9qdLcNACIqUU2lZ2NrTx0SCstgCfdJ9NPfmTmzvmAQnL3wSEdSeXq/Eh0DwHlumskXwirMWG8FqgHqIH77BO/fbjQvPP52zzjiW/VYm/0ALRB8xxlDmhg8NaZkaY8S5HPdpTKQk4KNHMdTBUBbp8cqDlSTynlyimuYroqmyrZqXz5P7YILIhK2ivvnk1zQhmlTo6Qhwp6EjwJ2MyZ3fOiwzeXFNRlY5qppwr/KMUi8tyT+vdEV6ICpRakZVhesNqLAE4KkXPBd983puuPlenn15A6E3wC1bQcDgQ4RikD6zDukiN5KkJxMEKM2SOkeHO7qATRPxSJxX9wDFOceorhBbYm2BrwJGhNKC1HNIvRlTbea979ib8844kXWnncABeySLLEuifUPAR/JydryPLKAkw4QYs77QCFHS60fAlggXXXoL04Mep59yNKsHafc7Aavgq+RCY0229I8+LZNFUp5wwo4rice3syOkcc9eeMPrCHBnoSPAnYiFZgSpX2EySoCFBNiSn6ZlZbrAU+GgyV+NfI33HutKrCvxwA+eepGrb7idq2+6h1+/uJHB8j0xUyuoUKqoYLJgJGjK+juXwqHg8/ak/xo/6GQhKLnWOnkBT0Z5sf0arROzzi/+qkDQiDUFxrhEMHWFCbOYegsfOHAPvnzOqZx7+rHssQwIYGIqNqciTMidHSlHGUmRHiSNowGqqgIMYnuMvFL2hA2zys13P8h/vuQ6fvijp7AxcvhhB3P26SdzyjGHs+eK9FV6BiRADDUuG7P6epQ0jyoo4+Vxm9fbZu5i/LftPaXDHw8dAe5EKMnyvcGYBKElkvaiyvmlBSsvoxC1xlghEpkLQ8T2UUpq4LEfv8pFl3+LW+98gFc319ip5RRTyxkRmZ2dw5YDVCxxNALjKKeWEWPEz82kNaS1LLZUT9XqLGhGJlgtTkQ4sVUdi5q2Hzh9ibwsjTAY9BgNtyBxxMBF6tnXOPj9B/Kl805n3UmfYfc+2Ah+OGR60AeNaF0jZY7/YkgT5bKNvohic3Enogglw5ym2zSCG295gIu/eRXf+/EvGfZ2o1y2gp41bHjlN/RjxWEfeR/nnHISZ59xFKv6MHBQjSLOKGVhs25Skug7FzkaraGZzL0uZDkZH/MmP2jpSHBnoiPAnYgmApzE4hXf8fNZcLgEIXiPFJYKoSap4h750TNcctV6vn3n93ht04hyag1S9Nk0N0KtYAqLGsmkJKm6qYqvEiXbwmCMoQ6BrbE91+g473dpWIkkvjaaq9EkziwxSBxhdA7LLB989978yQWns/bEw9mtSEtc8YFCUkBKqBOTmRRqaTCIy3JWiYRYt5XtuVgjZkBAmAtww22P8Dd/fxk/eurXUC7D9Fcwa/rUQGkNhQZKCYw2vo6MZnjfgfvw5xeex6knfo59dksHwCoIqZACkTCxD0yOjGUxSpP5EX9HgG8NdAS4k5EsoLLeTHPinhwpxKQEgSY6TNINjR5jUsW0rj1i+kRnmFH43pOv8g9XXMf6ux7gjVlPuWwFQWyWsCS3PREhanOZhjxPY7J2OdatJa5q+ibMOBqFvMSN2CwjiSEAFtvq5yLOlXifhcomLXfJ0hPxFb3ZGfo6x8cPfg9f+eLpnHjMwSy3IFQ4wGKwWNQno1WNIDYS6yGmsISRYssSok9zgwvLTDVCygGKZVOEW+96lL+7+BoeePTH0FtJf/kaam8YhSZkI1uAkXOLio01Tmvq2U2874A9Ofu04zh77fG8c++Cglxoih4rAWukjcwNNhO8ScFvUwAa68YJedc1toW/+7kzbjHs8LuhI8CdjjheEhpJfnmawh0BQvbrixrx3lM607avBbUYVzDr4cFHf8VFV97AHff/gA0jkKnl1MbhzSRZGSxJrtFWacUDubWs2Z4MxSQPv3EtlHlRn0SMUWLtMcZQFCUaharyoIorCnw9onQWJeIEykIYDmcZjeZYZuHoQz7Keacfx6nHH0TPJHPTnvH0jRCiR6OlcFNZkkI2ZRUgV41dL1V8rWF2OEcQQ9HrsdHDXQ89zt9edAXffeKXbBgJ06v3Yc4L0QsUfQgBwafiTC7oNDcjiYpRT4nHz22EahPvPWAPzjr1WM4+7UTee8AyHCChxlJjjUPz3cuaVIgKQbO/Yb5vKKimajX8/gTWEeDvj44A3wqIyY/PuNz+1Ugt8mzJECPGFAQMW4YVpihwVhgBdz30HJdffQu333MfM7ViBstRV1IbQx0VNeMiilFaOypRk+TAjQeghkWrvDAufSwkPxQkm5hGzaErJrWtIah6LIHSRlyswc8w2vIGA6Mce/RnufDcM/j8J99NT8CZpqKrGK1ThJiJNyj4CNbmeb7qQT1iLJU3DH2k7BUgMBfg1rt/wMVX38i9jzzJiB7FijVUWlKpxZY9Qp3a9kxZYMIoC7ltLqJkIlMFjYh6Bj0H1Wbq2Q30pOaAvVZz0vHHct4ZJ/LhA5fRJxWerIGeKajrilCPWDY1PdbjNNCxzyEiLJ5G6PDHQkeAOxk+pAgvXSMpGkPjmAARhkHBDog2PWNO4Y77HuPrl32Lh3/4LJuGinEl5WCKUYhESaJeH8PYWDRXjxsCVM0SllY1rW21t6nYLkp8+REg5/MUcRYRS1RN2sGY3seagIsVVofo7EaWF5Hjj/wUf/Hl8/n0we9opSxk2bGEJJMx5Iq2Eeq6pihLfLN0Juc9tSZK0brVeODWu5/k65dexYOP/ogRfdz0bsx4g+0vo1ZB6zon3hLpi0ZsjnzTFhhi0/HSCJwR0BrqEf2e0LPKlk2vY2LkHXut5IQjPswXzjqVj3/oHTR7sgdIDIlOjaA+jX8XmzV/kvapxghiuyhuJ6IjwJ0AbSOsdOFBJgINWXOnmZiEqo5QporuxgB33v8U/3Dp5dz/3ceJborYW060UxiT+nMbd5Yk4csCZpIrs+bPjORWEBXIPa6N2DlFg+lPiTzzc2GigyFiW21fkfpj1afPsoI1CjrC+CF+9nX2Wj7g+M9+gq+cvZZPffRdTFmgjlirVLEiGHCmwOBSddmn4MiMS+IEjTkiFaqcG/XA6xHuf/QZLv3m1ay//T5MuQI3WMnsSOhNr0zi75k5KAvKnqWuR1iJGAPV3Byu6Of83XySj82aNUaK0qXhT77GFZZeUVJVQ+JwC85vYo8VPU4+7iguPHcdH33fKgqgyFrCAp97SiKopM4Wm2LdyXOhI8Gdg44A/8jYaneLoCimTYdHvI+tJx8WXp+D9Xc9yj9882q+/+NfEt0UbrCCkVpqERCL+iwMtEW2dYFeWaSiSXZkVkn9rlGyYFctrXakiXhyFDi2r9qaAA0eo027mCNqxBqhXwpaD6lm3sDokOUlnHFSWioe2bg215G+zctsDUlFgyViiWpweTeIQvARa6DyI8regFqhSpUGamD9PU/yjatv4jvffYKZShksW0NNicchbkC9eQacpez1CH5ECBVlLxWBfEx9vd6HHBE3y9JxDjQCoopKM44zWXIhgrFFKoLEGhNGEGbZbargqCMO4U/OP4MjPrpvO8DJaZXI0CbNZETwPjvZRG37kTv88dER4B8ZW/nhaYr4fIA6RLB9TJHkLDMBbvz2Y1x2zXrue+QxpFyGnVrOyBuCFLiioKpnsE7QaIkIxhXjfJxIKi233QeT+jzymjjr/DIB2omIREUI2shYxj41DQFCJlOUvgWpZhhtepV9V09x7tqTOH/dqbzvncuZMok7Q+3p9xxCEis7GdeXdcKZTRsnLSKNUrLyNbGYogLueOhpLr92PTd++16G6nBTK5ByiipavBRoUExREr3HoWissEIivKhpP4tLYm9NTtSTkVjbt5xW4ePHNOdqSamEtOy3FNZAGBKGG9F6M7svLzjp85/m3DNP4ohD38UgfUG0rugVrn2f0hUpWt8OAXaFjj8sOgJcFIv1u8KbSlhPCmEXe3gBASoBEUONa3NZL26G69ffz9U33MpjP/4FwZTYcproSrw6kIJoLDFUGNNEeAaMHcsEjc0E2LQeLNJ3nMqnjIlt7Mg8JsD5iXqjyUkl5c48hY1Uo83o3Cz77LGCM48/mi+etZaPvGcZLrejxSpgreCcwU+08SVhTiMldkTNn9R0U8QKCCCGaBzPvryJf/1v/j033vUg0a2kv3pvKi0YiaGqhegsphgQap9uBL7CkoxTDYpGiNYixiVpUKgRI0gcax3nmTYIKSWRiyImR4LRhyTxKXp4XK6uB5x4SirCaBNabaFnKo468hN88ZzTOfYzB9FvvrOChEBhLMIfmAB3cD4udXQE2CJu4/dJzO/bVNW0hGmkJhOV2/bfWdoyfk5IgmBjCBoZquKzbOKlGbju1oe49Iob+fHPn0HcNLacStXaGAkqRJOJDpMv1uYCNW2RQJv+kqZpfxJKmsbWTmeLORkvbc5LnEvbVwesmdD0WcFXNb3CEaOHahNu+DLvP3APTjjmGC4470zet88UMaaKbiFgtCE42mU1MunHHPP1adrtm3cMBOo6IEXBqzM1d9//KNffdDuPPPYEz706pFh9AMFNMaxqTGFQI/jagyswtkBj8voTLDbLh5olfkTHKYCtjn2rfJz4bVvnRHNjSMYNViNGAlZrhsMNxHoLn//UIfyTr3yBoz79fpbbZOKgAQryWFKhtdsCxktk0vEZnz+peCK2aUoco+W5NuJnAQFu5zVLFB0BAtsmvIWGRsCEv1uSZAAxmUM1J6nmpPfknTtVXRUjJrVCBfAGvMDzmwLX33wXl197Kz/+1UvU2kfNACl6GOMSgeYoRY3k4omZiFYWbKMsFu1NPHOSANWPCRvSsi6kuRvki9EZ8FVFz8HACvXcDMO5zbx3/zV88fTPccG6k9lv92mCQim50KogWmNkMr9pWsKZf9KlAezzH2i2XVGx1MCcTzWbCNz38C+49Nr13HjHI2wJjnIwhS1LagXKEh8NfmaY9H45jykiWbaTl+8xZpnQNnqZiRg1i/sbTu7r5qZCut8kuVG6+VgXMLGimn2DvngOP/ggLlh3Ksd+5lD2WA6NFXfZszQqqPHhiNR1SFZddoKoJ07AhRevzHt8sfPadAQ4gSVPgDpxwreRyGJ7pFlJ0lwiEYNpW8mQdJ2lauvYPcWKEFURLKMAIy8U/USCr26BS666hcuvu5GfPf0CvZW7M4wFtZT0p5czGib5hOaCQdoOmy9IAxpw+XOCQGiKG5PylYb4dP7F0OgDJSZXE5sF1r4egjhsWRCCgkQKA6V4Zt94GVNt4pAPvZvTT/g85607hX3XpM6I2ZFiNTLdt7m9SyGGiYt5vqRmsu9koti7aAQYQiAai0qy7m86ZWrgkSd+w9cvvYo77nmI12YD0l+N6a1gFMEUZeqVjr6dI+JjNkS0RXZ3mcyNTm4nrXh8HAM2hg7jKn7MudQxATYC81RQ8bHGGsVRY2NFPbuJnno+88lDOfvUEzj9pENYWeRUQZ2s/K3JMqhcbGlvZzmlYFqTidTgrAu3eyvEZlcyeWOhFbgvXS1iR4ATi7GtL0LmLSFaX7v2FXHihIcq6+6M5ItGPb6uKco+VbD4rDz51UuRy669kevX386vnnsF6U2hpk8oSiimmKtDsqNyLl8ZYf5SrCG52BBgJApEydFKsyTOXR80uTXGF/vkmMqyLBkNZ1HvKQc9jBGqqsIaKB3MbX4DqWf46PvfwYXrTuKcU49hr2XJrDmSohZr0qbWtaewkvRvqm20tSMChAXRS7uXQ7axckQMPkKttPOJmxvSY09t5O8vvpKb73qY12Y8FIM2fSCSpsuFGBExmLJHVEsczmKcbSOlrZe68yNDM0GUojqPAG3jgxhdOk9MWuarRKwTQjXCmcCgMIThLPXMDD3j+cRH38fZpx3HSUcfyT6r09JYoqIhpRrmRhVFUYCx880yNOV+m8H141N2vM1xq3+Ni04dASYseQLcCo1geMHPxREJBKKmpn+XK5lCWl5FFDGWuQDRJtv5i6++kcuvu4XnX92I6a3AummiWrwaarIcryiTiCz6sS4Q8obkE1ZMCgkkR5vNxuctGPv0TcRXqpg87NKRL15xKcdmwFkh+BqrHqupcyMMN/Hpgz/E2acdzxknH83uy1N+zwJ1NaQ0NjmjqOJyvjCEgMbU8mXsZBV53m7eCosTYBz/MZuHRk1D2pVkYmqymqcCfvCjV7js2hu5496HeOalN3DTq8BNE6Rk6CPUEcoCUzjU162j9eRWjG90OyAHaWRDtASYdT1Ek9oIiSGlEgTww2QvBvREKK1SzW0gjjZz8AfezRfOPpVTjz2cd6wp0plUVxgJFDaTqqYUgqIEFYwstnXzI/0dL3GXLvlBR4BbXYnbJruEpI9tOitS21p7qYZIDAY1FjFCyMviJ5/ewFXrb+O69XfyixdewQ5WYnrL8OpQtcSQTDpdr0StI1RVGgRkDY0MxGgTbTQOwiYXWRqRLfPm32ob8zURY9O9EecRYF17in4PQsBR44jUW15nWQHv3H8N/+RPzuO0445g9SCRXgiKaMAKFNkv0Jg0vrIKnrLXy1uwde50O7u93bfzH4xpoJNNS3tfpxY560pQ0/ZJKzDyKTIsyrQ0/sHPXuSb193Mjbd+h1c2V9Tap798DWoKZoYjIkrRc8Rs1LAQTRsgmrpD5n2PeeeIH+/7dgyAyf+BWIvWVQ62DPg6u8lYNFQUpeDwzG16jTjazKEfejcXrDuFU4/5DAfu2ceSRzdqJEaPs8mA1WvKMdtFvWcWy113WAxLmwB14qcwL6m8WFYomZXqWFKimhryeyWKMgoRsT0CsMnD86/W/O1FV3D7vY/w1NMvUK5YQ296OXO1TwPanCP4JLRt2qJwNhUgNKTlWcjC5NZifuzsklpvUzM/MRuTNsOKFJDAPJdp8oWtTdN/JErEicdUabraCltxyAffxRfOWsvZp30OC/QtVKNAWVpsjjpygRvUp4FFjRM1hhiTuFiseZMSjm1dsONc2Lw5GnmCG6YEn6JAV6bdNww1QQzGJAfsnz27kSuvvYWrb7yb517ehPRXIP1phir4kKrLqpq/U14miiRbK8niZyZJcF7GMkXgbZ51ouiec6yBVFlPXYcG9SG5zSBEDdSxxjmDM1AaqOa2UM28zgfevT8nHvUZ/vzCs9lvjWUgaXe4WKd+adE8k6RJd0DTc6wTeZtt7v5OHgN0BMhkRg9oSXDBX+eT37yCgqUKkSAOKQxD4LGnXuay627m+tvu4eXXR5j+CmxvmjpaQkyVVjWChpqiMNRVhSkLBEMYjZCiwDlHPTeLtWlJZURIfqImWVmJTTM2tM4Xm4FmDHj2ZpIsx5DGYCHVodNzxCIELEPCaCPTJnDkoR/mK+eexjGHf4SVBa3BpxgleMVaQ4yKs5KuaQHwjEfTmfy7e5MX1zaKD5OHKOcRm2KFiBmzTCR/5/SVovqUGQAq7zG2ByJpFMCzI66+8XauXH87P3/2JRisoLd8BbN1MsFu3jARX0MeeZB7Y946jwSZF1nPI8FWARBbGVRDsgZDiB6Tb4C15lRJ3s9WoJBIqOYYbnqV9xywF+edcQpnn3Is792/zxQk5VMYUrqJ1UALmbeK2eFhWOIkuLQJsImOdKwHS5qrnMeZUOmPZ0Eovk5TxEQEr0K0ZV52vcHXrryJG26/j1c2zWGnlyOmIOSiRVQLaol5gFG6QGYx+HHRYuEZGRcKYQ1jF5EIEokx5K9gEeuIXiGEtDSNSuEcdT1CYqr2Ih4RJQw3scJVfOKj7+Of/MkFHH/4+3GAHwaW9y2i4xkeSvK3E/L2TN48FpHjLCrPaB5YVJ4x/z22LdVYWLUdhzJJcJNf344CVUYBojVEAy+8ARdfs57Lr1vP0y++QWWnMf3lFEXBaJRuJiIW9R7XG6TRAu37ByZb1yKaSbEpejHhqLPIXpAmko85pWGIWKKYNjUB2hYqDB6tR8TRFvZaPc1ZJx/NX3zhTN6/bx8Tm4JJqrR7H9MIVKCqPWXp0jjQJlMieYynCBp8MpFtblxLmASXOAGOq8CpZ1az+BTas0dgNBzS6ydPujomYqk0YsQwB3z/iWe46IobufmOB3ltNiKD1eAGSFFQh+zwAqTozBHbzE4EmQMJSW+GtLozM28K3PiCbzRpTXVVtPGcM/gY0aoC53BFL83PQKi2bKLX7zFVGvxoC1s2v8bqVVN8/tMf56vnncYRH38nUxaqCqbLlOurRkMGvX57E2ggC4XV2yG9SYzfoYmYGvJcPFe43RzhRIFE56syEwHqOHealrCgxjFbR2pjsBZ+/dKI9Xc9wCXX38GPf/U8QWEwvZKohiAOH5ToAxiHcS5pIzXQK0q0seA3Bj+5/fmmsO1vMLFEBhRLzAa12Zwsd/U0UWiS7lgTkdEs9ZbX2W+3AWedcgwXrDudg7PxQm6WQRXygDtqX1M62wqqY0yCaptziG1qYQmTHyxxAlSgiuPzIN2PA20rVkiqPnE9PEm7NxdS1XEEPPTYs3z98hu575HHeOWNTUwt341hEGp1lMWAekF+PbWXMdFrulConC/eLK2YeCA9d57AOZ/A0dF0cGAlVWRDIHqf+lUVpnsF9exGTL2F1VOGIz/5Uf78S+dwxMffgQsp99S8hciYk1yq+NBEmtJGQuS9ZbYpIW4wT+Iy+Szd6oF5r9peLnbyE2OOvlqNXlsxT7swRFAVjE2TgrP3AUZgpPDCZtKwqG+t58mnfkUw/TQX2ZTUFGCyHrK5TDRFXCZ358TcxcHE0nnRS6o1lFgos5H23BNiqhrLOOqPqpQ9R6xGGJSCyOY3Xmbv3XfnhM9/ivNPP5rPfnw/DKle1LPgUByBGEa5yyR9gkbAGOpo2g6fRdYcSwpLngAXDqiJocbXQ6wIrtcjRqHCUKlBsx/ffT94nq9943Lu+94TvDoLxdQqQt6NZW+aOgT8sEZciYhdMAUtLeHGM3IXTlVb7HRU5pkZyIQ0RAvElRA86n2SRohiNeCsYmPFcNPr7LbMccrRR/DVL5zFoR/cAwnQN+mTQ56p2ypr8jUTY+NeHPNWTESiwI40ZIt/qwU0qWbRJfGkPdWbq2kuKPbkpbFq0mUqQj2q6PUKIFJXI1w5SMcWeGWjsv7O+7j06pv43o9+QbADBit2T4470WBciYhQj0aginNFji7D1sS/SEpgoQ6yZc04QX6kvC2M738xpijXFQWGVAm31mJRRlteZ/fpyLFHHMqff/l8PnHQHqmWVwcGLltxhQpxgh95pCgxpiTZ1Mpk1njJYskTYJ0jOo0wqmqm+/kC0YAPQo1BXFrq3vPw01x6xU3cef932Txb0V+5B9pbztzIt3mh4D04S1H08FkPBzBumcpJ8hyieLOQABdgMtG+iDZOMGjloeilRv2qplcKfSrmNrzEnisMZ574eb76hbP40Dt3gwhFjFiT2M6r4Fyjpkl9smQz0lQscW1RqMm2pQs2kOItt02CmndhbTfiW/jdJsgiR3KL6TEndYONAU7IRKiZQtNNLeBUWNijjQg+eMQ4vDqCgTeGcP2tD/D1y6/j0R/9HOmtoL98N0ZBGA49tt/HuJK6SsUfY4SthNTbI8DmSwigARfT2iJmkp5fcU6VY2NSKyIhYnOeLyiUFky9mWrmDdas6HPkJz7Gl84+laMPfw9TQAyeIlYUzrY3gWEdKYoyF4qUabfoCKclgyVNgAAojEaRsp8WUcMqEICitIxImrLb7/spf3fx5Xz3Bz9lS23oTa1msGwlr22eQ1yRNHzOoaqEuk7OvyIpH1dOjJWM47jGxKwXy0vJRft3txEMTs7YjTGkz47JXkrrIaONr/Gu/Xfj5KM+zZ9ecDofeOcK+gC+psjLWutc6k+WZLradyZXun2avGYcGhUp+ltVxcfSGpi0sRpvdkNiccGSbxvfqcVkgWM+CU6+VtvPAKJpnxYYjxnVPK/NoWOizH5+UqQ82Kiao1eWqPcEEQIlXlKx5JXNcO/Dj/P1b17HI4//hFlvWLXHvszVylwVMWU/SaCS7H2su8z3ja2+d4v5SQMXU+Vac/W2dd9pdYja3jxdPmZhNAIVpHRoPWTVsimqmQ0MN7/GqoHhM5/4CH96/pkcfeQHEhFGqEcjpvplKhPJeEuWMvnBEiDAya+3MEfTVN1SNayk8spIBOPg9SF857s/5+8vuZL7H32cudowtXJ3glhqD7WPFP1+rsAqMbdrGZs83iD5z6mv24tVRYk0QrrGiTm2lcP5AhzmmR0YFSRfbm3DvYIrDNGPoJ7Bz21k3zUDLjjzJL549mm8e98BBVBVgX5h2zyWNULtA4VLXQUWQQmEOiXMWzPUif24Ve2jfYB525y3nHkEOEkIEyS2NcZLweZ7t6+b1K1NpgOa/Sjz37M57DH4XLEf59V8ruQaacoVabC8ilD5QFCT5EkkT8bb7nmSf7j0Gu5/9HG8SYJqrwU+5oqxkDwYiRNFqlyUiTpRFZ6PFO25FI3mKrCob6U4KukmGX0c70MN2LLEiUnFuXLAaGaGsrQ4iRQmMJrdiNMRnz7kw6m75LiPsdymm27fJKMKnZtD+qmnZ55sZol5D74NCHDhxTf/zqut1btZ9CT0ocK6kmEQxMGmADfc9hj/cOk1/OBHP6cWR2/5btQBajVgHSGOZTMWTbou60CS4SaNg7Cv89TFRrgrhOa2a3JkEyK2kUxMRIHzTkpNRGBEMvElKjQaCMN0sh+4726cdtLnOW/tCbzvgClKIHrFxEBZJpIIQRE7LmRoNiVduP+8JheUSQnFfHZpNqz5uXD5+rsS4GR0uaBosC0C3EYeUmN6vkx8XszdE5rzX4EJiUgMOGMQ0rLYiEsFMpda7GY83P+9X/P3l1zJfQ99n1G0FP3lRDtNwBLQtPAW8kCnRlY1roTJgm8dRYjSy9sfQD1Ww8QNM+vgjcUYh2JQ9RAiIorBJrsvFZwVVAMaKspCsOrxoy1INcMnP/ZBvnjWWs46+VCWF2BDsuzXUCFFarNrOuFb/eI8LBbNLozUfxtMnBc7GbscAS44hcht7vmEE4KCiMWQXHdRxVjB1wFnHdqoXqNSqSH0LEHgtRm47Z7vc/GVN/C9J35OjcMNlidremnGQ6aLTXMmbKzda4obTdQW27a0JlqLkk9oifmaVVCw0eSlbCK9xvYoHRYFDWiIWFF6RcloNKK0hrquYbiBD+7Z4yvnr+WMtaexz+6ORhpr2f7Q7e2l5HQbf1/0IMx74nYI8LfCIlnFrTbqd78Am7dafGlPjrRShw6S2oeHEWwmw/sefIrLr76G2+/9LpviMrydphxMUflI5Wts2WsW3ojY1udPJaYcsYApHDFEkpqPdmvSrJXxDlYZqwIXW1q3OeUsdjeaK/aSjGAdUM1tRvyQQw/6ABeeu5ZTjklWXBYgKIVJS3kNMZu+thWYPGs5tR4mH0La4k/KK26dAoHFDnmc/3t7XvwuBPqPh7cBAaYTJqoiMi4opDt6upv5uibGSFH0CGqpo1IUyVrpDQ8XXXkXl19zI7945kXmYonpTaMuzaDYGk0ENZ7vAM2CMd/tpcnRmXEk0LqH5OVMY2nkAU2mqs0Et2YZ6kNFrFOrVBjNMigsEofMbd7Ihz7wHr687hQuPO1z7Lkir6ojuQoM0oq4F3yJBZHYUsf8G8HERRrHy2XrUtZ1tko3sbJIEeH3nvgl/+Wi63ng+z/l1Q2bGSxfCcWAoVeqKFiXBkZZa9tuEFcYMAbvPeojqWumORjzSaKlhgnSa3OM+e/NCmOhiLqJRA2CE0Os5pIVl9R88N37c/5Zp/HFcz7LigIGQFUHSmcoRfD1CGdt62/p6xqbq+BjnaJCGxCMbygNOgL8A2GrnZwDJSRVdD2Kc5KGVkNiBU2l3lnvCa7EA89t9Nxyx/1cfMlVPPvCa2weegYrdkeKZcxFQx3I+bqmzLhgydUcQGkei1kpCOMIJkWPk5FiKmKQ5lAAuIIQk2W8lUA9qtLw7FxUsdZijSKjGcLcBj727n35yrmnccYJn2OPFemjrdCOny1d2h1mgYSlI77FsVUk3J5gY8GgD8qwrnBFL/Ub+5hmFEu6f93/6C+54tqbueWuB3llc6C3ak+inWLOR0xRpqKFgcIVKXIPAWPLMTHKBLG15xRJIpMjPBsBItEkgU8w5IpytuQXTeS0sCIdI9Yks1UnHuoZwtxmBj1hvzXL+OoXz+aME49mz9WJyJoFuQHqOp1DzuZFekxmsjZ3v0QfUj+yTBLam0G3BP6dsXBjJakxUhJcxnIN9ZHCpV5Oj+DzyfrcZuWS627iiutv4ee/foFBuZyoFikGBHoMPWCKcXN/mxNZQCRp5hdjJWH65OaOHOcJXydeoykJ7TQN36k1C6/x+eSS1rxTtKYazsFohkM+/B6+fM5azj7pE+w5gFJTgl0bO/X8xZuTlXaWREeA28P2CFDJ1vRFj0AyUzXG0RiUNelcTzoLfvDUG/zNRVdzw+33sWkIgxW7IUWfYe1pHKl91lwWtqSua0RCTo0YaP6b3BZNy2Ib0xI3pVJis2AGUu65jfya86nNnQpZRwOxpl9YCuOJ9RxSz+FnNvH+d+3PGacdzznrTuNde6Rot583oUmnpGkPaY5JiCHZ+U94KY63ecfEtsMUyx8RuzQBTmrTtIZQB3q9RDLDeiZNBnMFNfCr1zzX33YPl1x7Cz/91QtE12MwtYzoNS8cSqJYQnT5RGSiAjAmwNbyqCXArJ3bKpRviGdyi20mSIM0oxhNAKNtnqjf71HNbcFv2ciUrTn84A/y5XPXcvJRH2G6SHfoJnejqtn5JL27yPy8Xwj12Eq9I8DfCk00JQh1SNVdZxwRxQdPYXv46JOsCMMwCmqFCvjhL97gkqtu5Fvr7+KV12dYtnovopSMAth8PvoQ0nJYU9pEo0kkSCZBHYeBTV5vkq7bCI9xFTmhmQfTEGo+/jEla5xJUieJFaKRQdljdstG4nAz795vD75w9kmct/Y49t+jSIW0kPSHvaKfikW1Yq1gTBLQO8f4s+ctaRcUIyd+n7xadvbpuMsTIEQqX9FzPQxCNYrYwqAmtav94sVZvnn9LVx+/bd55pVNRDugnF7JsIqtXVPMA6vBpUy3ZHILeV6GNgTYzIholsDjv48FsGbe1qVwf4H6q9F6aYDoKVyau6HBM5rdTE8rPn/4oZx/+omcduyH6QtMmYb4UsRpG7lNnjHSuLZoyLkmt7CI0BHgNrFISNIQoM83poAiNDOcmzRIxOZjrhjmNCZhvUv55Z/+eoavX3oNt915P79+4VX6q/Yg2h41Fun1GPk6Ta2DVM3NVmXz/Qdh66rsJNJzx/m5hoya1IuMo0BI549WIJr1oxY0MF1Ctek1tHqDd+27mlOPOZwLzjmVDx24V4p0Uz0Oq1BY8vdmfA1sFfktToALRE47/XTcpQmwqQKHWIOxjLwSXZ8I/OTZOa65+S4uu+pmnn9tExQDev1lzI6GGHH0XB81wlyoMWUBGEJI4yWTG0xETG6IJ7aqOGn1KfN1cuPk9IRsg3EK0WQ7quQg0iQuIz1XIn6OetOrLC8CR3/yo/zJuWdy1OEHMV2kpW1hcrTrFWugMNLKK4xxaWmG5Jxf2i9oXkab7d+Rd/YJuNOxVU4l/YgT+VxBCDGNwixcQWwGxIujcZ/xQbMzNlQh5aEDaXn88+c2c8X1t3Lljbfy65c3E8ppvBtgXB+VJG6WPB96shKdKsBNYa+tfJDOhkZek/K8Zt5j6e8iJqVGmoVsNtBtdJHBR4gm5TN1iIlzTPct9ewmhpvfYL89VnHOGaey9qRjOPT9axJRjzzLnE09xlGbitsOl77blj3tXOziBAgh1qhYKsCL4Ze/2cjXLv8WV95wN69u8phyNdH2MEXJcDikX5Ro9ISqRpwluqR7IwLWpru99xCVorSEbGE+7tVsTrUUxbV3tix/aUmwWerGVIlLHnGKSgCpUElFE795E1M2cMJnP81ffPEsPnPw/kwbcJFUjXMOsYZRHSgK1whwCL5qOwOAtoIcs23XmAwbdAS4KHZIgLSOOEZMzoyk/GpKWxTJv1YzAeYI3FjLcDREipJgDEHgpy/McPFV67n65m/z/CubMeVypLcyrwigGaaVVr8uKwwcTe643WDJo6FEMc3xby7jRoGQo8k0uVWwtsAYQxUa/8Z089QoWBGQCtQTfY0z0LOC+BGzm15nr9XTnHLs4XzpvNM57P17UQLGJ8PyRs46seu2u6tl4udbAbsEAU7OS618ncYEGstsFbClZQg89etZLr7qeq5efwcvb5iF/nKknAJTUntBYzIlUFWsBKxRgnqCXbyXtfXB0/l2UNAM+7HpfismTxerEPXtdoYAQoE1JaGOaPT0S6HnAsPZ1wh+lhV9OOnIT3L+ulP5/Cffm071XCSxJNv5+TIWM265YjIu2N4SafzaSXQEmLENAtzGnxcPZRbbiVFbbagnGWrEXIh75tXAdTfewpXX38rjv3wJN7WKYjBNEEMVNR1OcWAMVhyx9ukmajTldQuHGgijEcY2TtzzF5lJAbEw07bwhk3bVommUQ5GTV7+e4zm1skwROe2sLJvOemoT/KnX1jHoR/ehz7pPC/ye1kSIVaVx1nJ18JiRZKuCvymMUl+o9GIstfDB5+MKnt9Hn3qVS664kauu/kuNtcw0h6VKVi2ag0zwyGxqnC9KQzJG62JlJSA2MnDYxhPeUv0ohLbu32LBXZGEQNGcRKIMRBDnaMFh4jBV4GeK+hZIdQzjDa9xKrpgrUnH8WfnHs6h35oD/rkxYuHXjsILkljxsUU2uhy0ilF3hT5pddNoiPAjN+VACfDmEV2YqOXCyFgXZpoV4c0vlQMDBVe2QjXrL+Xb15/C4//5Oe4ZSuRcopgeogtqYa551cEJ2TPx0DIw4NtUeYUDYxzf2kDm1bJdHlLNpMw+aY9uaWe8bCtIjluk278SOo6KSz0iIR6C4w2s3qZ4+jPHsY5px7HiYd/iD5QJw+Q1pNQaIa770jn1+kAfyvEnIweVSP+6n/7O/7jN25gC8twgxVQTDPUHtH2UNJSdmpZn2o0gw8jjE0tRT4YiC4pWmMFrSNzzEakiQBb0TIsIL7xEsM1olbV5J8qKU8XYyDWcwx6Duo56i0bWDPtWHvc5/nKF87m4A/sRgHEWhECRpMBgyG3NMU4MVFt23tjnHNccCIteWZ7s3gzbV9szYhbEWDWdTZylEYoE9NPk9eKMdKO8ajys1/aBNfcdAcXX3kDTz33Ct5OE0wakRpCwDiXUjEhgnNYV6bcXtB2eP14bvTEDZNxDjotPfPqYSFxt+2iiWxNHpeqjYhbA9Z40AqjFcSKUFcsY5ZTP/FevrzuZI488kimekJdR/qFIWQj1vENfJEbdeumvfPwlifA1rigHfgCVV2hqvz0Vy9y8fX3cMOdD/HsS28gU6sJ5XK89KjqmJamfgRaUZapna2uNVlH2eXE4TAl23KlFsYnzFYnybwD1eRdQGtPL7uL1CHN0rUSiWEO8bP4udfZa/WA804/ma+ct4537zdNEcEpEBRr0zjKpnARo8/5O1n8BFGYHMe4uNEmHQG+afyeBDiP/MbKhKbf2JAGzxtsWh+GgI8xrWqMMFvVFL0kzv/NBuWWex7i0mvX8/0f/RJvB0wtX83mkUdxuLJMkWRdkYZOFRN2awu+z8T2Gh3fKLc6r9tm6ZDPKYPFojGRoDEm5ZWNByqcVXpWqIdzMPMKa+oX+eKZJ/B//e//B3pFKo70ssXWvLnQkwTY6BTnqSZ2Dt7yBAjjpHPTTgYgCF7TAX3yl6/zNxdfzQ13P8RLc+CLZZhiihA1WUSFmPzgSPLWNEDckBLK89c8LQE2shYjC7R8+U4JmBjoGUddDamrQH9QMug5Nm98GeNn2W+vZZxxwmf50jmn8sEDVhNrT89Cz7h2bq4HQkhiV+ey5EY0bU+IWZBtFixZFxDg79xvCzv7BNz52MZ+2OYaeOLxeeS3MMWQenFRnyq8NnVpa1Q02qwjTX8ejWYxtkCKggp4eZPn5rsf5GvfvIHv/+hX+GIF5dRqostzSyzYwhIqj9jeeHMF5psmanvubn1jTysdkdwHrNmKK83yHF8XJEsxS6AwAeoZ5ja+wj67r+K4Iw/hL88/lU8cdADDucDUwKaPC6kJYf7+Xbj/uhzgb4VmtoFCcuswBiMmzfKwybvv4Z+9xjeuvolrbrubNzYOmV6zF6Ohp+hNE6pADIot05yOGDxuMMCHcRJ40TvlQgIUTW1JzR2zrlk5PQUamJvZgPgh++y+gtNPOZovn7OW9+1bJtfdGCmNYlBCnilRFv12WdISm6/T51jbRoCTB2h8zvy+HR4dASb8jgQ47/Vj6VPDP6ogBEQiqn48YAuHmHJ8Q/OB1M5Eyu1ZQ5A0ZGvDEG6//0dcfNV67rz/+8RyiqmVq6miUsXkWCPaDIvXfM5OEEt7Ux+T0PznRCR6rCTJTSuZaY1jA73CEuZmKagJsxtYM+U446Rj+MI5p/PxD66iT1qQ2PwSC1ko7rKj+OS+Got1drhb/0h4yxNgjJG6run1enjviSiFK/DB46wjamQ4HFL0BqhJ2qsfPvUy//lrl3HTnfez0RfYwUpi0aduFADWYiUPEcp3OqPZEHTiQlBI4ysZh/FW8t00eNCaorD40Ra02sJ799uds089hnPWHs/791mexh5VSmkkjbZVJrqd0klZV56iKBYdNuR9bBvxx399cyfNtq/fN7nkWyLYqhi0aJW3OS/i1hevbq0D0eaeKblqa8jOKRCiZlG1RVUp8znceBP64PEx4lwJkvo6Ngzhjvse5++vuJYHf/AzRqaHm16BSoHWKX8djGlVgG2RrLFPmzjmratR84VjmqUX05oVRHHGEGNNrCtKPFLNsmbKceLnP8U/+dK5HPL+1diQSK+qIkVpMALVsKLfLxFIfcKuKa7M37VvJTH0TibAiS6KbUY6tHeSRncVQmrCbkhQiMRQE0JImqfegJkIj/7kRf76kqu5+Z6H2Lhhjqm998dHofaptWc0qtOwHBnLXhImqmXt3TRiNVV7rXoIAQlD6moTH3zvfpx92omcderxvHM3RwGYkKK5nuvRNo7mLxaaPB4xN5anqq+IjEdyTqzOF5Xp7GDPdgT45rBDAlzQ6zoudGQHvTix/xYSIOkYJh/C9AfNNlvjAVPpJJMIYnK3SYzEmCRVdYgYV+CBDTU8+Nhz/O3Fl3P7fQ+DG9ArpwiSJg1WYoli51vqx0SAjRehwthiSwyiSUYjWqMkeZgjEKtZwnAL+6ya5uRjj+TL55zOIe9flcwSPBCqdL5ah+QbvDDWTKJhngh/TM4wSYtLKAKMWxU0Ui6rBhzDymB7NumRAWLAGUvT9q0aMJNGk+3c1zx0SDyisXXnrUyBlzy97cmX+a8XX8Wtdz/ISC29wUrqAGLK9FOyAFojiOCRtuJGXSeLlVAxcIY4uxEThlDP8b799+CfffkM1p7wOdasTMuaNFsjpKhPJ4nVjNONskBaswO8lYSjHWDhsrfFonKYrS+vhbNNoDlPts6XRYQ6JN/uIGBsHtNw9/f52qVX8dDjT7OpNkh/Gb6cotICTC8xsCugqjHqcUap61G25ErzarAFEm1OJwViGNI3gdmNL7Nm2nLWycfkWTKrmDIpqrMScFYIIZkAGyx1XaecnyRz1piLPE3uPkaIJrVshgBGlb7TlO5xRZ5GKFtdE1tJ0P4A2MkECMO5GTZuHrFqz93xJEurgU15CwlxopSe62shGZ2awqW7aOslFMEkCUly5k13wlEAbJIc3P3dX/Nfvn4533n4B9SUuN5ybLmc2bkKk51xfZ0GDEnh0nBsazDRQz0Lo82Y4RaO+OTHOGvtiZxxwifZvZ/6dEfVKLl8NMtzY9vCTdrRE7ti8qB2zLakMb78tlEskOS+Itl1uqqzw7ekc/qOh57i4mtv4a77H2XDKKLTu0ExRR2AOmCcJdYjCDXTy5YRY6SqA0WvR10FCucIozl6NjK3+WVWTTnOWXscX1x3Ch9/325IDdMFSFBCGGV3caWOFdYUEBuJT2oVTMJsk5f+BkQZjWq0N6DStLIpDcxunMFEz26rV47zo0uNAD2G12fhL//bf8nBn/os5553FnvtlnbQMqCem6FXpOHcrjAglugVbMpZBK3piU2DYqSp8iayEcBmV9wYY7Kzd5YZ4K6HnuZrl1/HnQ9+n5m6pJxagSt6jGKaEGatw1c1QqBPZLj5NfpaccShB3HuGSdwxsmfYlk2wu1B+xnOpj82SeB52NZe7giwwzxsnaLwMWCNbcOAECOCwUu6l1bAfY/8gouvvpF7HnycVzcPsVOrsL0Bla+RwmGtZVR7QpUS0eL6WPXYepZ69lX2WFFy+vFH8pXz1vLx9++N1VSbKfJpHLPfpDEgKDEk01RV017PIQSQIhU+VKmrWQprENdjjjRj5eUZuOmW+7jkoq/x1QvP5S/OOwmzVAlwhOHlIZx27j/j6WdfZq+913DeWSdxwbqTedeeyxgYiHUqiDZVddGUQ/PZXKBsBkhP5BmE1E5mJKYj50fgHD4IM9Fieo43PNz73V/ydxdfw8Pff5KZOc/y1bszqiOEZF6g1Sw9HXLkYR/mwnNP45RjP0pfgACiHmfSEHLn3MSAnZTHEeZ3sXQE2OHNYfEcrc/KgTSfOJ02VRj/rVkIPfazN/i7iy/n2/c8yKtbRlD0KaaXMTNXI8bS709R1ymiHG15nT2XC6cf/xkuPOc0PvGBPekBoZ5jUDg0RHwQrCvbFX9uI8aa3O0iKRdvJfUW11l5Y03KIlmT2v9e2ATX3HwXF135LX713G/Y/Prr/Md/97/y1bWfwuX898L+9bctAUIiwSGGF+bg7D/5P/PcKzP4GJnZ/CoH7LWatcd8jnPPPI2PfHB3ejmp7EcVU2WZ3FFSN1D2V0lVtbaiFDMVioL3tCVYY1BbMoJWQ7gFuOuBn/D3F13FQ999EmOXU408g9LwiYM/yP/uq+fwyYPfyZoBgMcScTiqCqw1YyF+4xZDY5m04MBtJaRtNAodA3ZYDIkI23kiKkleLYKPEWLESArPrE3qh6qORJsG3P/wF69w2XU3c8Mt9/HcyxuYWrkGSKLmzRtfZ8/dV3HacZ/lzy5Yy8ffsxsWmJsbMnBKaUjDslRB3FiOFdN147IyIWr2n5RUaIkxGbt6GY8offblIdfedAdXrr+LJ596FlzJihUr2PL6i/zV//wv+ct1h1NMEOAk6b2tCRBgKJZXI5x03r/kZ79+g2Wr98DXQwgjpJqjbyLrzjiRdacez2cO3ifZ3IfUvNEaBdiYqlck15VkXdt8u0Z7knZiXVdJ5GnSKCVswSgNT2AE3HjLD7jk0qsZ9Ke58ILzOPGYd1MCBRDqGQpRjElm+yK9tPyofDt1rYGSfPnsZCvb5Fef3A8dAXZYFLmlri325SXiIoQwGtU4mzwga18TpUCtYwQ8+fQMF33zGm67816efeZ5Dth/b44/5rOce9ZaDvvQ7gxIJFbVNf2iwBDxfpSiPGeasnH6oGZ4ezvEnUycIwaDHgHY7ME4+OWLkcuvu4Frb7yFX73wKkP6TK1cgy1KtB4xfONF/t3/9C/4y7M+TbkUI0BIBPjSCM748r/iR89soJhenaypiDgUh7Ll9VfZbdWAY4/4GF8+93Q+d9iBTAngs1ZYR6mXFjveWZGsZrdojPiYOzMLO5Y/kfzQSmuZGymunw7onE+eqCnVC+qVgRNMXluLJMGqj8lQtZiQkYRcuW6/cTuXYwJvWmDbYWljkYpwHqxk8gCtPEkYyLb1oU4T3cRQ1cpILaZMS9Cnfr2Je++9l0M++hE+efCBaVphXqKqDxRFyjHWVbJgawaKJXkWE+QraTmOYI1FgyLZUWkI/PjZikuvXc/V6+/guZdeY2rZCrwaTH8ZSMHc3BamCoGZN/i3/8N/yz9d94mdSoCLz7T7A2Lhl7I2hfkmV219yuCBMVQ+0N/zHcwMt3D9nY9w130PceTBH+ZPv3gWRx/5AQoBoz0sVZ6Tm3z3GB8rsAaT7dSaHGHajlx6DzDdS9VaH0csdwYlUseavhkgLlnGxJi2tZk366xpl70hJ6mtsfNO0q09+Tp0+O2RjHoXjkwlx2BCnUcfiEibpCudUAjMRdCgHHTgCj584GnJS7JWpor8agEK20pVizK11oWmL9gmAXjUmK5T41DjCKrU0VDk5ff3f/QCX7vyem6482Fe2lxTTK1ias3+jOoa4wqGtQIVxpQ4J2yamUlEvp3v/ccY0v5HJcCm0hNzA/d4hSrzJCMAwSfPvWEVKIoBRc8wque45cHHueeRH/LZIz7BWWtP4NRjPkrflPRMmmI1VbjU5pbvYJoD9YYAm10qCBhJprb5gVTViqhG+uJA67RCNSYngU07d2Oc8WNe1Cc7Ir0u4uvwpjDupbULXIGaYVc5VEjk1+S8SSSYxNhKTwyFS/m6dHOG0oJq6oUPIZm3TqakY0xpHGssmjWPAUM0gpdEnEGgErjvsee56vqbWX/b7by2ucKu2JPBqjXUFFQh0WrwOs7vkdpaTVES4lgc3gQLk33/4+/6h8MfPQLcCqlq0QrlsilVfkDAGrwE6loppGTZmv0IvubmB5/gtvsf5dMfeS8XrDuV00/5FMtLw5YIooYpDKNqlrIsUa2x0owMzx8nyYRAbNMWlETWQkOU4x2vE8sRWfCzQ4ediXnnoeSzV7IbCxFLmjSYiihpdWTE5Pw5eZygZPv8NOw8DUa3bYveXO7D96TGAg888MjPueib13DnA48zDIIppumv2YsRKdpT8TnCy8SW26o0D4FKV/hObELL2PkECCT32XR3ihhs9iMTFDGCcZYglrpWNlZgTB9ZNo2zwt0/fI4HHv+PXHzdhzjz5KM598wj2L1nmPGR6XIKiydqSNkShbr2FMYirsQ4COm0yDXkxsBo8d6L2FJzZyTQ4R8JW/XiLcTCc2zBuTfhqqIAEtF565OU/xFMaruUfPfP72MthDgkxjS3GDTNPMlT5aoA1jlq4I0RPPzYM/z9N67gzvsewhfTMFhD6BcYlDmfZDlYi3OuXelBno9NxGi2EdbUUrqzA4mdSoBCphxVkNgKmJtzoqk2hbpKlahygJgSP1IIgdpZBrvvj/E1Dzz+NA8/9iSXXnkNa4//DF84+xQO3L2Hw6FRMVpTuoKy7KHeE+s6dZNIoxyEpmcxdZjARDjYPp6W1dvPXXTo8MeGtub246Xz+BzNEur2DykiTPOFIxiLy0vQqq7SSqhI84+DhY0V3HL3D/j7S67iuz/8GVIuo1y1P6hlZIvkqKQRTAnOItgk1fE+u5pr60gjpMXepEPNzsTOjwDTbSv9Lh7FpgPTeKbVEXr9tHNDIFR58HMvzSat6ppQVQxW74mJNT96fiPf/48Xc9VNt3Puacdz9inH8O79VzCgYG40xCD0emnsc7KtDKQ0rs15BzsOzMcKGmByeT7GjjybO3R4U1gQCW5bLLDAZ098Jj+z4DUmr2ya9FIzRc6BBlRtvtYKAiHp+VQxtgfGMlJ4fQauv/0BLr76Rh5+/GeU07thV+2HRkulMDeqKaaKNKzdWFQchBS0YCyUFo0jJPpW7SUa8zxrR5CdTz87dQuUJEaO+WAJJtv1MD7iRZHuLt4nCypxGJt7fkcVzvUoBwOGowpQ+v1VrJpexUszQ/7X//D3XHbNes5Zezxnnnwc73vXaqYklevrOlBITO617aS3HAGa1hkoVZQnTs6FNo+dUUGHnQ0hYvI1tHhWrVlrpSc0kwSbm3lUi0+BICrwzEsjrr/lbq699Q6e+OXz1LZPudv+qC2Zq5TgA846yqkePtSgipoJg+G8lKauF704FJNnIO/8LOBOp+AoEAxESQPJIxbUIKrZSshDrMAGpDSI98RcVbJFga/qVKQVRX0giDBbB7yHle/4ML/evIV/+5+v4e+uvpOzTz2a804/lo+9bw3LCkuBhZiGwoSsnWns3ZK63TfeM2ljVVJSV0yOBjt0+D3RzvDYfj5561RhPgPzA8nZef7ArPS63MWh6VJXzTKabMcmQXBWqCz8/MXAdbfcw2XX3cyPnn6OYmolbnpP6hBTESVb8VM4oisY1h5rkhegM4agmmaXxCSTVvWkbJKhccBGIErKpi/mivPHxk4nQBgLzRWz4ERochdZwRcVVdNm4Hz2BTQWok+DSlUVcSW26LNpqEhvJXaPKTaEiv9y2U1c9a2bOPXYw/nS2adx2EH70xfX2oWTzwsjkIbMBJAFJ+YOD9pCu8ftOSYt5vTXocObXVVMrEeUlEefOKea7pGY+3NbMwURVGxK71j46a82cc1Nd3DVTbfz1PNvIFMr6e92IBVK5XPbFSAmacA0KtFHxBYgiq+GqTWvyDOMk78VYgWJkSipewTVvMqL7dS8nY0/IgGO+/waLWCrVW7SGZAOokha7k6KTjQNATfZzTY29ldIMhjN81FjPhFQxdhIDOkupWIpl+/GllDztevu56Z7n+SQg97N//HPL+DjB+3DcpdeZmMa62fEEmufh54qUQUV2qVG0BorBfORk8EyXhiPk9PNXmgeGVt8dSS41DFZoFjEZ3AC7Z8nChoiLhmtxjAW3/uAWAs+JNNfTX28FaAupYF+8dwWLrr8Vm65+0F++fQzDFbuxmDNvswGJcRIOxskLvjsXB3U6AkYjCkQM+mWbdJYzaYaLVlKliVvkp3Vm8ebUSTN739M7PQIsMmvpTlAjR1qw4hpTGUUQzOxSrQJnXOvZNuRmBCzfbnJM32JEVsU1HWkrqHXm6JcOeCV2RF3PPJTHvzuv+K4zxzCOaedzNFHHMTqMrXDGR/olT2UgIgZO9428gGx1L6m54r5ecAFB3Ai+7IIzHYf7fA2R8sY+TyYPHdk4XMm/iyT7e4y0UzgCL7COpfblhSsIwSoBXDJTPV7P3qJK29Yzw233sPLWwQtp1m2z7sIKmyaG2GKAmdLquEQ8sTDMXKtOW9AlPSfnRgnuzgWJowiC922dwZ2OgH+o6CdazohoWnvLkJoIk9jqH3qz3Vln8KUxFhw0z0/5Na7v8thH/kQf/7FcznhqA+xsm+pgXpYsaxfQAyoKtamSBSx9JwhzKsUpxPZ0CxFUkQ4Jsex9X+TgN7ZMxE67GQ0lvoTEdBkMaM5t8bn1MTT84BhmXgDa10udMSUo+tZqhzxPfz48/zDJddy5wPfY8tswA6W0xv0GMXI3OwIzaXaGCNRIqZX0kr5FhKYxB3mLXcFvD0IcDtwzlFXFSgURVoqpJ5Gg5qSGkMsLcYJD//4Gb73f/t/ctiH38/5ZxzP2hOPYPfBgCEQfLLiSvKmQFlaqlHEuGQ/pBN3ZY/BYJIEQcPEbbtZBiyULHRYspgUvi6CsRw1F+Da9agSY543LLlUF6H2INagzqA9xwbg9gd+xBXXref2+77LlpEwvXwvmCrwYpLBqURUIiKWokzXiPd5gphbGAFObvuun8Pe9QlQYVHnjPbAGMhN10E15SGtJaoynJ1FigIxJUOx0HfEUHHP40/zyBP/P/7LxVdzwboTWXfS0RywZioRYVVTpKQGTgJGFY2CWNNm9cZbY1InS2S8Zsjb1UV9HealuNsTohGFQepKmrSkElQkz73xKJo8MesKlYIaQyUG6+DlWbjnoR/ztSuu4rtP/pTNs55iajV2RZ/ZWBCjpSgK5qoZirJAjMGHQKgqMA6xBba0iQgn0J7Gb5MTeBcnwB3HUdF7xDqcSa4aOtGqI0UBxlENh8SoOaTvM1i9HKM1P3lhA//3//BfufJbt3HmSUdx/hmn8I7diuSx6lMeMJlUxiSREdr7dLN1Sk4eqozFoLx1pmJ12Lloda/AJPm158WkhZGkntpmtKWKMhx5iqKHR4gWNldwy7d/wtevuJaHfvA4teljB6spV/WJajGuQH1E60AVPeWgn8bNxkhRlmAcdV2j3uPf2hNz/1GwixPgm4AxaAjUURBjsb00VrOuPClHN4vp9VI0r8nOam6mQoyl6O+B66/mJ7+Z5Yf/8VIuumI9F5xxAueffiLv23dAiAZHQ2YRiblX2Girv4oIIZ/NTYkn3cubWPHtfwg6bAtN81psOS7l/5LaQVTmDdNq+Cjp+AApMH2hBl6Zhetve4C/ueQ6nvjFc5j+SnxvL4pyQIgQ6/RaX40QC3bgCCHpZQWLETO+JqxNltAdAb4NoIrJjdkaQlsQwblUQcOjqmmQs82N48ZSln1GdYU3BcoUg1VT/GbzJv7qr7/BJVdcw9mnncAXzlrL+9+5MlkS5WbzpsHbaEh5FbVbawnnYdfPo3T43SHZss2MO+NTAa/t1sinz4SOLyLEPCTpxQ3wrVvv5qIrruXxnz+DmV5NuWpvRt6CLamqSK83wMcK1YAtCkKsiRpImgaXZlVagymKtktEgUk39bcr/mgEuNAN+o9hdpg/KLtSpFxg2pa8TS35RKxziCo2T7TSeogTEDF4NYyC4srl2JUFv5mb5a+vup2Lr7uD0074PBecdRqHHbR7GhqtDhdGOFFEA9YkcWjwYItGja+ImPbnQmxrTGCHtwfGpsBm4lzI+tgAxgohBpCU46uDxyOITbbzEXj6uSFX3HAb37zxNn7z2maGCuWadxAoGIYkWpYYkuNzNZvEZQIaQ25DAyRVjMkTDFPFd+I6XeT8i21V+u2Bt38EuD2omahkTco8Nd2VFSJp7ZAE1w7sFLEoCaKICXztmtv41i13ceKxR/CnX1jHoR/Yk2mXTtQwnMXZiBOwrpioAmch+ISRqjYFmkWswTvs2ljM3HPs6tzYVQGaupDqUXJRMdZQ+RpciVfBA794YYZLr7iB69ffyfOvbqFYuSdb7AqMKwi2T117NIakYZ7w2SR/WkLuuPp976//GO+xk7GECHBhpXjCTy390jrdIqntLqmzQ55zmpYmUQ3YHh7LkMhgdcnmuTf45o13c/Od3+HoIw7ji2efypGHvJcV/Sl8gLqaY2DTNoRRhe31EQHvK1yWGTQuuAsvki4K3PWx0O18cupZDIrJrWYhVFhnsC4QIgTt4SVZ1P/o6Y18/YrrufnuB3n+lY3Utk9/9T5sCRaMEqUghqQLdDYV5OowShPjZNwTjGZj4KbpHdq+4N8GcaFwexfFEiLAxdH4kilp0EwQskV3dqgWae+kUWOu6hpQh48Voyi4/kr6g+XM1bPcdO9j3Hrvwxz+sQ/xlQvWsfboD1P0B9QCEgRbTuXPA+vKeSQ3aQsOf8Q0QYc/OLZ1LI3Ny1ED2DQ8fJT9+CLw+FOv87cXXcGtdz/MhllP7E9jV+yNGkdlbCK0mHKGIZsWiDFZr2IIzecu0BuaJsdo3ho9uTsLS5gAG7nBuIMktGYMLrfURaKkmpzEtEQ1ErOnWYVmxXzlPaZwRPpQ9AkSuP37v+D+x/7fHPKB9/CXXzmfEz77IZYXggN8gOhHOCOUzqE+TcVaOPsBWDRy6LDrYOGc28mbm6oiRvHBJ/N6k0ZZegvf+f5zXHLVTdz9nR+ycXNEyjXotEvjXFWoqiGEgC36WLG5oGeIIVBHUhW36JGm82Y0+btI9sBMPBknWtLM/FR9m/PbGm+PNM0SJkCAOBag6vhv83IkOepL7hXNpJCIoGCUsugzrBQxDmxBiAE1gl0+wIeK+5/8Fd/9V/+Gg961L1+54CzOOOlIdl8G1vbSQCaBKNoWPpqRh5MXS4NuSfzWxZst8k0O/YkqzI0CUpRgYBa45Z4fcMX1t3LXQ0+wcUaZXrkvo77BmhKxBSoepcbYAtfrEeuIqieqRQxI7nZKlb7cojQpz8+E9/agr98fS4YAzSI5jki+w+VzxOBzzrhZiprkU6ianTE0KbY0YLK+eW7LDHYwICho8GAg+kAUwdgebuVeEGuefH4T//qv/jPfuOI6zl13EmeceAz77mYYeugZwTk3rxAiIm3LXoddF9u7YfkoSM+xqYbb7/4Rl155PQ/84EdsrsH2VrB8zxVpnGQPotSgAT8agVhc2aMaeQqT7KViTLJ7Y0yynvIBQsAg81J1KjEZGEw4ac07wxZsrsk58nEk+PY6H9/yBLgwBJ9vl9N0WYx7JE1jh5WePX6f1nY/JX8n240aL7XWXFJjOolIMZ82JCnjO6pqquLaZdOpQDKaBWdwtsRrOkljhEoMhbHQs9R1wZPPvMrjf/XXfOPy6znrtOP4s7NOYc2KgoFLnxaDUmYXon+saK9zrf7tsPX+2trjcfK5mn0j5xVcY9bQyfh5tQpBLJg0jG0ocOX1j/CNq27gh0/+DIppbH8NzhnoDdg8HIERjLVpLo61FIOS4JOExhhHHUa4okiemDEQfVY8k3KBEjRvQiLKbaGRZI+/5ThFpNI8tm3y25Zk+q1gero9vDUIMB8Y2daV2hyAeXs5HwzjktGZj0wtm2I0nCXUNWXpqOpZnDOpXc0N0mi+qOAKHJFqbi5Z7mcSbQUxjWGBaZa+NGd62lZJZgZBBUKdXlWk/F3qnbTjbVUlkkZuUiwnuD5il/PTF2v+6m+u5bKrbubcM4/ngjPO4F37OJwTZr3Ss4ozWbeliopJc1rz6sYIaIztSW1M9vHP1mAhKtZK20m1VevdRE/n9vT+iyjBFvx7J0cECzd+4Q1zOy+VRZ7QtKaNv1VEU3YYaCr1Fo15T5hIyI9bkQk7SE0sZwuCQm1S1FUDL22Gm257kMuuvonv/vhppFxGMb034gqqCF6A2mfnoVQpxiTNXh00F+SS2zKul6Z9ZBPSpthhCDRGHI1L9Hxvypifly86I+k6cv2UlhluouwXVMMR/cE0oyqAzdXs4FN+UUmfsa2duws0krw1CJD5+0ulXZUu8iQzL2lLSMaP1jhmZ2YAKHt9YvBMDZZRj+ZwRY/K5zcueuADVRjSG5TUPktbADDzI85sszX2PmuiRDPu7lhgxrAwidwmwDEgkqzBC1CniNY8v2WWf/83V3Dx5es5/aSj+PJ5p/Phd60kINQRtPZM9UqiRqq6oleU4/GvxrSC6qQtrFtJhTUmaw23QVBvEzujPxTG6tCmSyNpQ2PUVKwyrSk9FkPQSAwRYxwmuzBjCgKJ0ALw/Btw3a33cMlVN/KTXzyLGazELt+baBxemuxyzJ+kbUSZPmfhsZqg6clzTuc/UycurDbdM34UYwy+qsA56A+gSm15Rb9PVc0xNdVndssM2B5YA7HG9ApiLtzNC3p/lx29k/GWIcDfDbljIobEVRKx5RRqDH6uxo8CYntItKSRfQUYCzILRIghDURqyKz92dzeI1YjNie4G5+2MGkDtOBO1yrlJ/s2Gf/eEKIYg1IQzQoGq1awoZ7lry9Zz9U33MHJxx3JhWedyic+tg/9XsnQK7EeMug7wKdCCwZrCrxCDELpSK18hDSLQQ3GJM2XgXmDndqkuKRe5N9ulbLgonurLHG2sR073LxFnjAZLQsWjT5F3dZibYqZqpBn3IpijcMQqImIEYbeM1d7BoMpauD5VyuuufF2LrtmPT9/9hVMbznF8j3RooeqRTR1/2gT2ss4b7f9ckXEtNPedOI8lPbRrYxWF9g7a4g4V+Dz9UAk5Q5dMkWdna2QssSW09kia4TQw2iFBsGYYjuV4rc+dm0CVNIdT1LvRlGW1HVNANz0FDaCizHl4mIkVKO0BnFCURTUdYWYkvHddKICTMSmDCCQ8oIGCDKOFhPe3H2vIcLJKmAqdAh1iFgGLNvzQDZXM3z9mlu54/5HOPITH+Er553BYR97B1ODKeZCWmlb4xD1RI0YMZiJo6i57W8c+cXkT7MLn6R/eKR9ttiKLbXDuvFyWZIVlbPNjcUSgyLOgrNs9p7ClRhX8sTzG7jsm9/i1rsf5JfPvpzGSi7bgxqHuD6jUZ3SFItUkFsVwHaPW0Qmb7Ay/jbzc3YT52qzYpHU6RRqT28wSPb2oyFgcYXFhBHez1IUJVhLvWVTcocuS0I9Au/p9QbUu7iMcNcmQJqTBUKMaSq9y7m4za/TWz7F7OuvMNWfZrqcItoelbFU1RCvgcFgilFdZTcO2y4JBZurZ6k7JEpMD2kjLHVZUd9gkQpzU3NpIr5FeDLd8dPErGhgU1Wh0dLf/QBeqWouvflB1t/7fY781Mf5ynnr+PynD2AZKbXUk9TEXvlZil7Z9hSLuLaHM8aIMaB4tvIh1AWH/rclyLcKob7p7RgfAFUdv0xyUat9dF55LOX0TPocjYB4jFGEmjpGnPQJCF5z/tA5Hvzpb7jy2lu44ba7eW3DCIoBLNuHkS3TPFwgagTnsnB5/pfQqMQmz7G9762GMCFpab+hLIjStyoQKqIpBdLrL2c0O8LYSOEEqx4nFaHejNQjCjfNsB5ip1cQ6pBmjBQlVgrCKILbtdMouzYBCkQP1lpsaQm+bjq66Q0M/91/8yesLoRLL7qIJ3/yC6ZWrsGYAT1bUpZTbN60ATtIRRAbc5VMJ/N2JAPK5O1MY2k/Xv7qds9RGEd+ZpE+0JgrvQrEegQGXH+KudEQ1LFi7/cwO7eZG+75Pvc+9BifOeQgLlx3Cid89kOsKkHFYntTIODzQBwjJlcgY5bQTBotkTNaExHhLpCo/sfEwmhrTH5x3nKzjZ0MeSx1pOgZVC0+1DhrqaOiTvAuDRv67uPPctk167n9vod4eeMoVXVXrqaODnEFagx1HdJSMxcs1IwtrxbKoCTncbcJAdTlc3P7KxGTI792Bk/+e6gjy6YGBD+DCXP0xLPx1RfZb99VXHDBhRx06Cf5H//nf8tvXp/BlNNgCohKPRpR9gaJyHfJ7F/Crk2AAIUQok/Dha1NQmI/i442cfB79+aYg9dw6hH/C7fcfhdXfus2fvjzF6lkgLKCVatWMDOq0sqm0fhJSLIGtWkYEyWNoWlCU1eNvxV5hHkuG+NfVVP1TnolZBmDKwpEhM2zQ5wbUK7o4wncfN8Puf3eRzjy0I/yxbPWctrxH2NZH3pAHSziA1OlBdKwapFG4tNUAptq4ES8kyOgHRZEtlFd3dmB4Da3o31g/sU5fl7MYwwmK6Mp6jL59WNZi2JMRDHM1T7dEK2lNo5Z4N5Hfso3r7qBu+9/lDdmk4aPwWq8Lah9oCj7jOoheA9lmfLQMU9r82EefTQT0ra5X+cV6Sby1813bucMp+9odOuXNZ8jClYC9dwc+I0QNrH/fnvx3335zzj9tGM5YJ9lbARKnUNCINYlSADbR9wgrZryvtxVsesToAYoHFSpHhdrT9+Ao2a04SUGrOHA3Uv+2RdO4rx1J3HNtx/mG1fexI9+8jRbZgz96VRxjVlTGNWOk7qNH9o8okvkZ3JJJLY+botjYSN868XVRIS5mqvBIxrzXV9RjdiiR4iRWHuCEaZ33w8bR9z3w6d4+PF/z998Yz/+4ktnccxnP8Heq8FimQ1KCThr0MZWq/F4m5cjGhu5LmWkqmuOikk9stJUYCUXyqxBRVIxqihQgVeH8J1Hfso3v/UtHnz0CV55fYZlq/eit2qKUTAEU6aovCypRrNQOop+ST2sQITC9amrqq3at8jL4TTmMo7Pk8XQqBKa1fI8sp9/lJu/tcIojVj14GcQP+Td+67kKxd8mfNP+zy7L0vP9kDYAqUGetYwVE0pJgFTOPyo6pbAOxf5Nh19FsYpiIUoWLVMuZIeEEZbsNayR3+Kr6z9FKef+inW3/ool135LX74o59TB4ebWkY0Bd4USd/nQ5LMhADGYmJS4ltriTEQYjZQ3dEWLlxyTZzQRkknZEhVWxWT9GVZrhBCIsLUMK+MfJ16hlfsRh0jj7/wBv+H/+Xfc9C7D+D8s0/h7LVHs+9yYRRLgq8p82BqrEViRGPEWItgqELSmTWxQRMVtvXHt5MZ5sRNJ+ZlpjEm53hDWxh32Z5Ma4UCiAE1gaCR2liisWz0cOvdP+aiS7/F9574GRURLQbYlavYoiUh5O5yjZjCEMMoDZpWT6h8qsyrJXrF2ILUbx633ta8vVt9D53/d9E6HbOtlsuCcQ5iUkmgihGwmdSjHxGrGT5+0Ls474wTWXvcEey7In1tHaXqdlkKU4D1FvGKLSxB03uFfKm91YXOO8IuToAkDgR0njdZTvjnxwa2QJwQ4oiht+xWOi48+VDWnXQod9zzU/7rxZfzne/+EO0tQ3orKYopZNCn8jViwJDcdKN6CIrYJEHRiaHRvyva/I/EPPN4ctma8kExE6VmegrY9DwrDFbvwxPPvspP/v1/4fJrvsU5px7HulOP5V179qlIVePgawyKy87YEOlZk7+TjAmhiXYnclG7OmKYf6OadNypKo8rDEaE2td4Iq4siFZQjahxqJR44LW5yC13PcglV9/Cw489xWxV0lu2mmAFtS4RECYVL1STRVUMNMsH0+hHtUlDpLDtt5KQLHI8yrKkqkYQ8g3MGFxZ4H0kjiogMtXvgx+h9RCtZgn1LIcd8lEuPHctJx31UfYYJCKQqBRRKApQTXYJoqkdTjROnA/5pjH+v10WuzYBqmCwiNp0/MUAIedRUsVWgRgNVg1WYEXpqElFg5ViWXfUBzjtqP+JW+79KV+77Fq+98Qv2bRpBun1WT61nNnhEBVwhUNMnypEYpRURc35s8ZSaxJv9sROxJf/0YQik1U8k26zSkjRIIKIZumPMBeU3so1SDXiiV+9zk/+t4v5+uU3ccapx3DButN41359Bq4gdRHUlBKREEBdysObcixAB0JIujJnd3RqLLbEeuvBmPTdffO9nEvfV6AoDFJHMJaigBg9lQYqI0TpocArm+GqG+7mquu+zQ9/9jTBlAxW7Mm0dXgsGnIeUUFFU44s3ZHbdriUUyyaLUrFFjORS/49MBrWKZ/YGyBGiFWVl9mKsdAvShhuwvpZXBhy0PsP4M8uPJdTjj+UKZcivoJ0kxeJ6bjHNO86Yog2FWowknSPpAq1yXlj1Td/rr8VsWsTIAaJBhGbyvqkkDxEJaq2VT1rS/BD0BTGWyOElOGgChbBsvZzH+DUz/333P/Yy/z1P1zCfQ89ypZXX2d62QpGQVAKgoloMFlUbdPy+PdAOnFMqtimvyRbfk225RHTzmhACtCQMlbtsHVFjE2OItEyWLknhUR+8err/KeLruOaW+/lpKOO4Nx1J3PIe1ZS2oJhFSnzPFhiyoCFPNHOoGk0QHpr6qpKk8K23vLf63v/MRG8xxauJT4F6hiIMdKzLg0AqjzeB0YCdVEixvD060PWf/t+LrnyZp56+iXUDJje410MK89sDBgjhBjTcrYpiMn8djREMPPkRmZcsW1udr9nFb4oCrzalDOsa7BpxU0IOALD116k1IrPfvLjfPXCs/n8p9/H8h7YkIpnaYs8MYywYvK6NgC23V9ZwUo042hW2kh218YuToCQIpDJW1DqcIgSiQRqSOYCpkyEFZNuEK9YCwOXXFeqYSAay1Ef25PP/4d/wSNPvsA/XHo1N97+AD4UxHIaKQt8jDlJEhAr22yFfFO+apKreJKiD6hxeGyMuTppUC0Ikg+TOEBRUTQGwKBRKV0fiIy8p7bQ2213Ygj86vUZ/vbqb3PV+ns49ohD+YsvncOnPrgKjyH4kHqNGZccmyV2Y9NV9MqttGRvBczLq8rCS3B+RNrMYUmvSzdIYywmDxUieigcwSU5yS9/E7jk2m9xxY238sxLb2DcNHZ6Nbgem6oqtcEFQz2sGUxPU1UVRhohjeSUjICkG3Nq7F1kW8XnfWu2OlfmYXvyFjVJzO9cIjKJ9EwkjuaoZzdR9gzHffbj/NkFZ3PUp99JT2Ag6aKPcRYTNZ9TSVyfND8CLkV2gdS/HGWUF71JMJ5SMlk32xD5LopdngBVkqpdieM77wQM4EOkIE2+aq53ay1Yi9YpQhikaUj4WBNwHP7hfTn43/zv+fMvf4lvXPUtrl1/J69t2MiylbtD2WPo61QfkfFgzN8JbWMv6YRSs0hiOZKHPKTlsMa2+IdYah9TR4EI9bCmNgbT6+GmdiNUIzYFzxU338vN376LEz/3Kf70i2fzmUP2o8Cmdi8F0Yg146V3iGFRg9Z529Tu4T8ettbx7Rje++TXaNJFXQWfdXYWUziGEX7xzAxX33AHl117C8+8sgk7vYpyxX5ETWMnRSV1S6jS6/UQ45ibG+Gc0DgGCYKq5PpRzvctrMwKjPuwf98EcsRZk4ZvSU2sZqhmN7JqusfxJx3Jl845gyMO2Y+e5HGsNSCeKB5jJKdXevm9UoogBkmtns2xb5YDElLOWATEZvuHhd9v18OuTYASiKZOJBKgcUJBDRKlDbskre+ASECSKwySpr4VaT5qqsjlG2CubQyAwz6wmsP+9Vf4s/PX8rXLr+HGOx7gxddexvZXYIoBikvFkIyFl2crhM4PzDdbiLT2LgBq8TJ54jG+iPIJmuyJGlEzGOfwlcfXEVc43GAFSOqKqUZz6YONYXq3PQl+xHX3fI+bv/M9jvjkYfzpBadz/CffyQqXLtiqChTOUJg0nUxV52nh0v55i53wTYth/qcsIEhFiJEcpQnVKLl3u+zMcv9Tr3HJ1eu5/a4HefHVDaidYmr1Pngsc1VARds0QC3JeGNUVeAjxdQUoR4mPbzmwlTrJRlTB474tslOlGSsENPzophF88fpyTvez6JKKRDrGcJwE6sGhpNOPYo/u2Adh3x4bwYmLXWtgMaIdQbEEXFUMSQn6SYNaR0NJ4okCUytjVIhYuZF3dD2GO/ia+BdmwAbTJoTtBj/W0xKRmtTSbWCwZG9ntsqYcp85OMalJ4V6qjUtfLx9+3GR//1X/DVPzmfy66+iZtuu5Nfv/gbesv3wEsB6iYKLzmjJyaJqhkT4FZuMZqb4NVk4XXTd8W4jN02sAc05wABRGxqUDfJmkRV8N4TQ0BEcf0+VoTRcIa5IPhaMOUKbOG49cEnuPv+hznh0+/jK2efxlGfP4yp0lLHpMqxCjbLb9JnN50y8yNUmYhidMExWDQ3Pk+3ttgTFjxv4p8wll3MGxguzb/NxO8T22EMHsEHkJ6jAu773lNcccNtXHPHI2wYKc5NY6b2QE3BSBzeR4wtsNamZa4VjDhilj9FY6mHs+n3fGa1X1oakkv7Jko67s22NZu33aVvvvGJSNvdRFswS/8u8YSNr7P78oLTzjqeL5xzOh99/yoGkoivjBFjSQfU2JQaImuscgrA5bHV2ggAYpMmgF6TGlms2vs2cRPaxQlQQG3yMUPynXhsNuAkNazH6DE23cUNIXf5Svu/BhZa3bPNAlVnwJaGUZ1Oxo+/Yzkf+xcX8BdnHcO1t9zF1666hede24KaAVMrd2euEqIrqGNaYokIGmuiavueRMVgiYQcIcbx+TUp55msDsccSeb/l5ydNtDqyELzXJccRnyEoB5xPUJMDtVGhFojxfQqRJdx23ef59sP/X/41Mc+xJ+cezonfO6j7DkNBktVeayAdY6okRA1dS+geO8pXIHic3aooQCT26NMirBZ8F0W/lwY6TQX1QQ5RMCrYkw6XlEjRJ81melGIJJ0kyKSZzun19a1JxoLFkYK3/nus3zjqhu44/5HeGOmprdiT+y0QzHUmURjVDC5sBBTf3lsid5mPkrLxFTUsMy7AbfRkkHopaUjIKJASPleSTdi9TGlGtQQUIwZO4M7MbkgEXEm4qsR/dLi6xHDLZvYY81yzj77s/zpuafx4Q/sn3ZnKuUn2VRDYM3gJVVM9rKcdLyZ5LamLbmRgUYBFZuJ0yHR5veLbbFkV8YuToAkOYoqJlNayEvEVsm/AKIgEtFFDpxM/GxO4UZc2nOCEWGU+40POnBP3vNPz+esM9dx7S13cfW3buEnv3oOU66gV6xKhQRSkhprMM6hIRK9T2JkZ1Dvk28f0lqaz19TNC13k12qTf3Xtc8xi0gR2n9nq5jGbDZqKp6gEChwU7ujvuI7T/yaB3/w/+Lg976DPz3/dE4/8TPsscxhgLmhxzmDc47a1xjnKFxBVVcUVrDGZuVHJCsm8jbmssrkV5rcyfPGMY41cgufZ4BiUgMnhiCWOiilNYRYo8bhbFru1jko9oCWjpkAdz/wFF+/7FruePgxZr2jN70b5eoBdYhJQ9pG7UxEc/P97hbFpGRpwlW8gWoOr4SUKhEDJuTjJrhej9HsCCRinSOEurXd8nMziDH0SkNPFMIczNYcuOcqjl+3ji+fcxofO3A6if1jIIRA6dJqQERpvIwWBtvzTpX2wfmrqNacWPN2k76HYCbysKn7alfGrk+AfyA0J4k1NkU/IWCcwVpLCGkp6hDet1fJ/+krJ3LBGSdy9Y23ccX1N/P4z3+FcVOUy9ekZLI1aWarWHqDHnUMVPUIEGwcL9/SxZESzuSljlFFchN7onmb5pQ0uUJqtpdMX9iK1y5fc0Q5Gs6yfNlyQr9PrFfx6LMbefjf/Cf+5ps3cvapR/OF047jHbs7HBDqEc6AxJoQs8aM1OJauJRrSg7ZEawj+gCuP68FLy3wG+LT+XyxcJWV3h1QJCqEgGIxja8jEGKNRZEwSrZUYnFFSQVsqOCabz/K9bfdx50Pfp+hh6kVe1KKw1eJKVUaz5+Uqli4jN8uJqr4xKa9LDCeNJjdgxoCSVuKiab9qqORpxwsQ/BU1RBRnxynVbA9wYil9kOGMxs4cI8VfOG04/niGcfxof2nKYBhHRnl2dVlUaYQIMbWfGN7WYYOHQHuECGGFOG4dCoZMfR7/fSgKuIFiXDAKvjnXzqBC84+gSu+dSeXX3sTP/r5c6B93GA5pphmrg6MfIUpe0jRS8WKOoyZweScnzFJnhE1pwHHUcTYiQbejI6sWU61HNjk0PIF0puaZstwiFYeipJi2R70Bit48tnX+dl/uoiLv3kFXzjzJC4882QO3Hu6jfRUIsNhhTEFvSLJgerRiLLvEinEFNEsxDjO0Im/ybZpR3MMIwLO5RtBSlSliCfl6Mqyj+Rl7mub4ea7HuS/Xn4d3//pM4zMADPYjX5vijqkm5nL2zaZAtluz/aimF/hbQRZJld4o0Ss2GyEkVMdcb5syxYF1cxmsEqvl0TIVgJWIlU9JFQj3r3fXpz5pS/xpXUn8O7dk36vVAjDOfr9PlGK9v2892313gdPsUNB+9JGt3d2gNbPL+eeVJUQkkSkMCn35ELKcNejmj37Bf/8/GP4wtpjuOv+H/L1S67mkR/+hBE9Vu+2N5tGgboaQq8HMeVpxDZJ9JygjobUOZAM2RuuUHK+Tzy0TsDb2vD8mgnrLRgvchKfWkZ1BFNglw0wGqlHc6m62J/Gh4Jnt9T8P/76Si667tuct/YEzj3teN57wApKY+j1U7FE64gaoZwajGktVx6NGRdKbKa/5Lsj7daY7Mg4HsTTbKQBKZLdPOQigGIIqUijkdnZGjs1xQzwqxdH3HrPd7nyxm/z6I9/zkgtUyv3poiK2gLvPb72IJYq+ylaNy5WzGv9a1ratoquJ/+tJOedvGfVpkmCzaMKxIiRiJrYzpeJYlIHkwiBGrusIFQ1MXr6zrL59VfoxREfeMde/NkFF7DulGPYZxXEAIVC9BUjoBwMkoggpBkhIkLhEhlGjUmHmKpj2z9XljA6AtwBpKke5+KKycOJoEkg5xNcYXk//X1mVLF7v+S84z/GCZ/9GLfc+SiXXnsT933vCbBTLF+9J3Pe49FshplP0nzBNNpAaX0Im61pmLBdN+cHf7dqXCBJZIw41HsikcKmSmcVIkWxjNmqor9mBS/ODfl3//VKLrnmJs46+RjOO+NkPvreNUwLBE2dJRGoasU5ScvhBU4kY5iJ38YkKAufJxA0znPhHmlFoYaeNXgVmBKeeT1wxQ23cclVN/LzFzbi7QAZ7M7U1AqqqiLEgBAQbK7aJsdscRb1NZPR6LwZKjIxFGsbMO3m5tyszHeWTgU40hI+51/VJB2nEZDgMVFwVFSbN1Jpzac/8E4uPOtUzjn1SFb3YGChHo5IptMWilS0CaS0g82qhrquCTIuAjbnaYdtoyPAHcC0J7akWRyq7RIjxNg2iRsRUENdVUy51Gc7OzdiTa/HOScfyumnHMq3v/MT/uaiq3nohz9Fg6U3WJ76SU2RczY5aR0VNWMiBMYFHR03p6ftIlcoJ5fGE1hQUZ5PQ8lS349mcKXDWstoOMQ6R9nvM5qZg7KkxqZ5Uqv35TU/4v/7jRu4+tsPceyRh/GVs0/h4x/ak77AqIapQvK+UQrT5B+FyWS5SEvtY8TxVjVBb4AsDaKVMFlbEoDXIrzw+hyXXHs7V91wG0/9/FkGu+2NWb4HxLRPZ7bMgrU4l4o5IYQ2qtOQrMeaCLk1wd0B4U1sJahgYnaQEYOyYGAXMVeE8meoSalLSZ8TfGBQOEabN9A3FZ8/9P18ad2JnPC5w9h9mUBUrKY0R9FPgmUfQ2qXtGBVk5RLUnW8KAqaso1qcv/Z5lCsDkBHgG8KITuKLLyj2nyBGxF8FXGFbUWzGiNTg4IYPaUKViynHPlBjj7yf+Q7jzzL337tMh754Y/ZVIO6AcYViC0J2R4/apJtL5xxbBorfh1XgbffLrUg+bcAMdSUZVpm1nWNK/t4jYRRgN4AYkwRlBgqrwg9BnscyMt1zdevvZ1b77qXz33io1x47jo+94kDU0nGQ0m6CIv2Y83YAlHT9qQgtolqm+Rkejzkamwg4kOS3wAMgaee2cg1N9zCpdfexAsbalx/NVN7v5sQDT6SCk9qkp1ZjOlvMRVdiqLAqBJqn6a3LbZftmdJNYG2eDXem+OfbRU+dYeIJi502TzDhxFmNIfOeY78yAf48vlrWXvMR1ldgolKqTVBKwRL8itMI1FTFJt2W3I6M9TB46zLVd80hLVxle6wfXQE+CawWEtYm7HLJ5nL84XbS8emJYoYwSJ4X2GMZRrLSZ88gKM+/n/he08+yyXX3MTNd9/Phi01xfRqgvawRT+LozWlonwS4gqWoB4jJWoLYlUlKw8ajRmMSww6FnPN2/Dmgm3yXmmZqRiicUTN0ZrVJL+RgMRUiU5iWcucB6QHK/blNT/LVXd+j9sfeIzDD/4gX/3iOo474v2p2olQV55e6Qg5T2iN4PPM4km9kUqOZhpC8TFJkIxFnVABP/z1Br5+2bWs//Z3eGPjEC2mcSv2JIpL8h7RFBVFD3kJGEyu1FpAlTrkfKRxyY27WXjP04lsmzha/Vyu7HqRpEf0QzSmnKKI4HOvtrE9YlUhRjAacb6mntnIlIscfuhB/OUXz+bTh72PVVPpcMWQiFKlSIX+xiI/uwS1NxEBk+c+J/LL29VE2hM3xWZbO2yNjgB/b0xUEfPPOPF7rgbQd2nGnA+eatazcqrPZz92AId//L/h3u+dyJU3fptb736Iuc2bsboSEcegN2A0mqMoC6rRCKKnnFpOHQStK9yyafxwM/PlJPk3k2ZK6HZmSiQVl+Sl2/YvkLYY02jZsnYuuGWYqQLv4ObvPMZd9z3AsUccygXrTua4z36S1QPHKKZBTn0neUZuow+MWCHPL3atWakiWJeWlDVw//ef5h+uvJ5b7nmITUPoL1uDWb46RXyaTE6tGnTCs25+P3VOEUgmd405UtoxKWwlI5qY7qeZlIL3iDEUzhJ8lSVTDucs1XCW/qBHmJshzm1hamA4/LMf58vnncmxR7yTKaBvYHbks5TFEWLS9ltJonNRHesNO/yjoiPA3xFZFLODx/MzxODrGudMNh2IGGqcKrEWjj3sXXz2sH/Kg0+cylU33s5t9zzACy+/gYbl9GwfraEspogKdV2ni9sZ/GhzinqaPJnqOJUWdf6qd6tCYHpNmHdh+bGouumKwC1YPY/1bgChAuNKhhH6u+1PTyI3PfgT7njwcY447CN86ey1HPuZQ9ljhWHGg69HTA+S155K6h/x6ulhUo5OCqLl/9/en4dbcpVnnujvW2tFxN7n5MlMDUgCgZAYxCxATAKEkNCA5llIQmKysV3t4unusm9XV9W9Xa4en+ruex/f27e6b9tVbWxsC9A8TwiEQIh5NGDEjJkRklKZec7ZO2Kt9d0/1orYsU+elISmlHTi5RH75B5jx17xrm98P+5v4LNf+zEfvOQKPvPlf+D+HQ1b9n4WY2dpouCKEd6vprq7PN9Ysen7tN00ufOCuX7bWZti933WFAH30ZGfpkFGrXCsapwVTTuHThqaOlJUY4xEgp9SmEhVKsv3/pS9NxUcc9xreNd5p3PU4c9JhnsAE+tU7Ow81hTJ0o4Ra9PEOFTWF8fonPdZEfd6v++AB8dAgI8WrUvZW6T9MTu+CRRFkcsT0uS2alSkeE1sqMRhUEKAI19+IG94+bv55vlncMW1N3P1Tbfz6/vuxZZLKBVNNGCT1HmUNKnMqMwF8pMiSbZSYhuEfyjEbmRi5yxpG9A389aHzr6fkARGfR1wVcmkaVgJnk37HgR+wh1f+wGf/dr/ymEvPIhzTj2e0084mmfuU1HTlgVbwOOzmrK6NOb7+k98gw9fezNf+Prd/Ha5YWHTPoz2rljxEMXhm0DdTBhtGhGmyxhpy5jTu65N/KRHWmGHpLVI1z3zMMUd1kz1SzG2TJA+YIsSY5VQN1RWqAhMt/2GLYuWtx//et7/rvN49UsPYEyKj+JTWVDIrrIzlhAaAJxrC8zTZL8o8xuqrP1yAx4xBgJ8jDBf1jtzggujoJ52uKyIxWtExGAKhwViUCrJA9wnU15z8BYO+8/P551nv50rrv841938ae7+yS9Y3LovUYokw1QWGBw+xk7nDsDkjHQ3he4hrhOrfvakda0N1lhQqeQ3lf4YCKlXWEO6WItizErTEINlvPezCfUqX/vpDr7y//ogf3Xp9bz3vFM5+9TjeOZmskCBwdmCncANt32FS67+GJ/6/NfwsoAbLzFaWmA1BKw1mKqAJuKqEmMM9bROMlcoSKvK09bv7R5J2SSkomqRB6XAWSJBuhrQdH+qEHAqufc2sFBVqI2sPnAPzz1wK2898RguPOcEXvWiAxkDBo9DCE2gsCVGDM6N8ymO2DymIFmbETFKMHSuev523W/V3fkgFuzu0WtB3MAYCPDRYrcGVn7AAijB1yAWY9PsXt/WFWorvCBApFwoacIExPHSZ2/l0H92DheedTpX3vBxLrvmJr73019gygWsWaRwBnEFIWpKZKyZK/twbIT1lD5iK3WU/zWP5HK19YlFkVreQqhBFRWX55UYVmvBuEVUHaPNY374qwf4f/z7/4O/+fAVnHfGiZx2yvE8Y+8FrrnmDi679ma+8o8/pHGbYGF/qmoTPsA0porIJkYI01zuUSVLKGvTpfnKNiUM8jlIP8Hs2CNZJIFUZvQQP97s/HRkOl8u0yloh8CoNITphO2//hnPesYW3nXRKZx/1tt5xfOWKEnE5+udVIXDSoXNCiwhKCFmIi1MPkohBJ/nzthZVrcfsdQ2HvuQhz/gITAQ4KNG6w62BbBmrp80hgZjFevaTLLHYJLSBxDEoxrxPlXuO+dw1mFjxDcBZy0v2r/gv/q9Ezn/9OO58pY7uPKGj/PdH/2cSe2RagkVm+oV85yQFKB/cLSxvrkYX/eimL9JP52ztkw5XbAhTpNm3igNifL1ND3s0pxjmjSWM1LgRnthxpv58X2r/M//6Rr+41WfwlrLL3/xK4ytWFw6CBPS+EltIiHEnO3MM18MWGepm9XcgFzMuh1o5aUirQRZJ0dFIq7QFmdbemnf2PsP1rWIcjueGNOJxsYYknBBWGXnjm085xlb+aPzz+Od55zEC541wpES6aGeYI0wKjcn6TJnEQPe11hXziQsIQ1wEsXaglbux8U8nKqfyRV6YYmH6cIPWBcDAT4qzF84be6g75AYa4nqMZ06h3Zxtk7bTwxlkZSlVSGEVGBdOgfR433EupKD97X884vexgVnvY0bPvZl/vrDV/H17/8EigWKokRsheLygKhUEhHoF/a2RNGaQia1MrQH3k6kA2YDtdvXt/JK80XXMSrWCI2v033Wpi4LybJS0E0Ya6JgTIlbGFPXNb9eTlZqtfdBhBDYPvHJsjKWOkSMs11mmNzh4EOTwpJVgWCI+buYtlRHQ6e9Ry80oL1uku7g1yG9tco6SXLLZzHTgFNFQo2va/DLHLzfZs5777s494yTOGQ/gyOJvRR52nRZlRCFpvEURUkIWbWldEzqVWwx7r6fMS6VM8WQ0ubWzgRee8fU39yEWSvfgN8dTw8C7Po2Ex6JbPojw67WQktu0nuOFTf3eEuWaXaEy95MjvMIGNfL7InF2STeaoIwFnjWArzv1NfwjpNew6U33cVHrrmRr371O7hqM268BW8KPJaJr7FFchHVe8QWaEyDlkbFAo2fEmjQGBApkv5bFCSartwl1SwnS3CWHc7I2dagkU6dRkFD7JUEae55BbFp7kYTI+Jc1kpMs0wAKIpuAwHSYKv+hd91ldB1Ds7+HYGAzZ0e6YgNKi6xYGytqJgz5LllrEmKz6qaajWtTYXtWdMw+AYxQuUEG2ri8jZMs8LzD9yPi8+7mHNOOoZnbCkYFekXC14pbVJxSTJbAgYKkwo2rbVdXemozL3T0tellBxlyOfTtiVH82tsHo8kjrexY38tnh4EuEexPgk+1HP6jz1oKMfMKLPIMa3kUib1kfed8SbOP+NN3PbJb/HBv72ML/3D9whSYsdb2Lq0xAM7lsG6NDhHDMGnprLGT/H1FFNFkuMciT6C2iTKaYToG7rA2cNQnlkXMkusdC9X6WYQz8lPPdKYVu+42kl7iuTC7UzmrZpOBEgWNkrW4AtE71MXnsmjQb1PLYGFAT8l7NyOhFVecvD+XHT2ezj1+Ldy4N6C1EplI4bUVufsLGH8cGYrr/+oeThPGvAYYCDApxD6RbkiQmmgiTCKcObRL+PYo17GHZ/7Dh+85Eru/OI3eWD7NjbttT91MEwnShClqCxiFO+XWdwyop40xJDl7gVQJUhExGU/uLVp+27irOWrG+35MI65j9/dSl/vMyKGOh+3yTWNPVc3f5/UERMRTZFald6sE0vq4BglBejaN5TWMq5KYrNKeGA7ZZzyoufux8XnvZMzTzmGZ2xKobsQlIVSIKbscDdcvj2CoQ/3SY+BAJ/k6FsR2s9wkowzC9joEXVsMXDSm17M2970b7jjiz/m7z56Pbd96ktgRmxe2AsP1CF1KsQwYWUScLKQen2NYl2q/YvRJ0K00gui9S/m3+3C3h0JPnqkiWjJw03F0DOF5jYImBM6qjNbO/ciY4TQTEEDgSQmUDmIkx3U9QojGg5/8fP4o/dewDFHvoSlMl0wdd1gncFEn7LeUefIb63AwoAnLwYCfAphLZGICCHEnEBRovdYjSy4ihNfdzBve90HuOurP+av/v4KPv2Fr1NPFCkWqMZLBLeJECUXV5vcgJqk2kVAtFWlznEsnXVPzOJy6wkBzOPxi8fmLHRLMpKTAf1MtqahP5r7fzWTeFSb+VGx1lE4B36K+GW0XsE0y7z58Jdz8Xmnc9pxr8Rqark2MVIYQ1VImjJnLBoFsWYgu6coBgJ8kqN/YUluxerX+qU2sIDGSGUUMRbVgA+BkS1526sP5s2v/lO+9I8/50OXXs/HPvUF7t3+GygWscWmRFvW4STNJYkx4rLQqw81xkA7Ec50kcdeYH6Pjsk0a+JjKSvfykGIJk5PXnD/nCXRBSuKE4+pp0x2bmNkPEe9/lW898KzOeqIF7KULeymXmVUlUAgNgHnUulNUgtPG8TaGkx4eDHAAXsWAwE+xdC/oJSI2qSu3E4Wi3WTOjJMSaynFAKuqDjiJQfyqj/7I/7xRxfwd5ddx4233skv7/slC3vvz2qYEtRhjE0N+CFlTI2WvS6QXA7TznXMcbcZ1rrFa4jxd+WB3fQu7wqXSFhT3te2CY621zkyG+oDSGq6w2iDiw1h5QGcTjj9rW/g/e96B68/7FksSC7PizXWpCLlNmPvCgdEgo9YV0BOfqzn9g7k9+THQIBPAewuptTWFPrYoLg0BtQWiTw01dMZAiFOsBHGbsQrD9nCy//lxVx8zulccf2tXHnLx2m2LxMax3hxK9GMWZnWqLGUVYWPDUjM7m9P7w7zsKy/xy/+10P7fWF+XjAG42aip0YjVmtMmGJig2PK6ce9kQvOPIU3v+4gRqSyyCJ6JDa4wiXRW2PSlD8x+BhwxmCtdIKja7X3hhjgUwcDAT4FsPZC6s8pUaA0LrmpaK7fa5/o0GwhOaOoBgq1qIHXP38zh/8X5/Kud5zAh6+4gRtvvoOf/uoXxGILC+WW1IImhmiKbGGFVCNSZ6293NnSElyU1M+KKhiTmvhj7KmptMf0EF92t1zZ6hdmNzZqEhIwqVjcOUP0EBRsUeJ9ADEYsaAei0eanZh6B1tGcOxb38B7Lz6Pw1/6TApS10ZLftZEcAWowUjR9VoHIA0UD5gcK12vD3cgvqcOBgJ8SsMga2fpytpOAUDTkPLkvkaI3ZhrXvHMzRz6gQt533nncPl1t/LRq27lJ7/5DUW1hG881hWETlHZ4sYLqeyjTsolnd6gs0kUQZUYAjGErsbu0UBE0BBmpBpCN++izJ0VESE0aSiQWEsIadh4YS3BT5A4JSw/wP6bR5xx+olcdPYpvPQFe6fEBukicNJ2Y5CLupPb3LZKt5TWyRIMJPe0wECATzvErrOtu1y7C7i9cBVCxGrEqlBaw0FbS/70/afy3gtP5SPXfIIPXXYt3/3pzzDVEgtL+xCkYGVlSiDVy1EHyi1LmBiI6vHeExqPcw7rUsF1bPyucly7swR3Z/mFXLCcMxpCxFiDiFJPVxBxmfiK1DwRFWMixCn19u3QLHPIs/bl9LNP4x1nnshLDkoCBQJI8ElF0GQZrXZAVXsmNd0lrB3//chLggY8uTAQ4NMJPYWQdLOOTJIAatGsZ5fkSAybS6gjbCngj9/5Nk4/+Riuu+V2Lr/hNr55909oqBgtbKUJU2wxwi+Mkkq1n2CLoicUmoqCNabETNi9T/uwkFRtAmDn3GrJ0leikfF4xGQyoZmusrS4QLOynTjZzgufsx+nHnsS5595Ci969iidkyYQY0NVSJqw1qqBrelWiSS3d6bAsvbIdtWBHPDUw0CAT3V0iYgsZKC7sUh6F2qUFKdTcrdCCDSxoSxLXKnUUXneVscfnv82LjrrbVxx3ae48oZP8KV/+B4iJaMt+1BHi1pHlCKREr3YpKaMhBjFxFaRZc1x7XKc8wmVVnTVOkusNU1xE0E01TuaomBhPCI0q0weuJeqtDR+B5Pf/oYXH3wgZ510MhecdTLP3tdSAM2kYVQIRWGAEgWm0ymlq1Ifdg4d9LUVpS920Xa8DFJUTysMBPiURl+Nhjk1lzlIWwicNaV6F28gYp2lQLJ0l6XUQPSRBXGMS8P7zzmKs085ihtv/yofvuIGvvwPd1M3ymhpb7Almt3QkONzzjnEWurpFLPOQKnfBSHHEokR7z3GGFyOLU5WtjN2EWGFwgde8Jx9OPOk87jwzFN4zj65gS9EKgtVqaim8hXyIKOyGqcRy13vLrkvuiesMderPVh9TzcMBPiUxu5jUWuVnQUQSSRgYhbjEvBp8CQOB2KJXjA2CY66nMGtI+w9gnee9GpOf/urueOu7/LhS6/mE5/9Kk21F14LivGIalQwWa3x3qeF1St/aeW1dj98qZXq6ltcoCHgioJAInDrcgzTBwqJ+OV7OPwlz+P0E9/GmScey3P2EqyC9REhZYenkwnVaCGVDPk0CxkMjW8QW3QxUiMxSVG1mY+s7TgnKNbGU58owaEBjysGAnyaoU98u0YA81hF0eyCBqxIN1fWtKMifW6hsBaCxwkYLHUUFiKcduShHPWaf8ldX/0B/+mj1/H5b9zNtnvvZbzPvhSjESurDTEEnHMzG1Xmj2pGiO19a79IqjM0xuaB5oGFymC1ZvsD9zKqCl7w7P35vXdcyHmnvonNRSplqYBQr1KUeWlrpBqNAGiaRKYAPiiFK/DaTsfLR5r2hvYg1o0NGsgdOOxyhgc8tTAQ4FMeu9eKm+kSrnORCggWQ6Sy1UyP0Jr89HZpuJzIVSrAmVRJs08FJ73p+Zzwpv+S2z/3HT5yxQ188rNf5bcrSrm0L1JULDc+MbI1iAVtPGLzKKSmTjp8IWBsGo1pbSI7DeBGFbHxyQUWGBnF77gP4jLHvOpFXPSOMznhqFfxjDJZl+3YFQPYaoxqQAnZ2gOIFEVyxxUQK5n40j2qgshaF9d0Wji7nuOB+J4OGAjwaYyHE6padzbuOi9MHRZxbsE4hZW64cQjXsyxR7yYz/3Dz/nrj1zLrZ/+Avc/sI3ReAvRVajJtXlliZ80+OhZ2LSJleVlbFGgomiMhBhxzqUWv3qCiYFC0ywQwjJvOuxFvOf80zj2La9ga5mmqxUKEgKhzV6YVsTVzGKerUjCeiekL421G1IbQn5PXwwEOOBhY63cE0BZll2M7E2vOJDXv+I/42t3n8nfX3E91992F/fv3E4tI9xoC0TBFWkK2rSO2KIiSICmxhVlGtUxnVJZxWpD9CsU1Lzxta/gonNP4+1HvYSKJDlfhmSNpsxFUmBu+St2x9sWNDPHYn2rTvpVzgM2HAYCHPCwsFaMNf0BNou5C4G68ajC61/0TF76X/8B7zn/XD506fXcfPtn+eV9v8SNt6JukaiGqW+Su51b55xo6sGVKc3ObWxeKnnV4S/kj37vQt58+MEpttcoEhsWyhKtp8mqK5MGYJsA70cTtd+/saYAe1YsPmAjYyDAAQ8bu/QkE6nrFaqyRBCsLdIcjgCbgNc8fy8O+9fv4v0XncbfXHo1N338Tn51//1gFtm8sJXagy0LNNaYZoXVbfew96LljJPfzHlnvJ2jXvM8NCc2YtOwWDhCHQGfhgcVNhPoLJq3C9a6vkpXFjTE8gYMBDjgEUHzqMhRWeKbBpsHfWtMzSWVhbr2WGN5xUFb+Z/+b+/lj3//Ij569S1ccvn1/PCn36ccbcG4kqZeYcumkjNPP5bfu/hsDj90rxS585FFZ5AY8KHBFBZjXZLtjwEbU8yvE0PFzM1I230XB+s9a8AGxECAAx4FDL7xuGKUtAgz0WSNUKrSEaOnaTylKzhws+NP33cq7zzrJD58+dXc9vE7+fkvfsspp5/E2WecxCtfshcVqVzFijJ2goYaZ0tMuZCGueUhTVIVec7RmmJwWhocMOChMRDggN8Z/SSIKypijJisqp/+TgTkfY1zhtKk+b7OCrVXnrPV8i/ffw4Xn/F2llcDz3nOFipJgjPWJM+2NNmWs47g66RzKInuQk74ikIh7TzmAIQ8ZS6QlHJkCPYNeFA8PQhwjeDmoMf2+KLLBgON5mlyeRB4KqVLFplzaUi7qqJiEaCymemicsg+CwQxHZkhqZhZxCYplpQXxloBSeSmPSkqI33HF6JGrCghKNYKMUSkna87GIUD1sHTgwAH7DnkIrq2Q6KbpNljnFZ5xtKqcUl6hQYkybiSClKKZLXBzMQTgaj42GCtxRmL4rrSmyZ/nAaDNRCDp7Au9xC3RLp2QxzYcEDCQIADHgVijsGBwSFtJrZnkBtCN7gpSQ3kpIVxuUG5oU3V7iLkCvjVCW5c4kyRp9YFRGNShTEVEZgojCz4aJAoGDsr25Gu7W/wCgbsimErHPCIIRhs/p+sJZisTp0msGluy9PusSCpD7h9pO3YUEnZZc2WpRuPEnf6AFGIAXyMaBbMWgWmAtd+4ku85w8/wOe/8jUChiZ4tFUj3KOT6wY8mTFYgAMeFfqS/JqJDeb7kLtWs5yTCMxmC6fOYJsd4TzWMnu/Ktnow2Kco1HwMRmPAfjVA8r1n/wKV157I9/91pdZ3fEbPvCBDyR3vCiJMWBN1ipcI2o/YAAMBDjg0UB7t5m02sTE7G6zSybWrHl5K8ASVEBiEiUlubHiLKvTSFRBCggGfnE/XHvTx7jkiuv50a92sLI6Zd/Ne1FqZDVYPFBgaGKNMXZwfgfsFgMBDnjkkFyD11NLFlImt81lBPJ4SsnSopoUoy2kBEfrnmqWZehZlEFhEhqKqsAD3/jh/Vxxw+3c/Im7+PEv74NiEXX74DY1rOgqaEW1tFfymL1nkyvzBDftYpUDBvQxEOCAxwDZfW2JcD2TS4W2LK97CRGkzecaIpao2Y3OlqQ6w2e//TMuufIGbrr9s9y7IxCLJWK5D6YcpVigX6W0hmm4j8lqTQTGztEQqBjc3wG7x0CAAx4F1qhQr0MwnepKy0Ht7A3y9DY8sWmoTUWwljo/rwE++8Wf8dErr+VTn/8aP79nG5v22Z9i6yZ2ThpQiL7O1qVhqmDLEQZPSSJPl9M0D+fYB2xMDAQ44NGhdVl3o0Q9Z/H1SvyiCIKhrgNiCrypCMCyws0f/xZXXH0Dd372K1AsEO2Yxf0PoY6WeqWBsmRclEwmk+49IwbBYDSPX2rFnHt9wrOj2+WQBmxQDAQ44DHHQ9lWkrMiHgPlJqbAL7bBtbd8mqtvvI2vfOs7xFiwuOUZeHV4FUIQ1LosfwXTEMFZJCZ7csCAR4KBAAc8Zthd262qdnG41BaX7m8UfnLPlEuuupmbPnEX37j7R0i5wHjLgfgAUzUEcUQjBGyqnbFJkz/GABoRZB2ZroESBzw8DAQ44HGH9N3frOAMadrcv/v3/4Hrbv8isVxkYd+DWK6VlViCpKlthAijEsSlYSQxJ1sk0gUUGfq/BzwyDAQ44HFDOz4ytaTNLMBWNqtR4Se/vg+3aR+iW2BHLYiriGKJdU1RLYIxNE2TxlVaC1agqUE9ZVkQfP6QAQMeAQYCHPC4YE6EWWfi9IkMTf4b1DomTUQs2KKkiW1FtaEJNaw2UBaYIo3sVN8AEWchhmbWfzxgwCPAUAsw4ImFmdXDKBBC6EROQ2xFE0CcIAbKhREGTSUv6gHFWYNB0WT+De7vgEeMp4cFKDKnCag6uERPGqhirJ1NbzMmjbAUECNJt08bnI258deSqgAV0UjwPpuTSXAhSV9FIkl4VWmtyj34HQc8ZTFYgAOecOguZBWJbVvdGon7tc+NGFRa9Zhh+Q54dHh6WIAD9hzWUcECHnZeQiV2JBfXkl0XOEyElzqIldjpQCupw3jo8x3wyDBsoQP2CCK5g1hMJrr+Uuz9rZK7OWhbO0AMUcy6rXcDBvwuGFbQgD0LNaAut9StcWvn/F/DvJpCIs8BAx4NhhU0YI/CqMFkF9esCfh1dKit7NZ825vRwfUd8OgwxAAHPGFIoqe6a9xQDUaZKUeTcsFtNt8CQWKWiE7SW1ZjSv4PdYADHgUGC3DAE4K4m6zIjPTamN5sSa5NFrdD0E0XQUxoDUfTV53WhyPKMGCjY7AABzw69OZ9zP1T5h+fFSv3aCmSySykIUndYwaVkNroRHr3zxDEdu+59lFp/+spUw9kN2A9DBbggCcUrZU2W3hr4ni9xIYKPYn8NEDdaPuKZDGqzKv9rf2cdY4gcfKgojWAgQAHDBiwgTEQ4IABAzYsBgIcMGDAhsVAgAMGDNiwGAhwwIABGxYDAQ4YMGDDYiDAAY8rVBWyJH7UiDHzCtEDBuxJDAQ4YMCADYuBAAcMGLBhMRDggAEDNiwGAhwwYMCGxUCAAwYM2LAYCHDAgAEbFgMBDhgw4BEgrrl9auLpQYBr5gAPg7KfpNCnx3Ib8PTBsCIHDBiwYTEQ4IABAzYsBgIcMGDAhsVAgAMGDNiwGAhwwIABGxYDAQ4YMGDDYiDAAQMGbFgMBDhgwIANi4EABwwYsGExEOCAAQM2LAYCHDBgwIbFQIADBgzYsBgIcMCAARsWAwEOGDBgw+JJQ4AKaE/Fqvtb1jwJMPPqVzNIJH0lk//e3WeZ3t+z/2aItDpnuz4e1/w3YMDjCAWIaxfoOpityV3W7JoFvv56X/eDedjzSx/kekvI1+VDf5EnFG5PH4A+BtJ9YkBjQDAggqrm+bMBBYxxKIogRNJMWsEQgdorhUsHYQGiR9rfSdJz2p/MAtLd8wgOXE33svY9B+XCjY1d1kF3R59QdkMu7XqKCgQQJUpa44pFAEeEEMAWoOmdVNLH2PyBszUY88WUZjc/qunNnfZj/wrKDz0kWT5xeNJYgI8r1hHibH+C0gmgeB+IgCTmBIHpdIrm586WQvtjtjvabj6ru2U3i3rAgEeBTH5rPSeDZOpLaOoarE0kCIgkvjRpiecNHci02ceD09/vQh1P3m1+j1uA8jhbxGtPvYggSLbhGkz0IIp1JV4jKoKVghADZVURSTtl2i19964hLwDbe28lLbDkgudbWJf4nrxLYsATiV3WQXfHWoIxuzxvtjlLelnrtUq6sFXBlGWy5pxBAO8jhTMQawgeTAFieSLxeF/zvwue/BbgY3SyVNd5I1UwEUKDoBgB731aVKIE73t2Xuxtt6Y7tLXxll3iL0LarQc5+AGPBJJj1muYsiU/BWK32Uq31oREhNN6FRElbd4R6/LfcQq279fo+tfI0xx73AJ8IqCqrDcmpI0v40qaEFADpUuWoQjg+i9qqTCCGozsGplZ+++W8mRuV+89S3dd2AM2GtZfNbrmGettn+1qShZVuz77QWalKh3Q0EwnVNUCIeYl70ao95kcNy42lFmyNsYREAKWJkK0FRMfcvBYUO+JTc3cAu27tex+UbZoXzn3qX1LcIgJDngI9GevrbdaDMx8Sp35xaqABvCBWE+oqoqI0KihUYePhigOlT1HAU8G7t0QFuDuEFXAjKhjigi6YsQkKCZGxs4+qHG23mPt/rseCe5qDQ4Y0Md8adau9ybML59kGwqa2SStQG3fJPvBxo3w0bAaBesKvIIPMHaWGEIXxxaRJ6xK5clyGTwpLEBpM1I6f9/jD0MTFDHCn//vf8PNd3yVnVHQomQ1FNTeopj5eqp2L547vnSf9G6l+4T1P3fAgPWw2xJX5jO2u8zl1fY+7eKCQS2TYJgyZtVUrFJw1c138ZcfvBQc1O1Ln8DY35PB6uvjCbsSRQTJNXoAxvR2PNXZLN/Y2u8Pf75v+3pVhd7fa18vs3wZAFYM6tMO+LkvfYv3feBf8kf/1f/MDZ/9AdsQmsIyARqFxsfssnqQgGoKKnfFp9FDDGkhtotRU+nBmsjfgAE9mG6T7Yy2Nf/F0OS1lNeUb7q1pu31YgATURoiAQ80Vghlxf0Bbv38T/iDP/lf+Gd/+md89ovfpABiPcXaWQa4f2081LUnIr+7OxMjqorNr0u1hrsn4CdivvcT6gL3v6iqgggS84loCUsMSEpaxBge9xR9ZR3TCMV4CbO4L7fe+XU+/rlvcMJRb+ScU0/ghKMOxQkUJmXWjBpM3mUFmNRTRmWViqdRaBpwDjQiKKoGRVCBQFtMTU6ADLS40aFzf+vcHW2CwloLGgg+ICKYTFreN7iiAiCGiI8BcSOiGCbAcgO3ffrbXHX9x7ntM1+ibgKLi89gvLQXCoyrgsd7a26NEScGrEGjouq7613YPdGtZ8Q81tgjMcDOYgNMPyfQVmn2/36Mv3/bEQJgctuQFUPjI5MgVFv2o4nKLZ/5Oh/71Bd4+aEH84fvfScnHvNiaoWxOCR6NDRYaynKMdPgMTFgDRhnIfpE3HswwDzgqYM2k9sSoEjuwdDYGQfGFlhn0RhpfMAYgxgHGCb1FLGWWFQ0wApwyye/xV/+9Uf4+rd/QrRj7Hg/qkVlsv031E2qYtXg02b9OEJEiDHig8cYgzUGIya5wmFGvrKHooJ7hABbVjfkok0xGA2zk6KKWNeR5GONfjbYWoMXUGMRUxJMSU0isqJY4Mvf/Tl//K/+e4587Su58JxTePuRL2Nr5XDWUQdN8UvrsNbln9Bnn1chBsSmtIhR6Sr2H2Ej3YCnGZL7l5K3c+1oyfFNF4em5EaMEVUhREFclTyKCNNGkTIV7P92BT5+11f4649czee//h1wmygW98HHgmBGFDaCKVCTLEjzBLiYqprJHWLM4SGXOKAfBttTeEIJsB8DbBHjvAksxqCE2fMf5+Npao+WDjEWH5XgAwEHtiQ6g3UVvl7ljq/9gLu+8u957UsP4dzTTuCUY49k783SLlWCprLBCoeIB6tzPUoiA+kN2D0074rJLZw1X4pxuTXTphSHTaGUqFAruEK4bxVu/vhn+ZtLr+Hrd/8TjJYoth5EUIOXguiVGAMhNlgU6wqUFNeu7OMbYoo+YK3F2SI1GQRPjLHLCewOT0T8D/ZgGUz7Bb2fEaMRSRZZjOuTRWrbzZXvj+zz1sYVytKxAjRNkyy5coHag/c1QQNEsG4x7Zw64it3/4LPfeXP+eg1t/LO88/m5ONfxaIDk2N80xipJDXbed9gXbnu1xjifwMgW4CAMYrorvIDqoqKwStMAwSBwqbb5QjX3vwFLr3yBr749W8R3ALRbiHGMVZL6qZBJKIxQFFR2BI/ESZNDUBVVjzeMUBjDKpKyL3IzjmsTX0p3vsHf/ETgD1GgG0SxJgU5A0hoN5jyhHorpbiY//ZKe7gmym2qHBVSaPgQwBTIkWF0SZldzWAcaiUeBwsLfK5u+/hC3/2/+bvPvoczjrxaE4/6W08c0sy6+sQKUQwtuh2O5BdrUB97GOcA54miDHXNSfy8wo2X62/2qZcd9PNXHHjx/n83b+AcjPl5mcngguW4AO+jhRFhUokSIkawzRMQSOuKBCgnk4pR8Xj+jWstTRNk+OYubY2prhmUTy+n/1wsMeywJB2vsrB2AZs2AFNiXURohAFjBhivxFIAXzSA5SY8yWzXaTdy9paI1mHX2J+L5dr9awVPBBinMXmVNEQUEnuAhrx3hOMoFiKagFKwVLz5bt/zpe//h/4+8uu5aJzTufMk47hWXsZYjREHyiNQ1phIWkTMG35zO768+YfWrsVyNy3fThxlN016Q2YPzcP87ysq+7zUK+d/73at0jezGwtpLhgCv1EFbxavBGw8KttcNWNt/H3l13J93/0U+ymfTCb9iPaBWqE6EP2llP8PIQAIigeXC5Dw6RKM6BwLvcNh9Rv3PYcK6hEBItRsFG749euGyp0dbHtN0/lXnHW/w6EEDACrrAUEgmTncRmhZFO2GfzaLdZ4CciAwx7MAsMqYZ9bOEv//y/57JrP8FHr7qFu//pR1Sb9sYVm9CiQhCamH5IYwXxmbhU8Bqzrk8KDmNdkgMKU/ANZVu6Igpq0mLLmVnBg0aCxq582YhQEJmGGpMby/OnYYp2gToaBaNKg8VsOghf7Mu3f7XKn/1vf8/fXfkxLjr7JM468a3st7clKIgKJkaMRFT6Mc7ZRdORWraM0VyALfOKaimlkouuta3+n3uTfJLb+3qF2+uQqmxkMmzPzS7YtR9XNYU5ZqoXuSZPFBXJ1QXz8bR+YkOj77K5ykxnUjVgTXq9D57CFUx9ACkQl6KBP73Hc/3H7+SSq27k69/9CW5hK+UBL2W5TmtLfftpjmjS9zHtOgKyYCZGA4pgcpSxC1FrOu4oqY9EJeRifovxUBmhEPAxoMbQaiAZtLPsVDUlOUwb+04LTgUKidhmmWbnfeyzYDnt5DfzrnNP4/CXHoADrFmnFvHpGANcW6BsiFQqHLzfEv/i/Wdw0TmncfUtn+IjV93EN+7+CXa8BVcuUFULLE9r4tRTjEYANN5TFAXTZpJqoYyDoERNoo7GpMwZpEXW9vHO7Mms+6dmbqEbBaOxqz+MkvY10xJoD4qhrgPleAt+aolS8P1fbedf//v/nb+79GrOP/NELjjz7ey3FZwxNNGkILS4JLwqMVm5IRW5Wmtn2ZKYOlRCe+7WP6EPnVLeJZLQbgBm14c2GnYnRvGQafpcFK/kC3W+wB6yFyLZkkGRvBmrKt5PwVicdVgj1NFjTQmuYjmCKUoC8E+/abjs2pu49Y7P8oVv3A3jzWzZ/3kse1hZDeCqVD1BWotRsvVFJOZL22hMnKQRomY19bWbXhJSTX8lK7Ft63QGfN3gvSJjC6ZMz8px+sIVTFdWMFVJVVVMp1MgUJUlEmpiPWG6/T4O3n8vzjjzDM49+W28/IV7Mza7VdR8QvGEW4Bz3SAYrE87q6A8ey/DH15wNGef9hauvflTXHL5jXzn+7+kmWyiqjaho014jYTowVimKMaMQFM9VPpGiphINEKtTVawinmB9i96k1dpvmc3clXpWLULVnf/r2DUUxmIk1UKY7CVI+qYhX2fzQ/uXeHP/j8f5O+vvZUzTzyW8886mYOfAU0UXIQFW2QrtMHaZGWmdxdiv0gUnbn03a7YWzoyc636hNY9VXvP7ZuR+oRtsk9u6BoLurPwZuen5bl0O7MYo4CS2iVt1wrZvodJsa5MLib/viKRorD5SQ2N94gpmeaaf2/gx/dEPnr1zVx9021870e/QKpFyr2ejdiSaePRELACErX7/N1j1z6k9vpLhkH7WDH3GhVDICX21DooYjIAYgAcIgUahKgOW20iNA2TZkpZWgpR/M7f4le28bwD9+H8iy/gjBOP5fkHLdKmBKc+UFmzxxfhHq0DhK5pAkPEIIRmyoGLY95/zjGce9ox3Pixr/HRq2/m01/6JrFYpFzai9oKoa0j8pHoAxibmDQGGk2EJ0WZ3cgUw9CuFyNDcyzjESGCKEYizXSF6Aq8KwgBXFFSLS5g3Jjv/2aZP/+/PsJHr7mRM054K+869zReeOCIVQVqg7MBK5pjnSloLAiqEdV0ToTWItXdWC1xntjnTja7WoG9i3xDJ2EyqdHGincxiWP7NNqHu9hd7gdq87YBwWrsCpiTRL3BiBAjBA1dTVy67iW5prboXNEf/trzkWtu5Kqbb+d7P7sHLZao9n0uAYePkRgjMTQYIs4I1hmmfv5YuwOe+y79WRxryTJkRZhdYyQqs7KbgMxCSMZhiam0pklJwnFhGY0KVrf/luUH7uHVL3keZ558NqcffxQvOHCBAvB1RCVQFQWle2JFWHeHPa4Go+oRa7D59xmZsssR7FPCBae8itNPehV3fOH7fPDvL+fOL34LoWRxvBXfGJw4ao1E33QCj2ItKhYfevG9TLCmT4Bzu4/JcUKTl/VDG+cKBI2MlhbxMWXrjLN4BF8HxIwwpaMYLfKbncv85d9dyU233s6ZJx3LOaedygufW+HVJPXeCM4oLuWmSVZnTBL99Cy/B8mb7O4+WY8EByTksioDaxr1255ueue8lcZte4ik96AkItHYO98pUWAUjE2WeOOh1tQJ0QhEA9/96YQrr7+JK669hZ/dtx1vRjDaBykWWKlJkvbiAI+xFmOUED3T1Sm2SCGhBysN211FRaueNRPzsIi2lm26t4ngkeSyO4cPAtMVGg2UBiyRsow0K/ey4/4HeMWhB/GuP3knpxzzFp65FxQKWivWKFWZzkFQTwiau1n2LBHueQKELjYiarCSLvIQPc4bjEuZ4JOOeAFHHfGv+MSdd3PJpdfxpa9+m+WJUi0sYW1JaQyTGAm1JxKx2DwMCZKz3VfPyG6BwIMNZYqztb2r+Gm+aLAFK9MpiMWWFUGBJiVtcAUqltVmiokli5v355cPrPDn/+lSrvnYXRz95tfyvovfwXMOsCxYaFRopp7KRQpDJr8HOXH52FI6KcWBdmfwzVzi3b/lhoTMbKKuI6NblLs8teduSu//23skW+oycy1DQCUttGlQ1FjUwFTh57+G/+uSy7nhtk/yk5//kmrTXlRbnomvI7ZcZBoEGRVoE1Lls3EgSgip1dKNitxd8dBfU+cSZv3blIBJf6WREMkvki4J51yJ+il+ZRlTjCgqR+GnmFizurKdYBpeeejBXHD2ezjt+CPZbxOYCNYHRs4SJGDysKXgPWJ0sADXHkJ7sYeQg21GcQ4abSiCB6nYJIYzj3wRpx35Ir7w5R/xt5dezXW3fZqJWmy1lU2LW6hLCzGkrFvr3aoDabIFKHkhr19s3Q6LeTiICHWjmNEmYkjugClGUDpiSJm0NGXOInaB1RCwdhG31yI/vK/mu5ffyqU3fIILzz6ZC846hZccVFJVjhAg+glVkabZJX9EUWkprgc1ncUxd9s+vOZWdsl8PhlC0XsIbcJ9/q5ZnG9tL3c+1621mDbB9Or1e1lTZYIPERULDibA3T+ecNWNH+eK62/lZ/fuRMoxo30PxgdledWjYjEhJOLTlEBRFfCeqArOIcbgQ5OO8ZGGMbJQx8zVzwIegGRfJDTk5EmkskLpFD+9n2b5PsbiOeENr+Kd553BMUe+nE02RRI1NEjwjMoin2PpzqVxLrV/PUmwRwmw7f5A8wZnAJNmcxgsUT1GlbIoMKSgrG8ihRje+ppDeNNr/gW/950L+fCVN3Ldrbfz23u3UyxsZWFUolPFaqpRSqT2eMjPG8S5FIMUC9YS6xqMwTpH8HXKp4ngY8pOq03iCWoKxps3sepX+T//7mquvOZmTjnuzVx87skcdui+VDbFCVuxpNYRnqnmtAF3syvLzVmHuz37j/XJeOpBWtcvobXu2tt0f+88pwc7K9F0m03yKma93oJqCro0mpZGDXzjB/dxyWU3cv3HPs0921YoFrYii/sSxLK9SeIaUgqFc9R1nYUKYueJ41yXRFSNu/99NR2Tya62EjE6I7q5DVBTAk7yWM226MqkFCVVAQWBkTTEuMJk2704nXLcm17Fxeeeytve9AoWXCaSGLHqc4zToDEiJqYyGWaeSBSSEOvj3Ib3cLBHCTApvqS/U41Vg5F2B0q7hojrfmcLOGdQIjF4jDhe9+IDeNm/+T3eff5ZXHPjbVxx/W3cc8/PWXCGMja5YkGTtJCAlVT2qbl0wdkkHWRMataOMVKWFZNpk9PTu2cRQyTGHIVuzc38o6YiVDtbaia5Ro2mDzNqmIZIaUfY8V5s95G/veYTXH3LJzn5uCO5+LwzOOzF+zBCEBUqySUNuSwnFZjKLFbZXohdVL910UwXz0nH3L9unmIZkLWxrEfbSaPau+hz2VPPGlbA5WxuV/aSX9dafGIghEykIvioKUHiLLWmGN+3frzMh6+6nitv+Dj3bFthYWlvRnttYepT2EMVpCi732gaYlovmrOk3fduZmTYWm0iyToE1qtrVAIaAq6qIHq896k7I8DImq7sVMR2yRtpS7+wOAWpd+B33M+mpS2ccvyRXHDWybzh1c9lswXnwYVW3ckgtshHF2aNCOKRNvMMbeyIwNr1+MRjz7rAa4POGbtGVxIMoJJqkJIrG1idNozLilc9fy9e9IHzuPCcM7n8qqu57qorMWE1FVpKFpPMqX6j0hHDdDrB5NpCMUpVFexYnWKMSwslzgLZ6wWazVoLgfy8fra2qxmbf5VRw2rt2TTazKSeoKVhRZS/vuJWbvrkZzn2yNfy+xe9g5e9cK9UxCMm9USpZ1TMNgYNcWaJaMy7iRBDwNiCeRfX5Fjh79JJsrHQN6RDBGuEHJdIVpkAWUk8hlS/6X1MvWomZU1XInz97t/wocuu5fa7vsLPfvsAC0t7Y5cW2RkEh6WJEWNNb6pb/wB0fv0AaOxs1NlTW8vOrJvVd86h1uJDjSVgrSCiFBbqaUNZFmu+dK6PlTT4S3zD5lI584Qjec/73ssrX3YAFTAi6X0UpH55NYKYmXhJItT2uut/gKyh6D2LPUuAcz+WSdtpTw3adC1ja+wXYyGmPXhkY6p+UfBN4IXPKviv//l5nHvSkSyWDmt69k//83LEezQasQz4ZgpRCY1Ho6coS+rpBNyD9yvOModxVgC/y3fruSvZPYFUZG2LiuWpx5gSnGMqkU37H8Kyn3DpTXdx/cfu5ISj3sAfvPt8Xv3SZ7DgBEfJaj2lsAKmxNgsJyZkFZp0JZgHdTEG4uui/Mm3oC2L6a+2FJqOeTh0W78HOME3Na6oUm2eKQkKE4VvfP9+/uqSK7j+tk+xbaVhcZ9nUu29xMo0UJQVhUuEWVYVTWxS8qH1IDrSy3/E2fqfxRv7LmxLgushUk/TWk7hJnBFwXR1lSbC5qqYGZeZ9HI7SBcy2rxY8L/9P/8HnnfowThgWicSrazB+4aYLb6U52lrVlvrNDNyd56hnaj44LWLTxz2eBKkn4VNdcnSnatZDKY1n2cFLWqSVeZctmaCZ3OZ1J29wkue98xMl83u3dgIk+kqsjCmLEusFawTXHDU0ylFWdJkgda0Cz+MH01S7AXirM9zZveztifUuJIwXSYYiytKmpVlmjpgixJZ3Be1cM3tX+Km2z/Dice8kd+/8Exe+/KDWMoacJ72RzSIr3GiGCut6TJ/rnUNPz/FPOAHwyMTzzBkPbbZ9cn8aYkaEdGukm5ae2xRIWLwxlE3EVuUeOCr37mH//h3V3DD7XexEi2UWxjts8BUDb7xYBxebfY0lZom7/nrZa1CLobX3p3tYwYk9iLDu5yNLplXlRVNM6WqCiQGvK+pqoLCwGS6QlVV7SsQNbOaeRGMwsjAKw49mFTc4NlaphrAZrpKWVVJ9lLaOLskKbgOMvOEeh1X0jvqPb0E92wShDX5SOldmHMrcd4pNvN3oypYgdhMqIoRJakmT1rxgfymbQ+mETqRxmpUMAGCr3HGsLq6ShRHVVU0oWF3pNe2FIXeL9i5wzKbBTJD7PbxdkEghqaewOIYmhq/OsEUBa6oCF6Jaln2Srn5QNQvc90nv8In7riLN7/25bzvgnN4yxtfQmlgpwcbI0tlRa7dB5MHu9tibgPe5QfY0yvw0eJRqwaly1Hbt9KZMSS98zNtPGILpBwzBSYBSgtqDZ/4/Hf50Eev5pOf+xoP1JZyaV+kGLM6bZBo0GmDVAXOlTSrq2AcbjzGN80sy9/FGGdurvS+n/Z2rLTmWs9k9w5lEkRogCQ9VRlBcxmNAUZVSSR2s3SQVAeouQupvXQskSY0VChWk6hCWZS0679NbpgkKzI7ALW0rnkXplzL53t4/e3ZJAizRp21pDbDrgQktLtffq0IWMGYAmIgRo9zZSoZkFnUZG14WAjUzRQqlxVqQZxB1aQ+4+m0F6NZ77h6rrWute1295o+YoophQZxaXyhhoZ6GhGxqAhFOWayshNbjKgW96UJq9z2uW/zua/+I6971cu5+MKzOeqNL2LkLDsacCGyZVRQ16tUVYWP6ULSNp/zKPMGT0Y8Kum03LzfWScm1bC1svSqkRhiHpkKDyx7xouOxsJdX/8FH/zQJdz1xa+yY0WxC1soqgW8HTNZ9ZhqnFzPUUH0NT5OKMYVTRPwyw9AOVonGd+uq9w3n44K6CdoHtwTERGMJH3B0HjKqqAONbYwRJNi3RFYnaxQ5fh38nByeIm+p5CutC5j2y5yscSYQgQ6d0S9K7otI3qwn2cPb8J73AUWBbtWHLQrOehnONe8LpPgJCiVFWg8zjkQSUkOTUmDmLNcM0caIIJpIEZcztYVhWPl3nsZ77cJ4wwrKyudZHfrThjWS4TsGl+Mc9XV6yzWNk6igDapc6VuCDEi5QhrUz1kYRxhOqGsRvhmSnQVUS31yLHTCbd98Xvc9aX/jre98XDOP+sUjn/Li6mKggCpv7RJcxiSRZoXsPQD0+T7nxzxmEeL35kIJbUzRvF0iSnIRnoqj0lrx2RX17C46Pjkl3/K33zkam77zBfYGRxuvA+yZMEW+GgItacalcSQRDukLDDOEesJzeoqriyw4wrvPZrfv08SsX9Z7nbzZebG9O9fUx5jrUVEqKqK5eVtxPu3sWVpc3Kec5fRrOc55v9PtaEG0KRLlww7kwIBGhWxzKvjAP046tzhyzpfQx7aXHgisMcJcCbXtJs6vX7yYJ3HKysEVaqiSHEvsYgzhKRhuu5bzRCRoIws/Okf/z5b9tqPW+74HPfeey+Lez0DtSV1CPNB5i7b2yM+WfPYXLIl3Zh1dvooEbEG9Q1SWISC6JvUx2INTfAYhKARYwsmkynGlRTjEXXTYBdHeF3mti98i1vuuIsjX/tK3nXuaZx83KtZKCwaLTFqEp3NrtYsptTu27tfgO0p3/XUr+k7zd9n3Tfo/S3EfHGlXp3HFJpKxLsZ0/nD27BDL0c/O+5oiDSoKBpjjv/bdBtzv6+mAuZphNvv+g7/8W+v4M4vf5OplNjxXhTlEvW0oajG1L4mNCElN5opMUZcWeDDNH1kUWBIslEaQ/ZK5tkhVRu0Wd3kls5jlhgUZpZru77ECBIVq4popCyU6WQH0+kyhx7yLN7zn7+L8096I6JKWaS+89hlknvHkW9ta9Dl5Aht2Q7k7q382jWe1i6roTMp49xde9ob2fMEuJ5Z/2BE2Htdm79Lc0ZN+rUyCbXque3u1v9BBINqcnvFepyfcuSL9+W1/+37+Mb3zuGDH7ma62+7k/t2NrjRZqRcJBXlC0Vhu9G/Yg0xTKHIRNIGQ7KApLUuVcUrXbY46QGaLj6i0UOWDVcCJm+rqgrWZJmi/CVcQVShrj1GkyDEqtuCLC7hFrbyiW/+mDu/8b/w+ssO5d3nnc5Jx7yGhUKogEnwSZcNxZhUBhRV50JoxoAPHpdPnteIEzO3oNcmpuZ/HLPL4OuuAF3Td2vwuQc16eg9kvEGcx8ruVYupOJ5K5psWmnykab6uqDZYgEg4DCdi2nEECXJlNUhsuo942pEnb/prXf9I3/94av41Be/wVQris0HJEu8q64qsry7wZbgY5NOmAWvTfKpSYkG1bbIOimudJllbUulsqObz2PUkItUUzmWLSyinhgarBT4kDQGQ+0ZLYwIXjEERk6Y7NhO7Vd48cHP4qxTzuLic49n300wClDh0wjXwiHYOYES05IvtMG92YUkdMOU5tWJ6Cy9dVMz62yUe5r84ElBgNAlBh4B+j/bmjt287w1zzcK3lOoQ9Vy+As384r/5t2887zTuOL627j+1k/xs9/+jGK0xHhxK5PpSo6TmNS253qLBUn1YdYiWMJ0grEOldgTn1xzGO184HW9t5gtgNjtvrS7bi6biGrAR0K0FAv7EOOET3/9+3zze/9fPnjJM7ngrFM49e1HstfIJSKIYHxD5WyS3eqdDQEK6wihQURwbQiAtYVI7T0Pw+XcJeqd3kVZv67ykSL2rKFUTJD+Vfsa58rkxaki4jEa8T7iXEXQ1PXggdXVKeV4TGEdv16Gu77ybT54yZV8/mvfYsdqZMv+z8Z7WGkEVy0gJtKOcW0tMF3vnPRiYRHBaNvbw5okzqxmrh/nRhXjCpBIqBuIOWyiirUW1UBRgtY7cdqgzSqTZicvf8Fzed9Ff8TpJ7yarVXi4ZGAdRGtp0hR0V4Hc5nodb/Dbk78uuGppw5EH8/hG09iaG+Ha2LIu1maD6zGJG024Ls/XeaSy67huls/yY9+cR9mcSujpX1YnnrEuLRMjU3tcBgKl2I7MQbMaJyGuwP9Yta2sBrIgqvtv2PncqbHmK2mdtdti2GzIKWq5jhfGy6POAP4KfXqDkYSOPSQA3n3eadz2glvYP/FtOv5SaAqklyTaqrtwjdZeSQdQfA+jQRoCYv5UgYA240k6F3U7eNtalVyltFYQkxZzCiwrYEzfv/P+Mr3fo0s7EWjeZayRqyE1O6ldlYoLBETWxexNfGTK+hipFj9NZf/xb/lmFcejCURnUpJ45WRa3s3pkAEnwL+0cCy90QpcdYyBa655Rv87aXX8qVvfBcvFaPNexOiYbXxmLLEGMe0qYnRY21qgROdbSXdAC4g9LNOXfjE0BaNGny3NlpLuy2S779f8D5ZWEWRSDaEpH40XWVhocL4KVrvRJplXvHCg3jn2adwzilHsqlM79lMajaNkhqfjz5tcGK7jW8tDTxRU9n2NDYsAQJElEmeWh+jxzdTFquKEAPTaUMxXkAxTIEf/nKVK66/jcuu/xg/+Nmv0WLM4tI+qBmzY3lCYSxFUVDXHjWCKUqakIRbk6uWCbArbFWMSgp460ytpi1yTaUruXSmjcbnRwSSarUqI3Fp6IxzKIJvFApHaR1NvUIl4CcPUIYVDnvRczn3pGM56+RjOGCvRCGpE6ymsKnvNDGd71r6ep84R4AtWiHQGfm1Lh75/XIWKsezWgJUgW01nPn7f8aXvv/ICTDV6EWMKnb111zxF3/G0YcdjNWQMqE4QohJfJOIbya4wkC0BDWsykyd5ZpbvshHLr+RL/3DD5hqhYw247UkSPtZgpKqDAojFKVjkuN7qMkkaFIGnxzuUM3Hvw6hqGKpaQX1Eyeu6QzxIRcyhzS5w5ikBhPAjQxVXGX5/l9Tiufwlx3KReecwtknH8GWvOxcDFhpawqFWpUoqc9XgYqBADcsASrQ0AZjPQUghKRRqGSLRYimYjUnUv/pPuWjV13H1dffwt0//jWMD8CUS0Dq/1VLHmKt2DJV2icS6zuSbRQZ0CITIEBPVTh3lsTWgkoNm8BswRoFaUKauSqSZsi6Er9aQwy4xTHaTBkZiPUO4soDlNS89IWHcMZJx3H+WUfzrCUQbdDGU5YjiNmiJM4smDmY3Fjf5+V+WVJX7ZiJMdViqgqSCbAl9ftrOOP9j4YAY6Zlkhjp5B4uzwToNOBQmqBUroAAMXpM4aiDR3Fg4b4GLr36Lq667ia+8q3vIeUmqJaYaIG6ET4APmJGI4wofmUn1aikKgzbt2/DjpLormjOqLYK3GITqakSDbNQxxwCVn23ApNlaGaZYUCMohpSGNA3UE8ZjxcwFpbv+zVLxYQjXv0SLjz3TN7+tsPYlN+piB4JNYXNv0okiRJImkkS8vuX9AzUXhfWRsFAgPmCTVG9SAzT1PspQqxrFIMtx3iS0GlLmL+4r+HG2z7D/+9D1/Gze3ei4hgtbqIOkUbBupLaJwmu1jvaNeYlJDn/hL4bvC4BzmWj042jFapUfJ0SKrgyWVy+BhGMJPFKCTVjJ/jJTnZsu59Xv/QQzjvtzZx9yjEcuM+mZKRFqCysrq4wGpe9Y5u5dqIzS7WfpZwrM+puMwGiQIHvJaDvr5ML/OgIsOlE4e3KvVz2F/8tb33lQbiY2u1VFR8jhSnASusAc99OuPX2r/KfLrmKb3/vJ6gtKBaWmEbDRAXjqm7QFpBcTkkzeWNoiCG3S+qs8DeVX7WJuJbIZsmD2VmanZ9+dcDsKUlwNSVYI0XhCH7KqFAqAg/c92u2jCve8KoX857zTuHYo16WxDJorXqlFCX6KdYZ+gRbe48aizUlMQYKY59SMbvHGhuaAGfoR1/oEgzkuGCMkSZEXFESItRNoKosq5rcuI9ccwcfvvxa/vGHvyIUSxQLW4nGUceAaHZ5JebQWnuB2FkiYE2CpLsoWhe4Q69sJZcNiTHodAI2SXOp92BSE7yf1hibsuS+aRiNSprJFGMMRWmZbN+G9csccuA+nHfmiZx/1gk8e6uhIPcZ+ClWUouUUYMaASJiAl1tofZ6pbsrqWfFZgKMpKlpvvtuJhPgf8OXvn/PIyRAj5EpQqCIDre6jcv+4n/k6Fc+E/HJPccpAUPIm9dP7wvc+InPcdV1d/D1b/2Qxo6wxZhiVFH7BnGWECNeNLcTWrqKXxHwIR2K2HR8Zs2mpLP4XRvf3XWtpWBCIseKVo7K4Mkj3rI1lsc+aKTUgF+5j5GucvQRh/G+d57JUUe8hAXJpEfKx6gq1gi+9pSly4eVenRTzNJ2Q8pTcfPTowb0kWLDE6D2pI00dwC0rmZsJ2k5CzHVb7XF0d57AoIvHQH47TJc//HP8zcfvYGvf+dHSLlIsbiJJmiv37klv5nlF0xklz7jlijnLKl+jC0H0QU6hRJiKjUR0pQ5jZRlSchWoAbApBouY1Kx7mhxjMQI9Q5Wt/2a5+y/xB9enOY4PPeARUaQyks0Wcitu7tbAmwh7THn8yuaq9YEP2sl4P7GcsbvPQoCFA9aYyVQ4HAr93LZ//k/cvQrD0wZLEmHOAXu3QlX33g7H7z0Wu7+p3vAbcVWmwkmlQOl0x5wztE0NSKKGpm1ibUlViHk+GiWjzI5S5+t967EMn35ZCH3M1uk4murEcUSJUXhOgKMEcEjGiAGCjE0kx0slZZjjngl7zn/dN7y2oMYA4SIU4/EkEbCmiIV/ufChJScoWt19j7i2v5wiVmv70lSCLKHsOEJ8CErObqRmi3hzKbaRZRIQK3Bk8pMViJcc/Pn+ZtLr+GL//B9ZLQVKTcj5QKNGnwzi5HFWCfyczZdKO2qbeNBoenxYkuCPQsQ0uulR5QpINUdfivjlRIu7XfS3vt4ilFFGSJ++QF05QGevc8S5552PBedfyrPeWbJGAgxMDKpAMbXNUVZgk+9A53qjIbZ1dZaQRpT/Co7wj6fNUG4vyk44/f/73zpe/c+iiywQJgysoqb/IpL/sO/5W2vfmE3x/dXO+DSa27j8qs/wfd+/mtWpSSUJSpFGlkQe+evl53vu6Mq9BIZuWRLer+Dksi4TfrMrR+gcEmhI6RBSandLil8O1ck8VOS1W4Bq55YrxBW7mPvTSVvP/oILjznFN5w2HMy8YGJASekmdeQiXbe3V6zFNZ8sQEwEOCsrGN3dXj9gtA1BKiSLu7JdJWqGtOo4kkX8bYp3Hj7l/n7K27mK9/+ATu9Q8pNSFFhbEW9skIxrtIITyGphUQLRZm15xRXVUmmq6v+b9ESYRuH6z3Wrvp2+9/t0JMcZLcpU0mjVCKMBJrlbZiwzNZNhovOP52zTz+WFz07JXqsjziUytqcKIpJZNOY5I43HmldK5FkZdjUkqgCodfFcH9jHyUBJvKxhQG/jCz/kuv/9s85/NBncM9va6678TY+dNl1/HLbKiurBTLaTCwrGknz3GZiqLNz1xFgT99RRXqJrLZmNRNfp4EWMwm2f5PI0VpYraGoMGVJrBuMRpy11NMVoKEoXVI2MgKhZnnbbzlov704+k2Hc8FZJ3H4yw5kyYIGxapn5Cwa03l3WSU6ySPJLNy3O7SPP9TzNggGAlzz7658o73tEaCiM8sLckdDiosZDMEnF7meRtRaTCmsArfe+S3+8kOX8vlvfI9VHWFHm7GjzakpvpkgolhXEU1BDL2VmeaFAr6rxJvBzALsu8RxZjFNo2sfm/+ugsWI0DSpnGNUOJSIBk9hlZ33/pqDD9qft7/1CM4/43gOf+EWKqCZNFQOrLNMm1VGRYkh1atZW2aTKY3bEZPTIZI6MsjX6bYGznw0LnDvHBXOItNt/Hd/8gfoZAd/9cG/4+e/uQ+zaYlgSsQtEimp/SyUIUYJnSp0isTBLP42+6HXlKbI7H4T24a0RIBtAXZnMapNSak8KKtwjmY6Bd+wsFjRTB6gcpqKl3fey7P23sRpbz+aC848mcMOPQBLisdasketIa2XtvWsf/m2Ki7rrokB62HDEyDMZy/XboqzgunWTcqWk6SLe+JrSjdKxaa1pyxy4DmmkZleDCHPhPjkl37IX11yFXd+8R/YvqoUo02Mx2PqukHNCLUFU6+pZMS2qhwNsxLkuOY4DA/WU7tr/3H+Kl021+DrmrKqEFF8TBepqhJjCmZVztJMdlD4VfbdVHDCUa/hPe84g1e9ZN8065WUzoneM3YulaPk1jSMSXHVHgFGZp7+Y0GAbRJaV2vKSrCTe6FusG6MGS2wUxuitTgZA0IMYG2B0UjTNMkCxmTxzvnfvdW52zV7Pytelzx3A/FpTUgqBUo9via9WCxiLBoC+IbCQSFCqHdidEKYbOeAfRY5+dg3c8EZJ/LqFz+LtkfDAME3GHFYmzbhGCPWtJn1/JvuxpxbnxDXehMbFxucAOOa27bQdzZEU9fcI91t+/8O75O0eSpbSUnDNmkgJtWiBTFgYBX4/Nd+xN9fehW33/kFdqxCubgXgYI6GigWaFTwTUixo17WcNeFmwup+/d2s2zjnFvf1enJrJwFNbmPOs02iTGmum0RNM7kmAprGBlYXb4PUy/zrP238ra3HMF5Z57MKw7dyqKZXaxWlVKSGo/m6V99F7glQJHHogwGJKYJL+V4iXqyE9vsZLEqmTSOIIZYpjIkg0V8zFZ6+k6qSgxFtlK1Z71lhWaBOXWVLkufN8EswDATMEilS/06PmsLwtSnpJRRChMwNGgzoVnZxjP3WeLMU4/lvNNP4WWHLDECXARi6kQOIVBW41SC1TSURZlIMQacSW1wM/Kerd92Vc/orb+5rxdO2ZjY4AQIs/hadil6i6ePrrA3ZzVnL0+6QKmFTrIWW3ootcFGxECdFYGjSaMVVyN8/Vs/4f/40Ef53Fe+zT0PrDLatA+r3hBMRbmwKQ1mIl1c8z0YMwJba53MWX0SZ8mI3VgxLRuJpMJqzbVzgsUYQ4yp7q0qHFZAwoTQTPD1CksjwynHHMHF553Oa16+P440KyJNCEtBes1jHWMeKRl6YcnHog7QoWmzKMegASc1WnuiLGCrCh8n6Xm2wiKpTdGHlMkGpOuFSB+RKhf77Ytr4qvtWuiRYCK+nPWWtIF0BIjFiFJopJSaZuV+pjvv48UvPIjj3vpGLj7ndJ5/4JgCiA1UBAonKbtvZzVSud8Fk3+yGCLOtcQ1T2jz9ZizWPH6NuLGJT8YCBBgPWqZJUce9JX9Uo/0zNCN+0v1g61bJAiTJvVgFq4gRCUaYRn44j/ew6VX3cDNt3+W326bYMabwY4JYmfagj0yo7sYUslEe7Rd4e1c/Vm/nAa6sWLS+4btnFYxnXWjmhIYRVnSTCaYwubymSllVVBaR5zuRJfvZZP1nH3qcZx96nEc/rLnsKlIbrHEOomJSDpeyZPAWmyr4bRHVQcYWZC0CUymDYxHUO+EKFizSIgBSknZ9GhIwxMkEX7ukFANzGhk7Xmb/cZr0d9outEH7RmV1PViFDR6XAxosxOmD/D8Z+/NO848kbNOPY5DnjGiAsI0UlYGS+q/hjTqIa0fSS6vTUPQow+4oh1kpLk2s/sl8/3k9ZIJcG4Rm+4pc6/ZoNjwBPiQVQIP8YSgPsnxx7YkoZf9NLPF1viGwhUIUNc1ZVlSN0ooLKukkNm3f7zM31xyDTd87A7uW65x1SZwVdr9JQ2xVly+UFOSwcYpSMgJmZwY6QL268R9cmG2tERoXXZVW4tmZv2F3MqVLMH0fIsQ8tD30kCpDabZgV99gM1l4H/41x/gwjOOhsZTuqTRYqS1SiyBmVVz7xTOfP+jI0Dj06YSbFJLMUYhRIQRIcbkT4Yk6yVt53JrrWk9/3v2BA12bxnl7o28IXWF7KpdQsaiWASHEutVwmQnLzhwX847/QTOOe04nvuMFLGzMVD40BUsp80zjTCI7QjUdnPtr0PVJLPmyvmZOt3j7e/dc4LbrHSvRGbuNRsUG7sKkoexAB7iCVbyhWhaSysTieldQKoUtgTNZWFFRYyKNWBig8URjPC6gxd5xb95J79//hlcft3NXH7drdzzwH3YajPlaBPTAMGCMSWhyZZANKlX1VnUSkdUyd9MbndXWJ0LpclWEICEtiZwZkVCxOvsu8Te30GTayak0jZclZRFijHbtv+Sb373RwSOxolgsMTc9yY5ptaepcdm1zVEN1OFSbFMzRd7032YyUSuWbmmjZO2Mcm2agWVbFHNLO10I5BVfZKlH9JGIZrkyMSB1hgRSqMYX1OvLhP8Ci9/wSGcdeq5nHPyW3nW3mloHwFGJm8Opes+x+YiZYEuy7u2tq+9T6Sk2+vWok92/SqA9chyg2PDE+BjgzUXzEOgy8J2cycabBAaNSxYy2uev8gL//k5vP+ic/iPf/sRbvr4Z/jRL39GsWlvrBnRxIARh68brDG40YhpXUPTIOM040FDSK1xod9vmqzGFM20bS4zPxZnhNB/+q4H3/2pkpSSrSlALVoupGl2+X1TImHXc9MdzmMRftrlGNvv2CNwMb1zkO5PySyhHU6lGrqERnsWuvnOIdIGdjUGMJK74tJwekKNNcqCVVZ3/JbJzvs44tUv54KzT+XU497AXouwYNJ+RPCMizQaMoTILrLlj/h7r/f4Wvd3wFoMBPgEQGRWPN2/L3WrmmwZCRWRkOdE7OUcS/vA//QnF/D+C8/kihs/zoevvoEf/OJnjJb2Rd2IqRF8TCMBKEuQcarFCwGJqWtDY6CTChUhSW3mEpq5WGFrBvXjhbsyVL9TQrPrN1eO8UQrieiaGGc6CuY6Negni1rLKNfnzOn4eYQwy/RHwRpHEyKYIikvx4gqBCuIiZQopQaale1Mm50c/qKDePd5f8BJbzuCfZfAaRbaiFCaiNoUIgkI6sqhHnkPYyDAPYA2aZIq7YQQQ6roM+BsrhGhoRDLymrD8w4Y8Se/fwrnnHUyl197K5decxPf/8k/IaNNjJf2YRJC5oBAaABJQ9Fj9J18eZojq8yGIFkSUaSjaGNss+wwM2NqXZj8HAP4XDsYZy44dN9xl+/PY33Rrxfr7H1Ix9a9I5izTHdNdBglTYSLKV4amzrV47XD0aJitSas7mBlsp0jXvUK3vPOcznuzS/hGQupeFl8zJVMTYoR5h7iNdG5AXsQAwE+QWitwI788q0QsBo73VSI3fUYQ8O4KokCjcIhewv/4r1v5/zTj+fam27jihtu5Rvf/wWuWEJchS1GxKpkWntCaHDjUSoRAWbEZjqrKBFgViRuebA9hjUwa/xWpc12pgz32kLc9Jg8CIE+Rsg6ezMFnfaBmOJzsE42NPbIPX3XNhyguORA5x5rawzWBvATnIHKOeqVFerpKuNCecNhh/DeC8/huKNfxrh9nwAaa8rCgM89wrneMn1KEqVQFcTsvpB9wOOPgQD3AGYkmFSLU70XxODREFKblnMY44gxdwKYNPIzNHDQ3oYPXHQC559+Ajfe8SX+9qPX8OVv/iM6WkLNmGq8CVOMWH1gG4wX13x47yfXdqrPWsKLPdfSrNtRInmgkfZc4I708neMMa4pyUn381i7yeu6wbl4PR977MKf/Wx4elY75L59jkjaHBSIfoozMLaK6JTpvfdT4Dnuza/nwrNO5fi3vJRFl3guhsioMhib/gaThC7abh1NQ8kxBmvsY38eBvzOGAjwCcTa+JjkSeVtyYyxJdhUMtJN7zWkaW0xQlQWC0OIgRACz14qed+pr+W0Y17Ltbd+kr+97Dq+/aNfMF2dEMIWFvfam+VJDvbnLHD64JAyoW3zvsa55n8gF/a27XftnYkkEnlnF1h9FhZ4vM7aQyB/7ozOALTXBZPLd1rC67U2anbjVSV3gShY7QqhRSPj8RiaFertD7DkPMe/+TAuPOckjn7jq9lc5na4GKicASfEqIgRbDmiUXqx3zSrxbbSZaQawSRHNTjDewoDAT7O6BdJA3Pub3v1pnCV67xFn+NXrdOU5OjaQdUNEj3joiDEKUW07Dd2vO+sozntpKO57TNf568/ehWf/8r3WFneSbW4F5Ei1w9mQtOYLSFlpuycrD5R6QnJCKKtnuGsfg5sJkiDiEn9CTpvgRmTxDzDHqoybb9Di9SvOyOjzjLUNJ6zbTeU7BZbCZgYWPntrxjbwElHvY4/uOhc3nT4M1kEnHpiPc1qLKmMSDFYY6hDTOfF9De9djJzOjgRzaVSA/ntSQwE+Dhjvaxof2pYzCZeP1Lnei+xMG9dmQJMKpZNiiZgNVCrYZ+RcN6xr+Ttb30lH/vkN7j0mlu483Nfx1NSLuxFNAVTVYJIKhIWSV0aTYOrRsnC85r1/SIheIxppaDarEJMhKFJQEDjfBa400qMbW/t420aZjIXAc11gAAiSTAgtuda0QhFUSCi1JNpLpFMg6tKa4lNjY0eE6ZQ72RpbDjpba/i4vNO4+jXPZ8CUJ8+TwgURUUrkdXn+crOSK1/fxds6G1+hj1nPA8YCHCPY702POk/MDdIZ3ahteF0owElJAEClGmj7F1YzjvuME46+jA+9Znv8KFLr0pEKCWjxS14ClYjiEstVaYo0ZjKZ4hCjNlltgWxc8ZnNYSdSOyDSG09YWj7ddcmQcQQ2kFSUZLrGQJ13STryzqMScXHjV9FPJRaE1d2sHlsOfGkN3Phuafzupc/kzQYFEJTs1DY1IMS2IX8ZkTWHk+yBNui5Me2CHzAY4GBAPcwZlVofeLrJSbWMKT0LriUrbQ0PuKMUhgLVvF1g7OWrcZw+ltfzMlv/dd88ot38xcf/DB3fuk7qF1kqdzENApeXHZTNdW6FSVYR5yuJuUCV2TC823hX7rVkC2/PUiCbb9rv69Z2k6OXOpjDMRIGoNhwWbLUBs0ekyYsMk0NMv3s3nsePvJR3DxO07n9S99TndxmJDidc44LGmah5q4PpG1M2Da2F8bblBD6o1h14LzAXsMAwHuQbSXamc5dD2mD/aqnJlldhFZO5vs5RCcSwokkZrV1ZpivMCxr3sRb3jNv+PzX/snLrn8Bj7x6S8wXfG4xS1gSlwxpvGRWE/AFmBs7iRp1j0Kk493V5WZJxA6Ow5o3cv5FkSaGso09NzXdR6YnrK7pTRIs42FInLiqUfzz957ES95ziaENC4yNqsUxuKsQ/J0uCakwUJG5pMXc1Z7e8fa37EjwtkrB/d3z2IgwD0MWdd9g7m6vQ5x9hpSSe3UTyldSVtqEbu+X4gqLI5LfAz46CkxvPXwgzjy8P+ML//DGVxy1fVce/tneWCyHRsXcW5ErQ5QIoYwibleuqXaZP0ZQDQJiT45sDsaidhRQVhdxoeaauRYcIbl7fcSmwkHPHMvznr7ybzz7FN4/oFLBJ96dS3gRLFFhaonhDpl6cXh2jGZCN1QuPb89Htwu3+vOUSJs3/2nztgj2AgwD2KuCZ7Ot83m4qN+00Ns77dFlVupwqxARzG2G58iSGJfmrwjIuKCKxkafYjXvEsDnvFH3LhBWfy4Suv59aP38l92x9godhC1AWmQVCTM5etiyySLa42rmVyOc3jdHoeEonoW57prOK2pS8Gop+yacFhg2W6fC+Nn/DcZ+zFiSecyMXnncaLDhx1Mb5oknqWaIQQ8OopXJHEbkmiECEErJhkdffJr9146N+y5tzE+VtZ89wBTzgGAtyjmLcA1lpUbbvU2rYpWXOhCeBMri3Liic+JJkoUSgKS9NMUTGMiyLVpwEjhdcduh+v+1e/xzfPPZUPXXIVt3zyC/x2+05ssYAtF/AIQQyiFm3FA9r4WtcKtydhejtEluMnJrUWDSyWsLrtl1AnSapzTzuNc884kUMOWMCSkhka0uCmykRMK3hrwEqR2hTFJX0/FOvyjhQ1kaFzayz3B4tgmMHge5JhIMAnCfrk91B9outljmfqK+kZrh1VmZ9hrU2dJQoSAs6lWr4YPZiSw1+wHy/7t3/EH7z33Xzkyhu5/Lpb+M323zAabSLYgkBBwKEk1zfMqqMTSUuqKtTu/pmVqkmiIR1Nttj6sUNpS1nUdoecBFTXUImaXIs42xYiSVlFs9SX1TRf14Yao1Oa+3fwoufsz7mnnsfpJx3LC56VLD5VMDGmkhUxWJP0+JSIWJvHA9ikdNNuNNkcb6X+bbGrpbe27GXtBqb5t5qFMwbsSQwEuMcxs+L698g6tw8FkTyke71PyfqERqByLTkaMCUR8CE18L/4OWP+3X95Du9+x8lcdvVNXH7dTfzTr3+DG29htLiVByYNbjwmTsktfJNU82ct0aT/amDJ2TSU3ZXkmey0utBCGhLV9RNr6kxp8EBJq8vXqiqLpkFRURXjXGqx8zVutEAIiayEgCMyoiGs3E+9ci+HHXoIF57zDs446VgO2Erb6AERCgvYlExqjVhri+78GTtvZc9OMojd/fb0YL/jml9kt+8x4InDQIBPQshubh8eHvrC6r+fqmJFsNmjnfpAVMOhzxrzJ//sbC56xxl88CNXcf1td/KT39yHjQWuKonOEiMUVgmaLLswrYmSkgh13bBQuuyQz6nzpba7PGs35pIQ0UAnoCqSNSHS/ODQNGlAlAqxDrhRisv5ehURi0OxBOLKA3i/zCtf8lxOO+F8zj3tePZdMowd1BOPs4aqMAS0G4H5eBLTI/v9BjyRGAhwg6MVBvVNg7GWkTNM6hpCwcgYDtzX8q8/cC4XX3gul1/7MS679lZ+8NMfszBaQooRGhyrdYMrx+BaMoJxWTCdTrFVlT+pDfwraZhmAQIeSZPqSOMl2/m3KUEU0kCoymKcJagg0eC9h6ZmoTBI2E6c7ET9lNcf9lLeceaJnHTM69hvc1rcjhTjG1mlKAANabaultkPHuhpI2MgwA2Ouq4pqwJj6MQSRmVykZvgAYdYOHAf+BfvO56Lzjuej15+M9fdeCvf/8kvWfaOLVv2Aa2Z+CmFpvrpxiulK+aGIM36ICxt5kLFEMXixGZhggAYjOQWOxFi0xARiB4NkYXCYQtPXH0AmdzHm17zcs445UROOuHNPGNxpngoXlFtEI0UhYMY0DZxQSRqwGgxmGgbGAMBbnCU2UIzNvXL+qYmxohzDieGwkQmXllylgbYr4L/4r0n8t5zT+TWOz7LX330Or5190+o68DYRUoTKAQKmzUCuwytaSlvjTapIBRg8kB2Db2gmcEYhzVK0zQUVlgcCTvu/SmbrOf1r3g+73nHH3H0m1/H3osptth2/or3FCYm2am2UCaE2awWDanneehL29AYCHCDQ0nyWsYIgsEVs+xu8kUjI5cyzFYsRQFN4zlgk+Ock97Isce+kWtuvItbPnY7n7njNnbc/6s03jHLd6VsAyRWa9vDeuU/bTpYTSK/nJFQSVlav7KdhapkZCPL236DjOCEI17O+We+nTOOezUVbcyxxrpkORYiiEvpB9/UszGSRlJrHIYY8zjMwfrb0BgIcAMjJ0SRXPAcFIgxlcyghBi6SWXee6xNJS+VBaJnwTjcCN5z9ps4/6w3cc0Vr6PUZUIIlNZS+yZ3v0KuVuxXdad7TQR8lupyIC7rLUzBr7BUROoHfsnYCUe96RVcfN7ZHPPml7CUGlaIdc2ocrjS5vxycrPrpklWbFFC+lrEKLg8hEiswceAG5KxGxoDAQ7oOMkIaU4wqSzG2oKgEStJRgoSETorhKBYk6yo2EQ2FQUXn3tCTjwEpqsTqtEIn9+7VcZqNfEMdBPYNCc8ksVpcARcWEGm2yjNhOPedBjveee5vOWIFzMCfAM0SmUFW5V4XwPgnKOuPWVZUhZF+84JBqxxhEiq8XNmkKMfMBDgRkZbo9Yv2O0es60O4LzWnXUlELEuvaKkwRXSiZ8YSVPVqqrKQq5Z2yYmEdi+wdW+xkkiTWcDjpqVbfcwcp6Tj34tF59zCq9/1YvZVIIPijOCs6AqtOE8Z8vOqizKUa/9Nt3Z62TOCtuD2TcgYSDADQ6hm1jR5WYfCtpr6bLtv6R9v5yKkOyQtr26rWhpjwINsOCEkhrnd1KvrlBUhnOPP4I/fs95vPLQ/RibpMwymU4oncNlFxnVWZJlzfdpxZ/bZMv6dDcvoj9gY2IgwI2ONVL2nY32EMkB7Z5rEZ0L6s1erzPXV/AYMaQ5wkm81YkgK9uQ7ffwjGfuxxFveT3vufAcXvOyZ7EAmKgYHwhAWYwgW5MaA66b1ORmB9QegswOYR7rDYCCgQQ3LgYC3NBYo07Sweyu4XjXu9r+3bnkRk8nz6T3NyIIMWmpRlBRrCr7LsB7zz2Rd7/7Yl50yFY6lT3f4IwkYS5bdMniEEnCrzSo951O3xx0za2s/Z67ajgP2JgQnY2tGrDh0COGbhWstYbWsY56lTId2r8zd85igpA6P1I6RDHEPIVt6mvu/e0ODjjggNSG16T2YitpipsQcS7FIH0WWuiGTCZN+sSwsuY4svDouuM+Ze6JzNSjB2xEDAS44fEg4uxzK2NXUYD1AoZ9ufdUzxxJQz4bEnUVqdxGQGODUU3CzcHiqmL29hGsiUzrVZxzaRhUVsJGQZsGUxT0w4DzKY/ed1jX0GtTPwP5bWQMBLiB0dYB9jliTqp9vZWx1trq050kqusTkqFBVEEbMCWqRVKtNpkg1VOvrlIsLBGA1aCUViiy6Fb/3bwHY107AK6dd9Qdw7oy8+uoLj+Edz9gA2HY/gbMYc5+eih2kN5/GTMt69zLq+09DkjlMqq9breglONFYkzxvcpKzvAqwSctmTitQU2azdH77N3t3HP393T6dO1jAzY8Bgtwg+OhjLxH/Z7a2pkt+up42tXqrSW0XdzZ3czPGKy5AY8GAwEOGDBgw2JwgQcMGLBhMRDggAEDNiwGAhwwYMCGxUCAAwYM2LAYCHDAgAEbFgMBDhgwYMNiIMABAwZsWAwEOGDAgA2LgQAHDBiwYTEQ4IABAzYsBgIcMGDAhsVAgAMGDNiwGAhwwIABGxYDAQ4YMGDD4v8PkXaHSTu1lz8AAAAASUVORK5CYII="""

def show_splash():
    splash = tk.Tk()
    splash.overrideredirect(True)
    splash.geometry("380x400+500+200")
    image_data = base64.b64decode(SPLASH_BASE64)
    img = Image.open(BytesIO(image_data))
    photo = ImageTk.PhotoImage(img)
    label = tk.Label(splash, image=photo)
    label.image = photo
    label.pack()
    splash.after(3000, splash.destroy)
    splash.mainloop()

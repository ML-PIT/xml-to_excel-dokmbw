import os
import re
import ssl
import smtplib
import requests
import tkinter as tk
import xml.etree.ElementTree as ET
import pandas as pd
from datetime import datetime
from tkinter import messagebox
from email.message import EmailMessage
from requests.adapters import HTTPAdapter
from requests_ntlm import HttpNtlmAuth
from urllib3.poolmanager import PoolManager
from urllib3.exceptions import InsecureRequestWarning
import urllib3
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

urllib3.disable_warnings(InsecureRequestWarning)

# ---------- KONFIGURATION ----------
USERNAME = "itbg\\sp_farm"
PASSWORT = "!DokM1MLcgi"
EMAIL_PASSWORT = "quaSeu2i"
ABSENDER = "ml-projekte-it@mail.de"
EMPFAENGER = "c.mueller@mlgruppe.de, k.vosen@mlgruppe.de, r.panske@mlgruppe.de, a.borowczak@mlgruppe.de"
SMTP_HOST = "smtp.mail.de"
MAIL_BETREFF_VORLAGE = "{} Bewertung DokMBW"
MAIL_TEXT = """Hallo zusammen,

diese Mail wurde Automatisiert versendet. Bitte nicht auf die Absenderadresse antworten.


Gruß
Die Software von Andy
"""

# ---------- HELPER ----------
def get_value_from_rating(rating_str, label):
    match = re.search(fr"{re.escape(label)};#(\d+)#", rating_str)
    return int(match.group(1)) if match else None

def extract_block(description_raw, label):
    match = re.search(fr"<b>{label}:</b>(.*?)</div>", description_raw.replace("\n", ""))
    return re.sub(r"<.*?>", "", match.group(1).strip()) if match else ""

# ---------- CONVERTER ----------
def convert_xml_to_excel(xml_file_path, output_excel_path):
    def get_value_from_rating(rating_str, label):
        match = re.search(fr"{re.escape(label)};#(\d+)#", rating_str)
        return int(match.group(1)) if match else None

    def extract_block(description_raw, label):
        match = re.search(fr"<b>{label}:</b>(.*?)</div>", description_raw.replace("\n", ""))
        return re.sub(r"<.*?>", "", match.group(1).strip()) if match else ""

    tree = ET.parse(xml_file_path)
    root = tree.getroot()
    channel = root.find("channel")
    entries = []

    for item in channel.findall("item"):
        author = "***"
        pub_date = item.find("pubDate").text
        pub_date = datetime.strptime(pub_date, "%a, %d %b %Y %H:%M:%S %Z").strftime("%d.%m.%Y")
        description_raw = item.find("description").text

        seminar = extract_block(description_raw, "Titel des Seminars")
        ort = extract_block(description_raw, "Ort der Schulung")
        beginn = extract_block(description_raw, "Schulungszeitraum (Beginn)")
        ende = extract_block(description_raw, "Schulungszeitraum (Ende)")
        trainer = extract_block(description_raw, "Name des Trainers")
        themen = extract_block(description_raw, "Folgende Themen haben mich besonders interessiert:")
        zu_kurz = extract_block(description_raw, "Diese Themen kamen meiner Meinung nach zu kurz/habe ich nicht verstanden:")
        bewertung = extract_block(description_raw, "Lehrgangsbewertung")
        zufriedenheit = extract_block(description_raw, "Zufriedenheit")
        vorschlaege = extract_block(description_raw, "Haben Sie noch Wünsche, Vorschläge, Anregungen?")

        entries.append({
            "Erstellt von": author,
            "Titel des Seminars": seminar,
            "Ort der Schulung": ort,
            "Schulungszeitraum (Beginn)": beginn,
            "Schulungszeitraum (Ende)": ende,
            "Name des Trainers": trainer,
            "Folgende Themen haben mich besonders interessiert:": themen,
            "Diese Themen kamen meiner Meinung nach zu kurz/habe ich nicht verstanden:": zu_kurz,
            "Haben Sie noch Wünsche, Vorschläge, Anregungen?": vorschlaege,
            "Zufriedenheit_Ich würde das Seminar weiterempfehlen": get_value_from_rating(zufriedenheit, "Ich würde das Seminar weiterempfehlen"),
            "Zufriedenheit_Zufriedenheit mit Seminar": get_value_from_rating(zufriedenheit, "Zufriedenheit mit Seminar"),
            "Zufriedenheit_Zufriedenheit mit Trainer/in": get_value_from_rating(zufriedenheit, "Zufriedenheit mit Trainer/in"),
            "Lehrgangsbewertung_Teilnehmerunterlagen": get_value_from_rating(bewertung, "Teilnehmerunterlagen"),
            "Lehrgangsbewertung_Praxisanteil des Seminars (erster Eindruck)": get_value_from_rating(bewertung, "Praxisanteil des Seminars (erster Eindruck)"),
            "Lehrgangsbewertung_Präsentation der Inhalte (Nachvollziehbarkeit)": get_value_from_rating(bewertung, "Präsentation der Inhalte (Nachvollziehbarkeit)"),
            "Lehrgangsbewertung_Durchführung durch Trainer/in (Methodisch)": get_value_from_rating(bewertung, "Durchführung durch Trainer/in (Methodisch)"),
            "Lehrgangsbewertung_Durchführung durch Trainer/in (Fachlich)": get_value_from_rating(bewertung, "Durchführung durch Trainer/in (Fachlich)"),
            "Lehrgangsbewertung_Struktur des Seminars (Roter Faden)": get_value_from_rating(bewertung, "Struktur des Seminars (Roter Faden)"),
            "Lehrgangsbewertung_Allgemeine Atmosphäre": get_value_from_rating(bewertung, "Allgemeine Atmosphäre"),
            "Elementtyp": "Element",
            "Pfad": ""
        })

    df = pd.DataFrame(entries)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bewertungen"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    stripe_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        for cell in ws[r_idx]:
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif r_idx % 2 == 0:
                cell.fill = stripe_fill

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 15)

    ws.auto_filter.ref = ws.dimensions
    wb.save(output_excel_path)
# ---------- RSS FEED ----------
class LegacySSLAdapter(HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
        ctx.options |= 0x00000004
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        kwargs["ssl_context"] = ctx
        self.poolmanager = PoolManager(*args, **kwargs)

def download_rss_and_save_xml(server_num):
    server_prefix = f"{int(server_num):02}"
    base_url = f"https://{server_prefix}.ml-schulung.de"
    overview_url = f"{base_url}/trainerseite/Lists/Lehrgangsbewertung/overview.aspx"

    session = requests.Session()
    session.mount("https://", LegacySSLAdapter())
    session.auth = HttpNtlmAuth(USERNAME, PASSWORT)
    session.verify = False
    session.headers.update({"User-Agent": "DokMBW/NTLM"})

    try:
        res = session.get(overview_url, timeout=10)
        res.raise_for_status()
        with open(f"{server_prefix}_overview_debug.html", "w", encoding="utf-8") as f:
            f.write(res.text)

        guids = re.findall(r'\{[0-9a-fA-F\-]{36}\}', res.text)
        guids = list(set(guids))
        if not guids:
            raise Exception("Keine GUIDs im HTML gefunden.")

        for guid in guids:
            feed_url = f"{base_url}/trainerseite/_layouts/15/listfeed.aspx?List={guid}"
            try:
                feed_res = session.get(feed_url, timeout=10)
                feed_res.raise_for_status()

                xml_filename = f"{server_prefix}_feed.xml"
                with open(xml_filename, "w", encoding="utf-8") as f:
                    f.write(feed_res.text)

                root = ET.fromstring(feed_res.text)
                trainer = extract_block(root.find("channel").find("item").find("description").text, "Name des Trainers")

                return {
                    "server": server_prefix,
                    "trainer": trainer,
                    "xml_file": xml_filename
                }

            except Exception:
                continue

        raise Exception("Keine funktionierende Feed-URL mit gültiger GUID gefunden.")
    except Exception as e:
        raise Exception(f"[{server_prefix}] Fehler: {e}")

# ---------- MAIL ----------
def sende_auswertung_per_mail(excel_dateien, empfaenger):
    datum = datetime.now().strftime("%d.%m.%Y")

    msg = EmailMessage()
    msg["Subject"] = MAIL_BETREFF_VORLAGE.format(datum)
    msg["From"] = ABSENDER
    msg["To"] = empfaenger
    msg.set_content(MAIL_TEXT)

    for datei in excel_dateien:
        with open(datei, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="octet-stream",
                filename=os.path.basename(datei)
            )

    try:
        with smtplib.SMTP(SMTP_HOST, 587, timeout=20) as smtp:
            smtp.starttls()
            smtp.login(ABSENDER, EMAIL_PASSWORT)
            smtp.send_message(msg)
            print(f"✅ Mail erfolgreich gesendet an {empfaenger}")
    except Exception as e:
        print(f"❌ Fehler beim Senden an {empfaenger}: {e}")

# ---------- GUI ----------
def run_gui_process(server_list):
    erfolge = []
    fehler = []

    for server in server_list:
        try:
            result = download_rss_and_save_xml(server)
            datum = datetime.now().strftime("%d%m%Y")
            excel_file = f"{datum}_{result['server']}_{result['trainer']}.xlsx"
            convert_xml_to_excel(result["xml_file"], excel_file)
            erfolge.append(excel_file)
        except Exception as e:
            fehler.append(str(e))

    if erfolge:
        sende_auswertung_per_mail(erfolge, EMPFAENGER)
    else:
        messagebox.showerror("Fehler", "\n".join(fehler))

def start_gui():
    root = tk.Tk()
    root.title("DokMBW Serverauswahl")
    vars = []

    tk.Label(root, text="Wähle die Server aus:").pack(pady=(10, 0))

    for i in range(1, 11):
        var = tk.IntVar()
        tk.Checkbutton(root, text=f"{i:02}", variable=var).pack(anchor="w")
        vars.append((i, var))

    def starten():
        selected = [int(i) for i, var in vars if var.get()]
        if not selected:
            messagebox.showwarning("Hinweis", "Bitte mindestens einen Server auswählen.")
            return
        run_gui_process(selected)

    tk.Button(root, text="Auswertung starten", command=starten).pack(pady=10)
    root.mainloop()

if __name__ == "__main__":
    start_gui()

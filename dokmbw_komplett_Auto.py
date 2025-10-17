import os
import re
import ssl
import smtplib
import requests
import tkinter as tk
import xml.etree.ElementTree as ET
import pandas as pd
from datetime import datetime, time
from tkinter import messagebox
from email.message import EmailMessage
from requests.adapters import HTTPAdapter
from requests_ntlm import HttpNtlmAuth
from urllib3.poolmanager import PoolManager
from urllib3.exceptions import InsecureRequestWarning
import urllib3
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from tkinter import filedialog
import threading
import time as time_module
import schedule
import logging
import subprocess
import sys

urllib3.disable_warnings(InsecureRequestWarning)

# ---------- KONFIGURATION ----------
USERNAME = "itbg\\sp_farm"
PASSWORT = "!DokM1MLcgi"
EMAIL_PASSWORT = "quaSeu2i"
ABSENDER = "ml-projekte-it@mail.de"
EMPFAENGER = {"a.borowczak@mlgruppe.de",
              "k.vosen@mlgruppe.de",
              "c.mueller@mlgruppe.de",
              "r.panske@mlgruppe.de",
              "j.kujasch@mlgruppe.de"}
SMTP_HOST = "smtp.mail.de"
MAIL_BETREFF_VORLAGE = "{} Bewertung DokMBW"
MAIL_TEXT = """Hallo zusammen,

diese Mail wurde Automatisiert versendet. Bitte nicht auf die Absenderadresse antworten.


Gruß
Die Software von Andy
"""


# ---------- HELPER ----------
def get_value_from_rating(rating_str, label):
    """
    Extracts a numerical value corresponding to a specified label from the given
    rating string. The rating string is assumed to contain label-value pairs in
    a specific format, and this function retrieves the integer value associated
    with the provided label.

    :param rating_str: The input string containing label-value pairs in the format
        `label;#value#`. This string represents ratings or similar data, which
        contains labels and their corresponding numerical values.
    :type rating_str: str
    :param label: The label whose corresponding numerical value needs to be
        extracted from the rating string.
    :type label: str
    :return: The integer value associated with the specified label if found in
        the rating string, otherwise None.
    :rtype: Optional[int]
    """
    match = re.search(fr"{re.escape(label)};#(\d+)#", rating_str)
    return int(match.group(1)) if match else None


def extract_block(description_raw, label):
    """
    Extracts and cleans a specific labeled block of text from a raw HTML-like description.

    This function searches for a block of text in the provided raw description
    that is associated with a specific label. It uses regex to locate the block in
    the format `<b>{label}:</b> ... </div>`, removes all HTML tags within the block,
    and returns the cleaned text. If the label is not found or the format is
    incorrect, the function returns an empty string.

    :param description_raw: A raw string containing HTML-like content to be
        searched through for a labeled block.
    :type description_raw: str
    :param label: The label to search for within the raw description that identifies
        the block to extract.
    :type label: str
    :return: The cleaned text content of the block associated with the specified
        label. Returns an empty string if no matching block is found.
    :rtype: str
    """
    match = re.search(fr"<b>{re.escape(label)}:</b>\s*(.*?)</div>", description_raw.replace("\n", ""))
    return re.sub(r"<.*?>", "", match.group(1).strip()) if match else ""


def berechne_speicherpfad(dateiname: str):
    """
    Constructs and returns a file system path for a specific file by creating a directory structure
    based on the current week of the ISO calendar year. The function ensures all necessary directories
    are created before returning the final file path.

    :param dateiname: The name of the file for which the storage path is constructed.
    :type dateiname: str

    :return: The complete file storage path including the file name.
    :rtype: Path
    """
    dokumente = Path.home() / "Documents"
    hauptordner = dokumente / "Bewertungen"
    kw_ordner = hauptordner / f"KW{datetime.now().isocalendar().week}"
    kw_ordner.mkdir(parents=True, exist_ok=True)
    return kw_ordner / dateiname


# ---------- CONVERTER ----------
def convert_xml_to_excel(xml_file_path, output_excel_path):
    """
    Converts data from an XML file to an Excel spreadsheet.

    This function parses an XML file containing structured data about seminar ratings
    and participant feedback. It extracts the data, processes it, and then saves it
    as an Excel file. The generated spreadsheet contains an organized layout with formatting,
    headers, alternating row colors, and auto-filter functionality for ease of analysis.

    :param xml_file_path: Path to the input XML file.
    :type xml_file_path: str
    :param output_excel_path: Path where the generated Excel file will be saved.
    :type output_excel_path: str
    :return: None
    """

    def get_value_from_rating(rating_str, label):
        """
        Converts an XML file to an Excel file.

        This function reads data from an XML file and writes it to an Excel file. It
        includes a nested helper function for extracting numerical values from a
        formatted rating string based on a specific label.

        :param xml_file_path: Path to the input XML file.
        :type xml_file_path: str
        :param output_excel_path: Path to the output Excel file.
        :type output_excel_path: str
        """
        match = re.search(fr"{re.escape(label)};#(\d+)#", rating_str)
        return int(match.group(1)) if match else None

    def extract_block(description_raw, label):
        """
        Convert an XML document into an Excel file by extracting specific data blocks.

        This function processes an XML input file, extracts structured data using
        certain labels, and writes the extracted information into an Excel file.
        The parsing and extraction are implemented based on predefined patterns.

        :param xml_file_path: Path to the source XML file to be processed.
        :type xml_file_path: str
        :param output_excel_path: Path where the resulting Excel file will be saved.
        :type output_excel_path: str
        """
        match = re.search(fr"<b>{re.escape(label)}:</b>\s*(.*?)</div>", description_raw.replace("\n", ""))
        return re.sub(r"<.*?>", "", match.group(1).strip()) if match else ""

    tree = ET.parse(xml_file_path)
    root = tree.getroot()
    channel = root.find("channel")
    entries = []

    for item in channel.findall("item"):
        author = "***"
        pub_date = item.find("pubDate").text
        pub_date = datetime.strptime(pub_date, "%a, %d %b %Y %H:%M:%S %Z").strftime("%d.%m.%Y")
        description_raw = item.find("description").text

        seminar = extract_block(description_raw, "Titel des Seminars")
        ort = extract_block(description_raw, "Ort der Schulung")
        beginn = extract_block(description_raw, "Schulungszeitraum (Beginn)")
        ende = extract_block(description_raw, "Schulungszeitraum (Ende)")
        trainer = extract_block(description_raw, "Name des Trainers")
        themen = extract_block(description_raw, "Folgende Themen haben mich besonders interessiert:")
        zu_kurz = extract_block(description_raw,
                                "Diese Themen kamen meiner Meinung nach zu kurz/habe ich nicht verstanden:")
        bewertung = extract_block(description_raw, "Lehrgangsbewertung")
        zufriedenheit = extract_block(description_raw, "Zufriedenheit")
        vorschlaege = extract_block(description_raw, "Haben Sie noch Wünsche, Vorschläge, Anregungen?")

        entries.append({
            "Erstellt von": author,
            "Titel des Seminars": seminar,
            "Ort der Schulung": ort,
            "Schulungszeitraum (Beginn)": beginn,
            "Schulungszeitraum (Ende)": ende,
            "Name des Trainers": trainer,
            "Folgende Themen haben mich besonders interessiert:": themen,
            "Diese Themen kamen meiner Meinung nach zu kurz/habe ich nicht verstanden:": zu_kurz,
            "Haben Sie noch Wünsche, Vorschläge, Anregungen?": vorschlaege,
            "Zufriedenheit_Ich würde das Seminar weiterempfehlen": get_value_from_rating(zufriedenheit,
                                                                                         "Ich würde das Seminar weiterempfehlen"),
            "Zufriedenheit_Zufriedenheit mit Seminar": get_value_from_rating(zufriedenheit,
                                                                             "Zufriedenheit mit Seminar"),
            "Zufriedenheit_Zufriedenheit mit Trainer/in": get_value_from_rating(zufriedenheit,
                                                                                "Zufriedenheit mit Trainer/in"),
            "Lehrgangsbewertung_Teilnehmerunterlagen": get_value_from_rating(bewertung, "Teilnehmerunterlagen"),
            "Lehrgangsbewertung_Praxisanteil des Seminars (erster Eindruck)": get_value_from_rating(bewertung,
                                                                                                    "Praxisanteil des Seminars (erster Eindruck)"),
            "Lehrgangsbewertung_Präsentation der Inhalte (Nachvollziehbarkeit)": get_value_from_rating(bewertung,
                                                                                                       "Präsentation der Inhalte (Nachvollziehbarkeit)"),
            "Lehrgangsbewertung_Durchführung durch Trainer/in (Methodisch)": get_value_from_rating(bewertung,
                                                                                                   "Durchführung durch Trainer/in (Methodisch)"),
            "Lehrgangsbewertung_Durchführung durch Trainer/in (Fachlich)": get_value_from_rating(bewertung,
                                                                                                 "Durchführung durch Trainer/in (Fachlich)"),
            "Lehrgangsbewertung_Struktur des Seminars (Roter Faden)": get_value_from_rating(bewertung,
                                                                                            "Struktur des Seminars (Roter Faden)"),
            "Lehrgangsbewertung_Allgemeine Atmosphäre": get_value_from_rating(bewertung, "Allgemeine Atmosphäre"),
            "Elementtyp": "Element",
            "Pfad": ""
        })

    df = pd.DataFrame(entries)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bewertungen"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    stripe_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        for cell in ws[r_idx]:
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif r_idx % 2 == 0:
                cell.fill = stripe_fill

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 15)

    ws.auto_filter.ref = ws.dimensions
    wb.save(output_excel_path)


# ---------- RSS FEED ----------
class LegacySSLAdapter(HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
        ctx.options |= 0x00000004
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        kwargs["ssl_context"] = ctx
        self.poolmanager = PoolManager(*args, **kwargs)


def download_rss_and_save_xml(server_num):
    server_prefix = f"{int(server_num):02}"
    base_url = f"https://{server_prefix}.ml-schulung.de"
    overview_url = f"{base_url}/trainerseite/Lists/Lehrgangsbewertung/overview.aspx"

    session = requests.Session()
    session.mount("https://", LegacySSLAdapter())
    session.auth = HttpNtlmAuth(USERNAME, PASSWORT)
    session.verify = False
    session.headers.update({"User-Agent": "DokMBW/NTLM"})

    try:
        res = session.get(overview_url, timeout=10)
        res.raise_for_status()
        with open(f"{server_prefix}_overview_debug.html", "w", encoding="utf-8") as f:
            f.write(res.text)

        guids = re.findall(r'\{[0-9a-fA-F\-]{36}\}', res.text)
        guids = list(set(guids))
        if not guids:
            raise Exception("Keine GUIDs im HTML gefunden.")

        for guid in guids:
            feed_url = f"{base_url}/trainerseite/_layouts/15/listfeed.aspx?List={guid}"
            try:
                feed_res = session.get(feed_url, timeout=10)
                feed_res.raise_for_status()

                xml_filename = f"{server_prefix}_feed.xml"
                with open(xml_filename, "w", encoding="utf-8") as f:
                    f.write(feed_res.text)

                root = ET.fromstring(feed_res.text)
                trainer = extract_block(root.find("channel").find("item").find("description").text, "Name des Trainers")

                return {
                    "server": server_prefix,
                    "trainer": trainer,
                    "xml_file": xml_filename
                }

            except Exception:
                continue

        raise Exception("Keine funktionierende Feed-URL mit gültiger GUID gefunden.")
    except Exception as e:
        raise Exception(f"[{server_prefix}] Fehler: {e}")


# ---------- MAIL ----------
def sende_auswertung_per_mail(excel_dateien, empfaenger):
    datum = datetime.now().strftime("%d.%m.%Y")

    msg = EmailMessage()
    msg["Subject"] = MAIL_BETREFF_VORLAGE.format(datum)
    msg["From"] = ABSENDER
    msg["To"] = empfaenger
    msg.set_content(MAIL_TEXT)

    for datei in excel_dateien:
        with open(datei, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="octet-stream",
                filename=os.path.basename(datei)
            )

    try:
        with smtplib.SMTP(SMTP_HOST, 587, timeout=20) as smtp:
            smtp.starttls()
            smtp.login(ABSENDER, EMAIL_PASSWORT)
            smtp.send_message(msg)
            print(f"✅ Mail erfolgreich gesendet an {empfaenger}")
    except Exception as e:
        print(f"❌ Fehler beim Senden an {empfaenger}: {e}")


# ---------- AUTOMATISCHE SERVER-ABFRAGE ----------

def run_automatic_process():
    """
    Läuft automatisch alle Server durch und verarbeitet gefundene Daten
    """
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Starte automatische Server-Abfrage...")
    
    erfolge = []
    alle_server = list(range(1, 11))  # Server 01-10
    
    def get_anwender_kuerzel(seminar_titel):
        """
        Konvertiert Seminartitel zu Anwender-Kürzeln
        """
        kuerzel_mapping = {
            "Anwender-Schulung": "AN",
            "Assistent": "AA", 
            "Löschberechtigter": "LOEBE",
            "Registrator": "REG",
            "Anwendungsmanager": "AM"
        }
        
        for titel, kuerzel in kuerzel_mapping.items():
            if titel.lower() in seminar_titel.lower():
                return kuerzel
        return "UNBEKANNT"
    
    for server in alle_server:
        try:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Prüfe Server {server:02d}...")
            result = download_rss_and_save_xml(server)
            
            # XML parsen um Seminar-Titel zu extrahieren
            root = ET.parse(result["xml_file"]).getroot()
            channel = root.find("channel")
            first_item = channel.find("item")
            
            if first_item is None:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] Server {server:02d}: Keine Daten gefunden")
                continue
                
            description_raw = first_item.find("description").text
            seminar_titel = extract_block(description_raw, "Titel des Seminars")
            
            datum = datetime.now().strftime("%Y%m%d")
            anwender_kuerzel = get_anwender_kuerzel(seminar_titel)
            dateiname = f"{datum}-{anwender_kuerzel}-{result['trainer']}.xlsx"
            zielpfad = berechne_speicherpfad(dateiname)
            
            convert_xml_to_excel(result["xml_file"], zielpfad)
            erfolge.append(str(zielpfad))
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Server {server:02d}: Erfolgreich verarbeitet - {dateiname}")
            
        except Exception as e:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Server {server:02d}: {str(e)}")
            continue
    
    if erfolge:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Versende {len(erfolge)} Datei(en) per Mail...")
        sende_auswertung_per_mail(erfolge, EMPFAENGER)
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Automatische Verarbeitung abgeschlossen!")
    else:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Keine Daten zur Verarbeitung gefunden.")

def run_manual_test():
    """
    Manuelle Test-Ausführung der automatischen Server-Abfrage
    """
    print("=== MANUELLER TEST GESTARTET ===")
    run_automatic_process()
    print("=== MANUELLER TEST BEENDET ===")

def setup_scheduler():
    """
    Richtet den Zeitplaner für 16:30 Uhr ein
    """
    schedule.clear()
    schedule.every().day.at("16:30").do(run_automatic_process)
    print("Zeitplaner eingerichtet: Täglich um 16:30 Uhr")
    
    while True:
        schedule.run_pending()
        time_module.sleep(60)  # Prüfe jede Minute

def start_scheduler():
    """
    Startet den Zeitplaner in einem separaten Thread
    """
    scheduler_thread = threading.Thread(target=setup_scheduler, daemon=True)
    scheduler_thread.start()

# ---------- GUI ----------


def run_gui_process(server_list, versand_aktiv):
    erfolge = []
    fehler = []

    if not versand_aktiv:
        zielordner = filedialog.askdirectory(title="Speicherort für Präsenzbewertung auswählen")
        if not zielordner:
            messagebox.showinfo("Abgebrochen", "Kein Speicherort ausgewählt. Vorgang abgebrochen.")
            return
        zielordner = Path(zielordner)
    else:
        zielordner = None

    def get_anwender_kuerzel(seminar_titel):
        """
        Konvertiert Seminartitel zu Anwender-Kürzeln
        """
        kuerzel_mapping = {
            "Anwender-Schulung": "AN",
            "Assistent": "AA", 
            "Löschberechtigter": "LOEBE",
            "Registrator": "REG",
            "Anwendungsmanager": "AM"
        }
        
        for titel, kuerzel in kuerzel_mapping.items():
            if titel.lower() in seminar_titel.lower():
                return kuerzel
        return "UNBEKANNT"

    for server in server_list:
        try:
            result = download_rss_and_save_xml(server)
            
            # XML parsen um Seminar-Titel zu extrahieren
            root = ET.parse(result["xml_file"]).getroot()
            channel = root.find("channel")
            first_item = channel.find("item")
            description_raw = first_item.find("description").text
            seminar_titel = extract_block(description_raw, "Titel des Seminars")
            
            datum = datetime.now().strftime("%Y%m%d")
            anwender_kuerzel = get_anwender_kuerzel(seminar_titel)
            dateiname = f"{datum}-{anwender_kuerzel}-{result['trainer']}.xlsx"
            if versand_aktiv:
                zielpfad = berechne_speicherpfad(dateiname)
            else:
                zielpfad = zielordner / dateiname

            convert_xml_to_excel(result["xml_file"], zielpfad)
            erfolge.append(str(zielpfad))
        except Exception as e:
            fehler.append(str(e))

    if erfolge:
        if versand_aktiv:
            sende_auswertung_per_mail(erfolge, EMPFAENGER)
        else:
            messagebox.showinfo("Hinweis", f"{len(erfolge)} Datei(en) gespeichert. Kein Versand (Präsenzschulung).")
    else:
        messagebox.showerror("Fehler", "\n".join(fehler))


def start_gui():
    root = tk.Tk()
    root.title("DokMBW Serverauswahl")
    root.geometry("400x500")
    vars = []
    versand_var = tk.IntVar(value=1)

    # Automatische Funktionen Sektion
    tk.Label(root, text="=== AUTOMATISCHE FUNKTIONEN ===", font=("Arial", 12, "bold")).pack(pady=(10, 5))
    
    def manual_test():
        """Startet manuellen Test"""
        import subprocess
        import sys
        # Führe Test in separatem Konsolenfenster aus
        subprocess.Popen([sys.executable, "-c", f"""
import sys
sys.path.insert(0, r'{os.path.dirname(__file__)}')
from dokmbw_komplett_test2105 import run_manual_test
run_manual_test()
input('Drücke Enter zum Beenden...')
"""], creationflags=subprocess.CREATE_NEW_CONSOLE if os.name == 'nt' else 0)
        messagebox.showinfo("Test gestartet", "Manueller Test läuft in separatem Fenster!")
    
    def start_auto_mode():
        """Startet automatischen Modus mit Zeitplaner"""
        start_scheduler()
        messagebox.showinfo("Automatik gestartet", "Automatische Verarbeitung ist aktiv!\nTäglich um 16:30 Uhr")
    
    tk.Button(root, text="🧪 Manueller Test (alle Server)", command=manual_test, 
              bg="orange", fg="white", font=("Arial", 10, "bold")).pack(pady=5, fill="x", padx=20)
    
    tk.Button(root, text="⏰ Automatik starten (16:30 Uhr täglich)", command=start_auto_mode,
              bg="green", fg="white", font=("Arial", 10, "bold")).pack(pady=5, fill="x", padx=20)

    # Trennlinie
    tk.Label(root, text="=== MANUELLE SERVERAUSWAHL ===", font=("Arial", 12, "bold")).pack(pady=(20, 5))

    tk.Label(root, text="Wähle die Server aus:").pack(pady=(10, 0))

    for i in range(1, 11):
        var = tk.IntVar()
        tk.Checkbutton(root, text=f"{i:02}", variable=var).pack(anchor="w")
        vars.append((i, var))

    tk.Radiobutton(root, text="Versand aktiv (Online-Schulung)", variable=versand_var, value=1).pack(anchor="w",
                                                                                                     pady=(10, 0))
    tk.Radiobutton(root, text="Nur speichern (Präsenzschulung)", variable=versand_var, value=0).pack(anchor="w")

    def starten():
        selected = [int(i) for i, var in vars if var.get()]
        if not selected:
            messagebox.showwarning("Hinweis", "Bitte mindestens einen Server auswählen.")
            return
        run_gui_process(selected, versand_var.get() == 1)

    tk.Button(root, text="Ausgewählte Server verarbeiten", command=starten).pack(pady=10)
    root.mainloop()


if __name__ == "__main__":
    start_gui()

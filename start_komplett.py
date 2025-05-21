
from tkinter import (Tk, Button, Text, Menu, Label, Entry, StringVar,
                     Toplevel, END, OptionMenu, Radiobutton, filedialog, WORD, IntVar)
from tkinter import messagebox, scrolledtext
import os
import random
import string
import shutil
import openpyxl
import csv
from datetime import datetime
from pathlib import Path
import xml.etree.ElementTree as ET
import pandas as pd
import requests
import smtplib
from email.message import EmailMessage
from requests.adapters import HTTPAdapter
from requests_ntlm import HttpNtlmAuth
import ssl
from urllib3.poolmanager import PoolManager
from urllib3.exceptions import InsecureRequestWarning
import urllib3

urllib3.disable_warnings(InsecureRequestWarning)

USERNAME = "itbg\\sp_farm"
PASSWORT = "!DokM1MLcgi"
EMAIL_PASSWORT = "quaSeu2i"
ABSENDER = "ml-projekte-it@mail.de"
EMPFAENGER = "c.mueller@mlgruppe.de, k.vosen@mlgruppe.de, r.panske@mlgruppe.de, a.borowczak@mlgruppe.de"
SMTP_HOST = "smtp.mail.de"
MAIL_BETREFF_VORLAGE = "{} Bewertung DokMBW"
MAIL_TEXT = "Hallo zusammen,\n\ndiese Mail wurde Automatisiert versendet. Bitte nicht auf die Absenderadresse antworten.\n\nGruß\nDie Software von Andy"

def create_menu(root):
    menubar = Menu(root)

    # Menü für PW-Funktionen
    pw_menu = Menu(menubar, tearoff=0)
    pw_menu.add_command(label="DokMBw Passwörter", command=select_file_and_show_output)
    menubar.add_cascade(label="PW-Funktionen", menu=pw_menu)

    # Menü für DokMBw Bewertungs-Tool
    dokmbw_menu = Menu(menubar, tearoff=0)
    from dokmbw_komplett_test2105 import start_gui
    dokmbw_menu.add_command(label="Bewertungstool starten", command=start_gui)
    menubar.add_cascade(label="DokMBw Bewertungen", menu=dokmbw_menu)

    root.config(menu=menubar)

# --- Passwort-Tool ---
def show_permission_error_dialog():
    root = Tk()
    root.withdraw()
    messagebox.showerror("Permission Error", "Bitte schließe die Excel-Datei und klicke dann auf 'OK', um es erneut zu versuchen.")
    root.destroy()

def create_backup(file_path):
    backup_path = file_path.replace('.xlsx', '_backup.xlsx')
    shutil.copyfile(file_path, backup_path)
    return backup_path

def generate_random_password(length=8, special_chars="!#?", special_count=2):
    special_indices = random.sample(range(length), special_count)
    characters = string.ascii_letters + string.digits
    password_list = [random.choice(characters) for _ in range(length)]
    for index in special_indices:
        password_list[index] = random.choice(special_chars)
    return ''.join(password_list)

def generate_fixed_password(length=10, special_chars="!#?", special_count=1):
    special_indices = random.sample(range(length), special_count)
    characters = string.ascii_letters + string.digits
    password_list = [random.choice(characters) for _ in range(length - special_count)]
    for index in special_indices:
        password_list.insert(index, random.choice(special_chars))
    return ''.join(password_list)

def generate_passwords_and_save_to_csv(file_path):
    try:
        output_messages = []
        create_backup(file_path)
        output_messages.append("Backup-Datei erstellt.\n")
        workbook = openpyxl.load_workbook(file_path)
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        directory = os.path.dirname(file_path)
        fixed_password = generate_fixed_password()
        output_messages.append(f"Festes Passwort für Dozent generiert: {fixed_password}\n")

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            nutzer_column = None
            passwort_column = None
            csv_file_path = os.path.join(directory, f"{base_name}_{sheet_name}_passwords.csv")
            dozent_cell = sheet['A2']
            current_value = dozent_cell.value
            dozent_cell.value = f"{current_value} {fixed_password}"

            with open(csv_file_path, 'w', newline='') as csv_file:
                writer = csv.writer(csv_file)
                for i, column in enumerate(sheet.iter_cols(min_row=1, max_row=1)):
                    if column[0].value == "Nutzer":
                        nutzer_column = i
                    elif column[0].value in ["Kennwort", "Passwort"]:
                        passwort_column = i
                if nutzer_column is None or passwort_column is None:
                    output_messages.append(f"Spalte 'Nutzer' oder 'Passwort' in '{sheet_name}' nicht gefunden.\n")
                    continue
                base_password = generate_random_password()
                tn_counter = 1
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    nutzer_value = row[nutzer_column].value
                    if nutzer_value and any(nutzer_value.startswith(prefix) for prefix in ["TN", "PU", "FA", "LB", "RG"]):
                        nutzer_value_simple = nutzer_value.split('@')[0]
                        password = base_password + str(tn_counter).zfill(2)
                        row[passwort_column].value = password
                        writer.writerow([nutzer_value_simple, password])
                        output_messages.append(f"Passwort für {nutzer_value_simple} in '{sheet_name}' gesetzt: {password}\n")
                        tn_counter += 1
                output_messages.append(f"CSV-Datei für '{sheet_name}' erstellt: {csv_file_path}\n")
        workbook.save(file_path)
        output_messages.append("Passwörter erfolgreich generiert und Excel-Datei gespeichert.\n")
        return '\n'.join(output_messages)
    except PermissionError:
        show_permission_error_dialog()
        return generate_passwords_and_save_to_csv(file_path)

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        return generate_passwords_and_save_to_csv(file_path)
    return ""

def select_file_and_show_output():
    output = select_file()
    if output:
        show_output_in_window(output)

def show_output_in_window(output_text):
    output_window = Toplevel()
    output_window.title("Ausgabe")
    text_area = scrolledtext.ScrolledText(output_window, wrap=WORD, width=80, height=20)
    text_area.pack(padx=10, pady=10)
    text_area.insert(END, output_text)
    Button(output_window, text="Schließen", command=output_window.destroy).pack(pady=10)

# --- GUI Setup + root init folgt ---

if __name__ == "__main__":
    root = Tk()
    root.title("DokMBw All-in-One Tool")
    root.geometry("400x200")
    Label(root, text="Willkommen im DokMBw-Tool", font=("Arial", 14)).pack(pady=20)
    create_menu(root)
    root.mainloop()
